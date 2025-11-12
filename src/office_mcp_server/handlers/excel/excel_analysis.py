"""Excel 数据分析模块."""

import statistics
from typing import Any, Optional

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from loguru import logger
from scipy import stats
from sklearn.linear_model import LinearRegression
from sklearn.preprocessing import PolynomialFeatures

from office_mcp_server.config import config
from office_mcp_server.utils.file_manager import FileManager


class ExcelAnalysisOperations:
    """Excel 数据分析操作类."""

    def __init__(self) -> None:
        """初始化数据分析操作类."""
        self.file_manager = FileManager()

    def descriptive_statistics(
        self,
        filename: str,
        sheet_name: str,
        data_range: str,
    ) -> dict[str, Any]:
        """描述性统计分析.

        Args:
            filename: 文件名
            sheet_name: 工作表名称
            data_range: 数据范围 (如 'A1:A100')
        """
        try:
            file_path = config.paths.output_dir / filename
            self.file_manager.validate_file_path(file_path, must_exist=True)

            wb = load_workbook(str(file_path))

            if sheet_name not in wb.sheetnames:
                raise ValueError(f"工作表 '{sheet_name}' 不存在")

            ws = wb[sheet_name]
            data_cells = ws[data_range]

            # 提取数值数据
            values = []
            for row in data_cells:
                if isinstance(row, tuple):
                    for cell in row:
                        if isinstance(cell.value, (int, float)) and cell.value is not None:
                            values.append(cell.value)
                else:
                    if isinstance(row.value, (int, float)) and row.value is not None:
                        values.append(row.value)

            if not values:
                raise ValueError("没有找到有效的数值数据")

            # 计算统计指标
            stats = {
                "count": len(values),
                "sum": sum(values),
                "mean": statistics.mean(values),
                "median": statistics.median(values),
                "mode": statistics.mode(values) if len(set(values)) < len(values) else None,
                "std_dev": statistics.stdev(values) if len(values) > 1 else 0,
                "variance": statistics.variance(values) if len(values) > 1 else 0,
                "min": min(values),
                "max": max(values),
                "range": max(values) - min(values),
            }

            # 计算四分位数
            sorted_values = sorted(values)
            n = len(sorted_values)
            stats["q1"] = sorted_values[n // 4] if n >= 4 else sorted_values[0]
            stats["q2"] = statistics.median(values)
            stats["q3"] = sorted_values[3 * n // 4] if n >= 4 else sorted_values[-1]

            logger.info(f"描述性统计完成: {file_path}")
            return {
                "success": True,
                "filename": str(file_path),
                "sheet_name": sheet_name,
                "data_range": data_range,
                "statistics": stats,
            }

        except Exception as e:
            logger.error(f"描述性统计失败: {e}")
            return {"success": False, "message": f"统计失败: {str(e)}"}

    def correlation_analysis(
        self,
        filename: str,
        sheet_name: str,
        data_range1: str,
        data_range2: str,
    ) -> dict[str, Any]:
        """相关性分析.

        Args:
            filename: 文件名
            sheet_name: 工作表名称
            data_range1: 第一个数据范围 (如 'A1:A100')
            data_range2: 第二个数据范围 (如 'B1:B100')
        """
        try:
            file_path = config.paths.output_dir / filename
            self.file_manager.validate_file_path(file_path, must_exist=True)

            wb = load_workbook(str(file_path))

            if sheet_name not in wb.sheetnames:
                raise ValueError(f"工作表 '{sheet_name}' 不存在")

            ws = wb[sheet_name]

            # 提取第一组数据
            data_cells1 = ws[data_range1]
            values1 = []
            for row in data_cells1:
                if isinstance(row, tuple):
                    for cell in row:
                        if isinstance(cell.value, (int, float)) and cell.value is not None:
                            values1.append(cell.value)
                else:
                    if isinstance(row.value, (int, float)) and row.value is not None:
                        values1.append(row.value)

            # 提取第二组数据
            data_cells2 = ws[data_range2]
            values2 = []
            for row in data_cells2:
                if isinstance(row, tuple):
                    for cell in row:
                        if isinstance(cell.value, (int, float)) and cell.value is not None:
                            values2.append(cell.value)
                else:
                    if isinstance(row.value, (int, float)) and row.value is not None:
                        values2.append(row.value)

            if len(values1) != len(values2):
                raise ValueError("两组数据长度不一致")

            if len(values1) < 2:
                raise ValueError("数据量太少,无法计算相关性")

            # 计算皮尔逊相关系数
            n = len(values1)
            mean1 = statistics.mean(values1)
            mean2 = statistics.mean(values2)

            numerator = sum((x - mean1) * (y - mean2) for x, y in zip(values1, values2))
            denominator1 = sum((x - mean1) ** 2 for x in values1)
            denominator2 = sum((y - mean2) ** 2 for y in values2)

            if denominator1 == 0 or denominator2 == 0:
                correlation = 0
            else:
                correlation = numerator / ((denominator1 * denominator2) ** 0.5)

            # 判断相关性强度
            abs_corr = abs(correlation)
            if abs_corr >= 0.8:
                strength = "强相关"
            elif abs_corr >= 0.5:
                strength = "中等相关"
            elif abs_corr >= 0.3:
                strength = "弱相关"
            else:
                strength = "几乎不相关"

            logger.info(f"相关性分析完成: {file_path}")
            return {
                "success": True,
                "filename": str(file_path),
                "sheet_name": sheet_name,
                "data_range1": data_range1,
                "data_range2": data_range2,
                "correlation": round(correlation, 4),
                "strength": strength,
                "sample_size": n,
            }

        except Exception as e:
            logger.error(f"相关性分析失败: {e}")
            return {"success": False, "message": f"分析失败: {str(e)}"}

    def goal_seek(
        self,
        filename: str,
        sheet_name: str,
        target_cell: str,
        target_value: float,
        variable_cell: str,
        max_iterations: int = 100,
        tolerance: float = 0.001,
    ) -> dict[str, Any]:
        """单变量求解(目标搜索).

        Args:
            filename: 文件名
            sheet_name: 工作表名称
            target_cell: 目标单元格 (如 'D10')
            target_value: 目标值
            variable_cell: 变量单元格 (如 'B5')
            max_iterations: 最大迭代次数
            tolerance: 容差
        """
        try:
            file_path = config.paths.output_dir / filename
            self.file_manager.validate_file_path(file_path, must_exist=True)

            wb = load_workbook(str(file_path))

            if sheet_name not in wb.sheetnames:
                raise ValueError(f"工作表 '{sheet_name}' 不存在")

            ws = wb[sheet_name]

            # 简化的单变量求解 (二分法)
            variable_value = ws[variable_cell].value or 0
            min_val = variable_value - 1000
            max_val = variable_value + 1000

            for iteration in range(max_iterations):
                # 尝试中点值
                test_value = (min_val + max_val) / 2
                ws[variable_cell] = test_value

                # 重新计算 (注意: openpyxl不能自动计算公式)
                # 这里只是演示逻辑,实际需要Microsoft Excel API
                result_value = ws[target_cell].value

                if result_value is None:
                    raise ValueError(f"目标单元格 {target_cell} 没有公式或值")

                diff = abs(float(result_value) - target_value)

                if diff < tolerance:
                    wb.save(str(file_path))
                    wb.close()

                    logger.info(f"目标搜索完成: {file_path}")
                    return {
                        "success": True,
                        "message": "目标搜索成功",
                        "filename": str(file_path),
                        "target_cell": target_cell,
                        "target_value": target_value,
                        "variable_cell": variable_cell,
                        "solution": test_value,
                        "iterations": iteration + 1,
                        "final_diff": diff,
                    }

                # 调整搜索范围
                if float(result_value) < target_value:
                    min_val = test_value
                else:
                    max_val = test_value

            wb.close()

            return {
                "success": False,
                "message": f"达到最大迭代次数 {max_iterations}, 未找到精确解",
                "note": "此功能需要Excel应用程序支持以执行公式计算"
            }

        except Exception as e:
            logger.error(f"目标搜索失败: {e}")
            return {"success": False, "message": f"搜索失败: {str(e)}"}

    def regression_analysis(
        self,
        filename: str,
        sheet_name: str,
        x_range: str,
        y_range: str,
        regression_type: str = "linear",
    ) -> dict[str, Any]:
        """回归分析.

        Args:
            filename: 文件名
            sheet_name: 工作表名称
            x_range: 自变量范围 (如 'A1:A100')
            y_range: 因变量范围 (如 'B1:B100')
            regression_type: 回归类型 ('linear'线性, 'polynomial'多项式)
        """
        try:
            file_path = config.paths.output_dir / filename
            self.file_manager.validate_file_path(file_path, must_exist=True)

            wb = load_workbook(str(file_path))

            if sheet_name not in wb.sheetnames:
                raise ValueError(f"工作表 '{sheet_name}' 不存在")

            ws = wb[sheet_name]

            # 提取X数据
            x_cells = ws[x_range]
            x_values = []
            for row in x_cells:
                if isinstance(row, tuple):
                    for cell in row:
                        if isinstance(cell.value, (int, float)) and cell.value is not None:
                            x_values.append(cell.value)
                else:
                    if isinstance(row.value, (int, float)) and row.value is not None:
                        x_values.append(row.value)

            # 提取Y数据
            y_cells = ws[y_range]
            y_values = []
            for row in y_cells:
                if isinstance(row, tuple):
                    for cell in row:
                        if isinstance(cell.value, (int, float)) and cell.value is not None:
                            y_values.append(cell.value)
                else:
                    if isinstance(row.value, (int, float)) and row.value is not None:
                        y_values.append(row.value)

            if len(x_values) != len(y_values):
                raise ValueError("X和Y数据长度不一致")

            if len(x_values) < 2:
                raise ValueError("数据量太少,无法进行回归分析")

            X = np.array(x_values).reshape(-1, 1)
            y = np.array(y_values)

            if regression_type == "linear":
                # 线性回归
                model = LinearRegression()
                model.fit(X, y)

                y_pred = model.predict(X)
                r_squared = model.score(X, y)

                # 计算标准误差
                residuals = y - y_pred
                mse = np.mean(residuals ** 2)
                rmse = np.sqrt(mse)

                result = {
                    "success": True,
                    "regression_type": "linear",
                    "coefficient": float(model.coef_[0]),
                    "intercept": float(model.intercept_),
                    "r_squared": float(r_squared),
                    "rmse": float(rmse),
                    "equation": f"y = {model.coef_[0]:.4f}x + {model.intercept_:.4f}",
                    "sample_size": len(x_values),
                }

            elif regression_type == "polynomial":
                # 多项式回归 (2次)
                poly = PolynomialFeatures(degree=2)
                X_poly = poly.fit_transform(X)

                model = LinearRegression()
                model.fit(X_poly, y)

                y_pred = model.predict(X_poly)
                r_squared = model.score(X_poly, y)

                residuals = y - y_pred
                mse = np.mean(residuals ** 2)
                rmse = np.sqrt(mse)

                result = {
                    "success": True,
                    "regression_type": "polynomial",
                    "coefficients": [float(c) for c in model.coef_],
                    "intercept": float(model.intercept_),
                    "r_squared": float(r_squared),
                    "rmse": float(rmse),
                    "degree": 2,
                    "sample_size": len(x_values),
                }
            else:
                raise ValueError(f"不支持的回归类型: {regression_type}")

            logger.info(f"回归分析完成: {file_path}")
            result["filename"] = str(file_path)
            result["sheet_name"] = sheet_name
            return result

        except Exception as e:
            logger.error(f"回归分析失败: {e}")
            return {"success": False, "message": f"分析失败: {str(e)}"}

    def anova_analysis(
        self,
        filename: str,
        sheet_name: str,
        *group_ranges: str,
    ) -> dict[str, Any]:
        """方差分析 (ANOVA).

        Args:
            filename: 文件名
            sheet_name: 工作表名称
            *group_ranges: 各组数据范围 (如 'A1:A10', 'B1:B10', 'C1:C10')
        """
        try:
            file_path = config.paths.output_dir / filename
            self.file_manager.validate_file_path(file_path, must_exist=True)

            wb = load_workbook(str(file_path))

            if sheet_name not in wb.sheetnames:
                raise ValueError(f"工作表 '{sheet_name}' 不存在")

            ws = wb[sheet_name]

            # 提取各组数据
            groups = []
            for group_range in group_ranges:
                cells = ws[group_range]
                values = []
                for row in cells:
                    if isinstance(row, tuple):
                        for cell in row:
                            if isinstance(cell.value, (int, float)) and cell.value is not None:
                                values.append(cell.value)
                    else:
                        if isinstance(row.value, (int, float)) and row.value is not None:
                            values.append(row.value)

                if values:
                    groups.append(values)

            if len(groups) < 2:
                raise ValueError("至少需要2组数据进行方差分析")

            # 执行单因素方差分析
            f_statistic, p_value = stats.f_oneway(*groups)

            # 计算组间和组内统计量
            all_values = [v for group in groups for v in group]
            grand_mean = np.mean(all_values)

            # 组均值
            group_means = [np.mean(group) for group in groups]
            group_sizes = [len(group) for group in groups]

            # 判断显著性
            alpha = 0.05
            significant = p_value < alpha

            logger.info(f"方差分析完成: {file_path}")
            return {
                "success": True,
                "filename": str(file_path),
                "sheet_name": sheet_name,
                "f_statistic": float(f_statistic),
                "p_value": float(p_value),
                "significant": significant,
                "significance_level": alpha,
                "groups_count": len(groups),
                "group_sizes": group_sizes,
                "group_means": [float(m) for m in group_means],
                "grand_mean": float(grand_mean),
                "interpretation": "组间存在显著差异" if significant else "组间不存在显著差异",
            }

        except Exception as e:
            logger.error(f"方差分析失败: {e}")
            return {"success": False, "message": f"分析失败: {str(e)}"}

    def t_test(
        self,
        filename: str,
        sheet_name: str,
        group1_range: str,
        group2_range: str,
        test_type: str = "independent",
    ) -> dict[str, Any]:
        """t检验.

        Args:
            filename: 文件名
            sheet_name: 工作表名称
            group1_range: 第一组数据范围
            group2_range: 第二组数据范围
            test_type: 检验类型 ('independent'独立样本, 'paired'配对样本)
        """
        try:
            file_path = config.paths.output_dir / filename
            self.file_manager.validate_file_path(file_path, must_exist=True)

            wb = load_workbook(str(file_path))

            if sheet_name not in wb.sheetnames:
                raise ValueError(f"工作表 '{sheet_name}' 不存在")

            ws = wb[sheet_name]

            # 提取第一组数据
            cells1 = ws[group1_range]
            values1 = []
            for row in cells1:
                if isinstance(row, tuple):
                    for cell in row:
                        if isinstance(cell.value, (int, float)) and cell.value is not None:
                            values1.append(cell.value)
                else:
                    if isinstance(row.value, (int, float)) and row.value is not None:
                        values1.append(row.value)

            # 提取第二组数据
            cells2 = ws[group2_range]
            values2 = []
            for row in cells2:
                if isinstance(row, tuple):
                    for cell in row:
                        if isinstance(cell.value, (int, float)) and cell.value is not None:
                            values2.append(cell.value)
                else:
                    if isinstance(row.value, (int, float)) and row.value is not None:
                        values2.append(row.value)

            if not values1 or not values2:
                raise ValueError("数据不足")

            if test_type == "independent":
                # 独立样本t检验
                t_statistic, p_value = stats.ttest_ind(values1, values2)
            elif test_type == "paired":
                # 配对样本t检验
                if len(values1) != len(values2):
                    raise ValueError("配对t检验要求两组数据长度相同")
                t_statistic, p_value = stats.ttest_rel(values1, values2)
            else:
                raise ValueError(f"不支持的检验类型: {test_type}")

            alpha = 0.05
            significant = p_value < alpha

            logger.info(f"t检验完成: {file_path}")
            return {
                "success": True,
                "filename": str(file_path),
                "sheet_name": sheet_name,
                "test_type": test_type,
                "t_statistic": float(t_statistic),
                "p_value": float(p_value),
                "significant": significant,
                "significance_level": alpha,
                "group1_mean": float(np.mean(values1)),
                "group2_mean": float(np.mean(values2)),
                "group1_size": len(values1),
                "group2_size": len(values2),
                "interpretation": "两组存在显著差异" if significant else "两组不存在显著差异",
            }

        except Exception as e:
            logger.error(f"t检验失败: {e}")
            return {"success": False, "message": f"分析失败: {str(e)}"}

    def chi_square_test(
        self,
        filename: str,
        sheet_name: str,
        observed_range: str,
    ) -> dict[str, Any]:
        """卡方检验.

        Args:
            filename: 文件名
            sheet_name: 工作表名称
            observed_range: 观测频数范围 (如 'A1:B2')
        """
        try:
            file_path = config.paths.output_dir / filename
            self.file_manager.validate_file_path(file_path, must_exist=True)

            wb = load_workbook(str(file_path))

            if sheet_name not in wb.sheetnames:
                raise ValueError(f"工作表 '{sheet_name}' 不存在")

            ws = wb[sheet_name]

            # 提取观测频数
            cells = ws[observed_range]
            observed = []
            for row in cells:
                row_data = []
                if isinstance(row, tuple):
                    for cell in row:
                        if isinstance(cell.value, (int, float)) and cell.value is not None:
                            row_data.append(cell.value)
                else:
                    if isinstance(row.value, (int, float)) and row.value is not None:
                        row_data.append(row.value)
                if row_data:
                    observed.append(row_data)

            if not observed:
                raise ValueError("没有找到有效的观测数据")

            observed = np.array(observed)

            # 执行卡方检验
            chi2, p_value, dof, expected = stats.chi2_contingency(observed)

            alpha = 0.05
            significant = p_value < alpha

            logger.info(f"卡方检验完成: {file_path}")
            return {
                "success": True,
                "filename": str(file_path),
                "sheet_name": sheet_name,
                "chi_square": float(chi2),
                "p_value": float(p_value),
                "degrees_of_freedom": int(dof),
                "significant": significant,
                "significance_level": alpha,
                "observed": observed.tolist(),
                "expected": expected.tolist(),
                "interpretation": "存在显著关联" if significant else "不存在显著关联",
            }

        except Exception as e:
            logger.error(f"卡方检验失败: {e}")
            return {"success": False, "message": f"分析失败: {str(e)}"}

    def trend_analysis(
        self,
        filename: str,
        sheet_name: str,
        data_range: str,
        periods_ahead: int = 5,
    ) -> dict[str, Any]:
        """趋势分析和预测.

        Args:
            filename: 文件名
            sheet_name: 工作表名称
            data_range: 数据范围
            periods_ahead: 预测未来期数
        """
        try:
            file_path = config.paths.output_dir / filename
            self.file_manager.validate_file_path(file_path, must_exist=True)

            wb = load_workbook(str(file_path))

            if sheet_name not in wb.sheetnames:
                raise ValueError(f"工作表 '{sheet_name}' 不存在")

            ws = wb[sheet_name]

            # 提取数据
            cells = ws[data_range]
            values = []
            for row in cells:
                if isinstance(row, tuple):
                    for cell in row:
                        if isinstance(cell.value, (int, float)) and cell.value is not None:
                            values.append(cell.value)
                else:
                    if isinstance(row.value, (int, float)) and row.value is not None:
                        values.append(row.value)

            if len(values) < 3:
                raise ValueError("数据量太少,无法进行趋势分析")

            # 使用线性回归进行趋势预测
            X = np.arange(len(values)).reshape(-1, 1)
            y = np.array(values)

            model = LinearRegression()
            model.fit(X, y)

            # 预测未来值
            future_X = np.arange(len(values), len(values) + periods_ahead).reshape(-1, 1)
            predictions = model.predict(future_X)

            # 计算趋势
            slope = model.coef_[0]
            if slope > 0:
                trend = "上升"
            elif slope < 0:
                trend = "下降"
            else:
                trend = "平稳"

            logger.info(f"趋势分析完成: {file_path}")
            return {
                "success": True,
                "filename": str(file_path),
                "sheet_name": sheet_name,
                "trend": trend,
                "slope": float(slope),
                "intercept": float(model.intercept_),
                "r_squared": float(model.score(X, y)),
                "predictions": [float(p) for p in predictions],
                "periods_ahead": periods_ahead,
            }

        except Exception as e:
            logger.error(f"趋势分析失败: {e}")
            return {"success": False, "message": f"分析失败: {str(e)}"}

    def moving_average(
        self,
        filename: str,
        sheet_name: str,
        data_range: str,
        window: int = 3,
        output_cell: Optional[str] = None,
    ) -> dict[str, Any]:
        """移动平均.

        Args:
            filename: 文件名
            sheet_name: 工作表名称
            data_range: 数据范围
            window: 移动窗口大小
            output_cell: 输出起始单元格 (可选)
        """
        try:
            file_path = config.paths.output_dir / filename
            self.file_manager.validate_file_path(file_path, must_exist=True)

            wb = load_workbook(str(file_path))

            if sheet_name not in wb.sheetnames:
                raise ValueError(f"工作表 '{sheet_name}' 不存在")

            ws = wb[sheet_name]

            # 提取数据
            cells = ws[data_range]
            values = []
            for row in cells:
                if isinstance(row, tuple):
                    for cell in row:
                        if isinstance(cell.value, (int, float)) and cell.value is not None:
                            values.append(cell.value)
                else:
                    if isinstance(row.value, (int, float)) and row.value is not None:
                        values.append(row.value)

            if len(values) < window:
                raise ValueError(f"数据量 ({len(values)}) 小于窗口大小 ({window})")

            # 计算移动平均
            series = pd.Series(values)
            ma = series.rolling(window=window).mean()
            ma_values = ma.tolist()

            # 如果指定输出单元格,写入结果
            if output_cell:
                for i, val in enumerate(ma_values):
                    if not pd.isna(val):
                        ws[f"{output_cell[0]}{int(output_cell[1:]) + i}"] = float(val)
                wb.save(str(file_path))

            wb.close()

            logger.info(f"移动平均计算完成: {file_path}")
            return {
                "success": True,
                "filename": str(file_path),
                "sheet_name": sheet_name,
                "window": window,
                "original_data": values,
                "moving_average": [float(v) if not pd.isna(v) else None for v in ma_values],
            }

        except Exception as e:
            logger.error(f"移动平均计算失败: {e}")
            return {"success": False, "message": f"计算失败: {str(e)}"}

    def exponential_smoothing(
        self,
        filename: str,
        sheet_name: str,
        data_range: str,
        alpha: float = 0.3,
        output_cell: Optional[str] = None,
    ) -> dict[str, Any]:
        """指数平滑.

        Args:
            filename: 文件名
            sheet_name: 工作表名称
            data_range: 数据范围
            alpha: 平滑系数 (0-1)
            output_cell: 输出起始单元格 (可选)
        """
        try:
            if not 0 < alpha < 1:
                raise ValueError("平滑系数alpha必须在0到1之间")

            file_path = config.paths.output_dir / filename
            self.file_manager.validate_file_path(file_path, must_exist=True)

            wb = load_workbook(str(file_path))

            if sheet_name not in wb.sheetnames:
                raise ValueError(f"工作表 '{sheet_name}' 不存在")

            ws = wb[sheet_name]

            # 提取数据
            cells = ws[data_range]
            values = []
            for row in cells:
                if isinstance(row, tuple):
                    for cell in row:
                        if isinstance(cell.value, (int, float)) and cell.value is not None:
                            values.append(cell.value)
                else:
                    if isinstance(row.value, (int, float)) and row.value is not None:
                        values.append(row.value)

            if len(values) < 2:
                raise ValueError("数据量太少")

            # 计算指数平滑
            smoothed = [values[0]]
            for i in range(1, len(values)):
                smoothed_value = alpha * values[i] + (1 - alpha) * smoothed[i - 1]
                smoothed.append(smoothed_value)

            # 如果指定输出单元格,写入结果
            if output_cell:
                for i, val in enumerate(smoothed):
                    ws[f"{output_cell[0]}{int(output_cell[1:]) + i}"] = float(val)
                wb.save(str(file_path))

            wb.close()

            logger.info(f"指数平滑计算完成: {file_path}")
            return {
                "success": True,
                "filename": str(file_path),
                "sheet_name": sheet_name,
                "alpha": alpha,
                "original_data": values,
                "smoothed_data": smoothed,
            }

        except Exception as e:
            logger.error(f"指数平滑计算失败: {e}")
            return {"success": False, "message": f"计算失败: {str(e)}"}
