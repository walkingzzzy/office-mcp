"""Excel 自动填充操作模块."""

from typing import Any, Optional, Union
from datetime import datetime, timedelta

from openpyxl import load_workbook
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
from loguru import logger

from office_mcp_server.config import config
from office_mcp_server.utils.file_manager import FileManager


class ExcelAutoFillOperations:
    """Excel 自动填充操作类."""

    def __init__(self) -> None:
        """初始化自动填充操作类."""
        self.file_manager = FileManager()

    def fill_series(
        self,
        filename: str,
        sheet_name: str,
        start_cell: str,
        end_cell: str,
        fill_type: str = "linear",
        start_value: Union[int, float] = 1,
        step: Union[int, float] = 1,
    ) -> dict[str, Any]:
        """序列填充.

        Args:
            filename: 文件名
            sheet_name: 工作表名称
            start_cell: 起始单元格 (如 'A1')
            end_cell: 结束单元格 (如 'A10')
            fill_type: 填充类型 ('linear'线性, 'growth'增长, 'date'日期, 'auto'自动)
            start_value: 起始值
            step: 步长
        """
        try:
            file_path = config.paths.output_dir / filename
            self.file_manager.validate_file_path(file_path, must_exist=True)

            wb = load_workbook(str(file_path))

            if sheet_name not in wb.sheetnames:
                raise ValueError(f"工作表 '{sheet_name}' 不存在")

            ws = wb[sheet_name]

            start_col_letter, start_row = coordinate_from_string(start_cell)
            start_col = column_index_from_string(start_col_letter)

            end_col_letter, end_row = coordinate_from_string(end_cell)
            end_col = column_index_from_string(end_col_letter)

            if start_col != end_col and start_row != end_row:
                raise ValueError("填充只能在同一行或同一列中进行")

            current_value = start_value
            count = 0

            if start_col == end_col:
                for row in range(start_row, end_row + 1):
                    if fill_type == "linear":
                        ws.cell(row=row, column=start_col, value=current_value)
                        current_value += step
                    elif fill_type == "growth":
                        ws.cell(row=row, column=start_col, value=current_value)
                        current_value *= step
                    elif fill_type == "date":
                        if isinstance(start_value, (int, float)):
                            date_val = datetime.now() + timedelta(days=int(current_value) - 1)
                        else:
                            date_val = datetime.now()
                        ws.cell(row=row, column=start_col, value=date_val)
                        current_value += step
                    count += 1
            else:
                for col in range(start_col, end_col + 1):
                    if fill_type == "linear":
                        ws.cell(row=start_row, column=col, value=current_value)
                        current_value += step
                    elif fill_type == "growth":
                        ws.cell(row=start_row, column=col, value=current_value)
                        current_value *= step
                    elif fill_type == "date":
                        if isinstance(start_value, (int, float)):
                            date_val = datetime.now() + timedelta(days=int(current_value) - 1)
                        else:
                            date_val = datetime.now()
                        ws.cell(row=start_row, column=col, value=date_val)
                        current_value += step
                    count += 1

            wb.save(str(file_path))
            wb.close()

            logger.info(f"序列填充成功: {file_path}, {start_cell} 到 {end_cell}")
            return {
                "success": True,
                "message": f"成功填充序列从 {start_cell} 到 {end_cell}",
                "filename": str(file_path),
                "start_cell": start_cell,
                "end_cell": end_cell,
                "fill_type": fill_type,
                "count": count,
            }

        except Exception as e:
            logger.error(f"序列填充失败: {e}")
            return {"success": False, "message": f"序列填充失败: {str(e)}"}

    def copy_fill(
        self,
        filename: str,
        sheet_name: str,
        source_cell: str,
        target_range: str,
    ) -> dict[str, Any]:
        """复制填充.

        Args:
            filename: 文件名
            sheet_name: 工作表名称
            source_cell: 源单元格 (如 'A1')
            target_range: 目标范围 (如 'A2:A10')
        """
        try:
            file_path = config.paths.output_dir / filename
            self.file_manager.validate_file_path(file_path, must_exist=True)

            wb = load_workbook(str(file_path))

            if sheet_name not in wb.sheetnames:
                raise ValueError(f"工作表 '{sheet_name}' 不存在")

            ws = wb[sheet_name]

            source_value = ws[source_cell].value

            target_cells = ws[target_range]
            count = 0

            for row in target_cells:
                if isinstance(row, tuple):
                    for cell in row:
                        cell.value = source_value
                        count += 1
                else:
                    row.value = source_value
                    count += 1

            wb.save(str(file_path))
            wb.close()

            logger.info(f"复制填充成功: {file_path}, 从 {source_cell} 到 {target_range}")
            return {
                "success": True,
                "message": f"成功复制 {source_cell} 的值到 {target_range}",
                "filename": str(file_path),
                "source_cell": source_cell,
                "target_range": target_range,
                "count": count,
            }

        except Exception as e:
            logger.error(f"复制填充失败: {e}")
            return {"success": False, "message": f"复制填充失败: {str(e)}"}

    def formula_fill(
        self,
        filename: str,
        sheet_name: str,
        start_cell: str,
        formula: str,
        fill_direction: str = "down",
        count: int = 10,
    ) -> dict[str, Any]:
        """公式填充.

        Args:
            filename: 文件名
            sheet_name: 工作表名称
            start_cell: 起始单元格 (如 'A1')
            formula: 公式 (如 '=SUM(A1:A10)')
            fill_direction: 填充方向 ('down'向下, 'right'向右)
            count: 填充数量
        """
        try:
            file_path = config.paths.output_dir / filename
            self.file_manager.validate_file_path(file_path, must_exist=True)

            wb = load_workbook(str(file_path))

            if sheet_name not in wb.sheetnames:
                raise ValueError(f"工作表 '{sheet_name}' 不存在")

            ws = wb[sheet_name]

            col_letter, row = coordinate_from_string(start_cell)
            start_col = column_index_from_string(col_letter)

            filled_count = 0
            if fill_direction == "down":
                for i in range(count):
                    ws.cell(row=row + i, column=start_col, value=formula)
                    filled_count += 1
            elif fill_direction == "right":
                for i in range(count):
                    ws.cell(row=row, column=start_col + i, value=formula)
                    filled_count += 1
            else:
                raise ValueError(f"不支持的填充方向: {fill_direction}")

            wb.save(str(file_path))
            wb.close()

            logger.info(f"公式填充成功: {file_path}, 从 {start_cell} {fill_direction}")
            return {
                "success": True,
                "message": f"成功从 {start_cell} 向 {fill_direction} 填充 {filled_count} 个公式",
                "filename": str(file_path),
                "start_cell": start_cell,
                "formula": formula,
                "direction": fill_direction,
                "count": filled_count,
            }

        except Exception as e:
            logger.error(f"公式填充失败: {e}")
            return {"success": False, "message": f"公式填充失败: {str(e)}"}
