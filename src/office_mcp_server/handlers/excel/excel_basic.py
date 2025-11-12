"""Excel 基础操作模块."""

from pathlib import Path
from typing import Any, Optional, Union

from openpyxl import Workbook, load_workbook
from loguru import logger

from office_mcp_server.config import config
from office_mcp_server.utils.file_manager import FileManager


class ExcelBasicOperations:
    """Excel 基础操作类."""

    def __init__(self) -> None:
        """初始化基础操作类."""
        self.file_manager = FileManager()

    def create_workbook(self, filename: str, sheet_name: Optional[str] = None) -> dict[str, Any]:
        """创建 Excel 工作簿."""
        try:
            file_path = self.file_manager.validate_file_path(filename)
            self.file_manager.validate_file_extension(filename, [".xlsx", ".xls"])

            output_path = config.paths.output_dir / file_path.name
            self.file_manager.ensure_directory(output_path.parent)

            wb = Workbook()
            ws = wb.active

            if sheet_name:
                ws.title = sheet_name
            else:
                ws.title = "Sheet1"

            wb.save(str(output_path))

            logger.info(f"Excel 工作簿创建成功: {output_path}")
            return {
                "success": True,
                "message": "Excel 工作簿创建成功",
                "filename": str(output_path),
                "sheet_name": ws.title,
            }

        except Exception as e:
            logger.error(f"创建 Excel 工作簿失败: {e}")
            return {"success": False, "message": f"创建失败: {str(e)}"}

    def write_cell(
        self, filename: str, sheet_name: str, cell: str, value: Union[str, int, float]
    ) -> dict[str, Any]:
        """写入单元格数据."""
        try:
            file_path = config.paths.output_dir / filename
            self.file_manager.validate_file_path(file_path, must_exist=True)

            wb = load_workbook(str(file_path))

            if sheet_name not in wb.sheetnames:
                raise ValueError(f"工作表 '{sheet_name}' 不存在")

            ws = wb[sheet_name]
            ws[cell] = value
            wb.save(str(file_path))

            logger.info(f"单元格 {cell} 写入成功: {file_path}")
            return {
                "success": True,
                "message": "单元格写入成功",
                "filename": str(file_path),
                "cell": cell,
                "value": value,
            }

        except Exception as e:
            logger.error(f"写入单元格失败: {e}")
            return {"success": False, "message": f"写入失败: {str(e)}"}

    def write_range(
        self, filename: str, sheet_name: str, start_cell: str, data: list[list[Any]]
    ) -> dict[str, Any]:
        """批量写入数据."""
        try:
            file_path = config.paths.output_dir / filename
            self.file_manager.validate_file_path(file_path, must_exist=True)

            wb = load_workbook(str(file_path))

            if sheet_name not in wb.sheetnames:
                raise ValueError(f"工作表 '{sheet_name}' 不存在")

            ws = wb[sheet_name]

            from openpyxl.utils.cell import coordinate_from_string, column_index_from_string

            col_letter, row = coordinate_from_string(start_cell)
            start_col = column_index_from_string(col_letter)

            for i, row_data in enumerate(data):
                for j, value in enumerate(row_data):
                    ws.cell(row=row + i, column=start_col + j, value=value)

            wb.save(str(file_path))

            logger.info(f"批量写入成功: {file_path}")
            return {
                "success": True,
                "message": "批量写入成功",
                "filename": str(file_path),
                "rows": len(data),
                "cols": len(data[0]) if data else 0,
            }

        except Exception as e:
            logger.error(f"批量写入失败: {e}")
            return {"success": False, "message": f"写入失败: {str(e)}"}

    def read_cell(self, filename: str, sheet_name: str, cell: str) -> dict[str, Any]:
        """读取单元格数据."""
        try:
            file_path = config.paths.output_dir / filename
            self.file_manager.validate_file_path(file_path, must_exist=True)

            wb = load_workbook(str(file_path))

            if sheet_name not in wb.sheetnames:
                raise ValueError(f"工作表 '{sheet_name}' 不存在")

            ws = wb[sheet_name]
            value = ws[cell].value

            logger.info(f"单元格 {cell} 读取成功: {file_path}")
            return {
                "success": True,
                "filename": str(file_path),
                "cell": cell,
                "value": value,
            }

        except Exception as e:
            logger.error(f"读取单元格失败: {e}")
            return {"success": False, "message": f"读取失败: {str(e)}"}

    def read_range(
        self, filename: str, sheet_name: str, cell_range: str
    ) -> dict[str, Any]:
        """读取单元格范围数据."""
        try:
            file_path = config.paths.output_dir / filename
            self.file_manager.validate_file_path(file_path, must_exist=True)

            wb = load_workbook(str(file_path))

            if sheet_name not in wb.sheetnames:
                raise ValueError(f"工作表 '{sheet_name}' 不存在")

            ws = wb[sheet_name]
            data_cells = ws[cell_range]

            data = []
            for row in data_cells:
                if isinstance(row, tuple):
                    row_data = [cell.value for cell in row]
                else:
                    row_data = [row.value]
                data.append(row_data)

            logger.info(f"读取范围 {cell_range} 成功: {file_path}")
            return {
                "success": True,
                "filename": str(file_path),
                "cell_range": cell_range,
                "data": data,
                "rows": len(data),
                "cols": len(data[0]) if data else 0,
            }

        except Exception as e:
            logger.error(f"读取范围失败: {e}")
            return {"success": False, "message": f"读取失败: {str(e)}"}

    def read_row(
        self, filename: str, sheet_name: str, row_index: int
    ) -> dict[str, Any]:
        """读取整行数据."""
        try:
            file_path = config.paths.output_dir / filename
            self.file_manager.validate_file_path(file_path, must_exist=True)

            wb = load_workbook(str(file_path))

            if sheet_name not in wb.sheetnames:
                raise ValueError(f"工作表 '{sheet_name}' 不存在")

            ws = wb[sheet_name]
            row_data = [cell.value for cell in ws[row_index]]

            logger.info(f"读取行 {row_index} 成功: {file_path}")
            return {
                "success": True,
                "filename": str(file_path),
                "row_index": row_index,
                "data": row_data,
                "count": len(row_data),
            }

        except Exception as e:
            logger.error(f"读取行失败: {e}")
            return {"success": False, "message": f"读取失败: {str(e)}"}

    def read_column(
        self, filename: str, sheet_name: str, col_index: int
    ) -> dict[str, Any]:
        """读取整列数据."""
        try:
            file_path = config.paths.output_dir / filename
            self.file_manager.validate_file_path(file_path, must_exist=True)

            wb = load_workbook(str(file_path))

            if sheet_name not in wb.sheetnames:
                raise ValueError(f"工作表 '{sheet_name}' 不存在")

            ws = wb[sheet_name]

            from openpyxl.utils import get_column_letter
            col_letter = get_column_letter(col_index)

            col_data = [cell.value for cell in ws[col_letter]]

            logger.info(f"读取列 {col_letter} 成功: {file_path}")
            return {
                "success": True,
                "filename": str(file_path),
                "col_index": col_index,
                "col_letter": col_letter,
                "data": col_data,
                "count": len(col_data),
            }

        except Exception as e:
            logger.error(f"读取列失败: {e}")
            return {"success": False, "message": f"读取失败: {str(e)}"}

    def read_all_data(
        self, filename: str, sheet_name: str, include_empty: bool = False
    ) -> dict[str, Any]:
        """读取整表数据."""
        try:
            file_path = config.paths.output_dir / filename
            self.file_manager.validate_file_path(file_path, must_exist=True)

            wb = load_workbook(str(file_path))

            if sheet_name not in wb.sheetnames:
                raise ValueError(f"工作表 '{sheet_name}' 不存在")

            ws = wb[sheet_name]

            data = []
            if include_empty:
                for row in ws.iter_rows(values_only=True):
                    data.append(list(row))
            else:
                for row in ws.iter_rows(min_row=1, max_row=ws.max_row,
                                       min_col=1, max_col=ws.max_column,
                                       values_only=True):
                    if any(cell is not None for cell in row):
                        data.append(list(row))

            logger.info(f"读取整表数据成功: {file_path}")
            return {
                "success": True,
                "filename": str(file_path),
                "sheet_name": sheet_name,
                "data": data,
                "rows": len(data),
                "cols": len(data[0]) if data else 0,
            }

        except Exception as e:
            logger.error(f"读取整表数据失败: {e}")
            return {"success": False, "message": f"读取失败: {str(e)}"}

    def clear_cell(
        self, filename: str, sheet_name: str, cell: str
    ) -> dict[str, Any]:
        """清除单元格内容."""
        try:
            file_path = config.paths.output_dir / filename
            self.file_manager.validate_file_path(file_path, must_exist=True)

            wb = load_workbook(str(file_path))

            if sheet_name not in wb.sheetnames:
                raise ValueError(f"工作表 '{sheet_name}' 不存在")

            ws = wb[sheet_name]
            ws[cell].value = None

            wb.save(str(file_path))
            wb.close()

            logger.info(f"清除单元格 {cell} 成功: {file_path}")
            return {
                "success": True,
                "message": f"成功清除单元格 {cell} 的内容",
                "filename": str(file_path),
                "cell": cell,
            }

        except Exception as e:
            logger.error(f"清除单元格失败: {e}")
            return {"success": False, "message": f"清除失败: {str(e)}"}

    def clear_range(
        self, filename: str, sheet_name: str, cell_range: str
    ) -> dict[str, Any]:
        """清除范围内所有单元格内容."""
        try:
            file_path = config.paths.output_dir / filename
            self.file_manager.validate_file_path(file_path, must_exist=True)

            wb = load_workbook(str(file_path))

            if sheet_name not in wb.sheetnames:
                raise ValueError(f"工作表 '{sheet_name}' 不存在")

            ws = wb[sheet_name]
            data_cells = ws[cell_range]

            count = 0
            for row in data_cells:
                if isinstance(row, tuple):
                    for cell in row:
                        cell.value = None
                        count += 1
                else:
                    row.value = None
                    count += 1

            wb.save(str(file_path))
            wb.close()

            logger.info(f"清除范围 {cell_range} 成功: {file_path}")
            return {
                "success": True,
                "message": f"成功清除范围 {cell_range} 的内容 ({count} 个单元格)",
                "filename": str(file_path),
                "cell_range": cell_range,
                "count": count,
            }

        except Exception as e:
            logger.error(f"清除范围失败: {e}")
            return {"success": False, "message": f"清除失败: {str(e)}"}

    def get_workbook_info(self, filename: str) -> dict[str, Any]:
        """获取工作簿信息."""
        try:
            file_path = config.paths.output_dir / filename
            self.file_manager.validate_file_path(file_path, must_exist=True)

            wb = load_workbook(str(file_path))

            sheets_info = []
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                sheets_info.append({
                    "name": sheet_name,
                    "rows": ws.max_row,
                    "cols": ws.max_column,
                })

            logger.info(f"获取工作簿信息成功: {file_path}")
            return {
                "success": True,
                "filename": str(file_path),
                "sheet_count": len(wb.sheetnames),
                "sheet_names": wb.sheetnames,
                "sheets": sheets_info,
            }

        except Exception as e:
            logger.error(f"获取工作簿信息失败: {e}")
            return {"success": False, "message": f"获取失败: {str(e)}"}
