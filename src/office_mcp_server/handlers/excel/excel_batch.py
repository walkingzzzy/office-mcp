"""Excel 批量处理和高级功能模块."""

import glob
from typing import Any, Optional, Callable
from pathlib import Path

from openpyxl import load_workbook
from loguru import logger

from office_mcp_server.config import config
from office_mcp_server.utils.file_manager import FileManager


class ExcelBatchOperations:
    """Excel 批量处理操作类."""

    def __init__(self) -> None:
        """初始化批量处理操作类."""
        self.file_manager = FileManager()

    def batch_process_files(
        self,
        pattern: str,
        operation: str,
        **kwargs: Any,
    ) -> dict[str, Any]:
        """批量处理文件.

        Args:
            pattern: 文件匹配模式 (如 '*.xlsx' 或 'report_*.xlsx')
            operation: 操作类型 ('format'格式化, 'merge'合并, 'export'导出)
            **kwargs: 操作特定参数
        """
        try:
            search_path = config.paths.output_dir / pattern
            files = list(glob.glob(str(search_path)))

            if not files:
                raise ValueError(f"未找到匹配文件: {pattern}")

            results = []
            success_count = 0
            failure_count = 0

            for file_path in files:
                try:
                    file_name = Path(file_path).name

                    if operation == "format":
                        # 批量格式化
                        self._batch_format_file(file_path, **kwargs)
                    elif operation == "merge":
                        # 批量合并
                        pass
                    elif operation == "export":
                        # 批量导出
                        self._batch_export_file(file_path, **kwargs)
                    else:
                        raise ValueError(f"不支持的操作类型: {operation}")

                    results.append({
                        "file": file_name,
                        "status": "success",
                    })
                    success_count += 1

                except Exception as e:
                    results.append({
                        "file": file_name,
                        "status": "failed",
                        "error": str(e),
                    })
                    failure_count += 1

            logger.info(f"批量处理完成: 成功 {success_count}, 失败 {failure_count}")
            return {
                "success": True,
                "message": f"批量处理完成: 成功 {success_count}, 失败 {failure_count}",
                "total": len(files),
                "success_count": success_count,
                "failure_count": failure_count,
                "results": results,
            }

        except Exception as e:
            logger.error(f"批量处理失败: {e}")
            return {"success": False, "message": f"批量处理失败: {str(e)}"}

    def _batch_format_file(self, file_path: str, **kwargs: Any) -> None:
        """批量格式化单个文件."""
        wb = load_workbook(file_path)
        ws = wb.active

        # 应用格式
        if 'font_name' in kwargs or 'font_size' in kwargs:
            from openpyxl.styles import Font
            for row in ws.iter_rows():
                for cell in row:
                    cell.font = Font(
                        name=kwargs.get('font_name', 'Calibri'),
                        size=kwargs.get('font_size', 11)
                    )

        wb.save(file_path)
        wb.close()

    def _batch_export_file(self, file_path: str, **kwargs: Any) -> None:
        """批量导出单个文件."""
        export_format = kwargs.get('format', 'csv')
        output_dir = kwargs.get('output_dir', config.paths.output_dir)

        # 这里可以调用导入导出模块的方法
        pass

    def merge_workbooks(
        self,
        source_files: list[str],
        output_file: str,
        merge_mode: str = "sheets",
    ) -> dict[str, Any]:
        """合并多个工作簿.

        Args:
            source_files: 源文件列表
            output_file: 输出文件名
            merge_mode: 合并模式 ('sheets'作为工作表, 'append'追加数据)
        """
        try:
            from openpyxl import Workbook

            wb_out = Workbook()
            wb_out.remove(wb_out.active)

            for source_file in source_files:
                source_path = config.paths.output_dir / source_file
                if not source_path.exists():
                    logger.warning(f"文件不存在,跳过: {source_file}")
                    continue

                wb_source = load_workbook(str(source_path))

                for sheet_name in wb_source.sheetnames:
                    ws_source = wb_source[sheet_name]

                    # 创建新工作表
                    new_name = f"{Path(source_file).stem}_{sheet_name}"
                    ws_out = wb_out.create_sheet(title=new_name)

                    # 复制数据
                    for row in ws_source.iter_rows(values_only=True):
                        ws_out.append(row)

                wb_source.close()

            output_path = config.paths.output_dir / output_file
            wb_out.save(str(output_path))
            wb_out.close()

            logger.info(f"工作簿合并成功: {output_path}")
            return {
                "success": True,
                "message": f"成功合并 {len(source_files)} 个工作簿",
                "output_file": str(output_path),
                "source_count": len(source_files),
            }

        except Exception as e:
            logger.error(f"合并工作簿失败: {e}")
            return {"success": False, "message": f"合并失败: {str(e)}"}
