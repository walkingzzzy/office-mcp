"""Excel 单元格高级操作模块."""

from typing import Any, Literal
import re

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
from loguru import logger

from office_mcp_server.config import config
from office_mcp_server.utils.file_manager import FileManager


class ExcelCellAdvancedOperations:
    """Excel 单元格高级操作类."""

    def __init__(self) -> None:
        """初始化单元格高级操作类."""
        self.file_manager = FileManager()

    def insert_cells(
        self,
        filename: str,
        sheet_name: str,
        cell: str,
        shift: Literal["down", "right"] = "down",
    ) -> dict[str, Any]:
        """插入单元格并移动其他单元格.

        Args:
            filename: 文件名
            sheet_name: 工作表名称
            cell: 起始单元格 (如 'B2')
            shift: 移动方向 ('down'向下, 'right'向右)
        """
        try:
            file_path = config.paths.output_dir / filename
            self.file_manager.validate_file_path(file_path, must_exist=True)

            wb = load_workbook(str(file_path))

            if sheet_name not in wb.sheetnames:
                raise ValueError(f"工作表 '{sheet_name}' 不存在")

            ws = wb[sheet_name]

            # 解析单元格位置 (如 'B2' -> 列字母='B', 行号=2)
            match = re.match(r'^([A-Z]+)(\d+)$', cell.upper())
            if not match:
                raise ValueError(f"无效的单元格格式: {cell}")
            col_letter = match.group(1)
            row = int(match.group(2))
            col_idx = column_index_from_string(col_letter)

            if shift == "down":
                # 向下移动：插入行
                ws.insert_rows(row, 1)
            elif shift == "right":
                # 向右移动：插入列
                ws.insert_cols(col_idx, 1)
            else:
                raise ValueError(f"不支持的移动方向: {shift}")

            wb.save(str(file_path))
            wb.close()

            logger.info(f"插入单元格成功: {file_path}, 单元格: {cell}, 方向: {shift}")
            return {
                "success": True,
                "message": f"成功插入单元格并向{shift}移动",
                "filename": str(file_path),
                "sheet_name": sheet_name,
                "cell": cell,
                "shift": shift,
            }

        except Exception as e:
            logger.error(f"插入单元格失败: {e}")
            return {"success": False, "message": f"操作失败: {str(e)}"}

    def delete_cells(
        self,
        filename: str,
        sheet_name: str,
        cell: str,
        shift: Literal["up", "left"] = "up",
    ) -> dict[str, Any]:
        """删除单元格并移动其他单元格.

        Args:
            filename: 文件名
            sheet_name: 工作表名称
            cell: 起始单元格 (如 'B2')
            shift: 移动方向 ('up'向上, 'left'向左)
        """
        try:
            file_path = config.paths.output_dir / filename
            self.file_manager.validate_file_path(file_path, must_exist=True)

            wb = load_workbook(str(file_path))

            if sheet_name not in wb.sheetnames:
                raise ValueError(f"工作表 '{sheet_name}' 不存在")

            ws = wb[sheet_name]

            # 解析单元格位置 (如 'B2' -> 列字母='B', 行号=2)
            match = re.match(r'^([A-Z]+)(\d+)$', cell.upper())
            if not match:
                raise ValueError(f"无效的单元格格式: {cell}")
            col_letter = match.group(1)
            row = int(match.group(2))
            col_idx = column_index_from_string(col_letter)

            if shift == "up":
                # 向上移动：删除行
                ws.delete_rows(row, 1)
            elif shift == "left":
                # 向左移动：删除列
                ws.delete_cols(col_idx, 1)
            else:
                raise ValueError(f"不支持的移动方向: {shift}")

            wb.save(str(file_path))
            wb.close()

            logger.info(f"删除单元格成功: {file_path}, 单元格: {cell}, 方向: {shift}")
            return {
                "success": True,
                "message": f"成功删除单元格并向{shift}移动",
                "filename": str(file_path),
                "sheet_name": sheet_name,
                "cell": cell,
                "shift": shift,
            }

        except Exception as e:
            logger.error(f"删除单元格失败: {e}")
            return {"success": False, "message": f"操作失败: {str(e)}"}

    def insert_cell_range(
        self,
        filename: str,
        sheet_name: str,
        start_cell: str,
        end_cell: str,
        shift: Literal["down", "right"] = "down",
    ) -> dict[str, Any]:
        """插入单元格范围并移动其他单元格.

        Args:
            filename: 文件名
            sheet_name: 工作表名称
            start_cell: 起始单元格 (如 'B2')
            end_cell: 结束单元格 (如 'D4')
            shift: 移动方向 ('down'向下, 'right'向右)
        """
        try:
            file_path = config.paths.output_dir / filename
            self.file_manager.validate_file_path(file_path, must_exist=True)

            wb = load_workbook(str(file_path))

            if sheet_name not in wb.sheetnames:
                raise ValueError(f"工作表 '{sheet_name}' 不存在")

            ws = wb[sheet_name]

            # 解析单元格位置
            start_match = re.match(r'^([A-Z]+)(\d+)$', start_cell.upper())
            end_match = re.match(r'^([A-Z]+)(\d+)$', end_cell.upper())
            if not start_match or not end_match:
                raise ValueError(f"无效的单元格格式: {start_cell} 或 {end_cell}")

            start_col_letter = start_match.group(1)
            start_row = int(start_match.group(2))
            end_col_letter = end_match.group(1)
            end_row = int(end_match.group(2))

            start_col_idx = column_index_from_string(start_col_letter)
            end_col_idx = column_index_from_string(end_col_letter)

            if shift == "down":
                # 向下移动：插入多行
                row_count = end_row - start_row + 1
                ws.insert_rows(start_row, row_count)
            elif shift == "right":
                # 向右移动：插入多列
                col_count = end_col_idx - start_col_idx + 1
                ws.insert_cols(start_col_idx, col_count)
            else:
                raise ValueError(f"不支持的移动方向: {shift}")

            wb.save(str(file_path))
            wb.close()

            logger.info(f"插入单元格范围成功: {file_path}, 范围: {start_cell}:{end_cell}")
            return {
                "success": True,
                "message": f"成功插入单元格范围并向{shift}移动",
                "filename": str(file_path),
                "sheet_name": sheet_name,
                "range": f"{start_cell}:{end_cell}",
                "shift": shift,
            }

        except Exception as e:
            logger.error(f"插入单元格范围失败: {e}")
            return {"success": False, "message": f"操作失败: {str(e)}"}

    def delete_cell_range(
        self,
        filename: str,
        sheet_name: str,
        start_cell: str,
        end_cell: str,
        shift: Literal["up", "left"] = "up",
    ) -> dict[str, Any]:
        """删除单元格范围并移动其他单元格.

        Args:
            filename: 文件名
            sheet_name: 工作表名称
            start_cell: 起始单元格 (如 'B2')
            end_cell: 结束单元格 (如 'D4')
            shift: 移动方向 ('up'向上, 'left'向左)
        """
        try:
            file_path = config.paths.output_dir / filename
            self.file_manager.validate_file_path(file_path, must_exist=True)

            wb = load_workbook(str(file_path))

            if sheet_name not in wb.sheetnames:
                raise ValueError(f"工作表 '{sheet_name}' 不存在")

            ws = wb[sheet_name]

            # 解析单元格位置
            start_match = re.match(r'^([A-Z]+)(\d+)$', start_cell.upper())
            end_match = re.match(r'^([A-Z]+)(\d+)$', end_cell.upper())
            if not start_match or not end_match:
                raise ValueError(f"无效的单元格格式: {start_cell} 或 {end_cell}")

            start_col_letter = start_match.group(1)
            start_row = int(start_match.group(2))
            end_col_letter = end_match.group(1)
            end_row = int(end_match.group(2))

            start_col_idx = column_index_from_string(start_col_letter)
            end_col_idx = column_index_from_string(end_col_letter)

            if shift == "up":
                # 向上移动：删除多行
                row_count = end_row - start_row + 1
                ws.delete_rows(start_row, row_count)
            elif shift == "left":
                # 向左移动：删除多列
                col_count = end_col_idx - start_col_idx + 1
                ws.delete_cols(start_col_idx, col_count)
            else:
                raise ValueError(f"不支持的移动方向: {shift}")

            wb.save(str(file_path))
            wb.close()

            logger.info(f"删除单元格范围成功: {file_path}, 范围: {start_cell}:{end_cell}")
            return {
                "success": True,
                "message": f"成功删除单元格范围并向{shift}移动",
                "filename": str(file_path),
                "sheet_name": sheet_name,
                "range": f"{start_cell}:{end_cell}",
                "shift": shift,
            }

        except Exception as e:
            logger.error(f"删除单元格范围失败: {e}")
            return {"success": False, "message": f"操作失败: {str(e)}"}
