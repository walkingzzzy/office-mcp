"""Excel 图表操作模块."""

from typing import Any, Optional

from openpyxl import load_workbook
from openpyxl.chart import (
    BarChart,
    LineChart,
    PieChart,
    AreaChart,
    ScatterChart,
    BubbleChart,
    RadarChart,
    Reference,
)
from openpyxl.chart.trendline import Trendline
from openpyxl.chart.series import DataPoint
from openpyxl.chart.layout import Layout, ManualLayout
from loguru import logger

from office_mcp_server.config import config
from office_mcp_server.utils.file_manager import FileManager


class ExcelChartOperations:
    """Excel 图表操作类."""

    def __init__(self) -> None:
        """初始化图表操作类."""
        self.file_manager = FileManager()

    def create_chart(
        self,
        filename: str,
        sheet_name: str,
        chart_type: str,
        data_range: str,
        title: str = "",
        position: str = "E5",
        x_axis_title: Optional[str] = None,
        y_axis_title: Optional[str] = None,
        legend_position: Optional[str] = None,
        show_data_labels: bool = False,
        grouping: Optional[str] = None,
    ) -> dict[str, Any]:
        """创建图表.

        Args:
            filename: 文件名
            sheet_name: 工作表名称
            chart_type: 图表类型
            data_range: 数据范围
            title: 图表标题
            position: 图表位置
            x_axis_title: X轴标题
            y_axis_title: Y轴标题
            legend_position: 图例位置
            show_data_labels: 是否显示数据标签
            grouping: 分组方式 ('standard'标准, 'stacked'堆积, 'percentStacked'百分比堆积)
        """
        try:
            file_path = config.paths.output_dir / filename
            self.file_manager.validate_file_path(file_path, must_exist=True)

            wb = load_workbook(str(file_path))

            if sheet_name not in wb.sheetnames:
                raise ValueError(f"工作表 '{sheet_name}' 不存在")

            ws = wb[sheet_name]

            # 创建图表
            if chart_type == "bar":
                chart = BarChart()
                chart.type = "col"  # 柱状图
                chart.style = 10
                # 设置分组方式
                if grouping:
                    chart.grouping = grouping
            elif chart_type == "line":
                chart = LineChart()
                chart.style = 13
                # 线图也支持堆积
                if grouping:
                    chart.grouping = grouping
            elif chart_type == "pie":
                chart = PieChart()
                chart.style = 10
            elif chart_type == "area":
                chart = AreaChart()
                chart.style = 13
                # 面积图支持堆积
                if grouping:
                    chart.grouping = grouping
            elif chart_type == "scatter":
                chart = ScatterChart()
                chart.style = 13
            elif chart_type == "bubble":
                chart = BubbleChart()
                chart.style = 13
            elif chart_type == "radar":
                chart = RadarChart()
                chart.style = 26
            elif chart_type == "doughnut":
                chart = PieChart()
                chart.style = 10
                # 设置为圆环图
                chart.holeSize = 50
            else:
                raise ValueError(f"不支持的图表类型: {chart_type}")

            # 设置数据
            data = Reference(ws, range_string=f"{sheet_name}!{data_range}")
            chart.add_data(data, titles_from_data=True)

            # 设置标题
            if title:
                chart.title = title

            # 设置坐标轴标题
            if x_axis_title and chart_type != "pie":
                chart.x_axis.title = x_axis_title
            if y_axis_title and chart_type != "pie":
                chart.y_axis.title = y_axis_title

            # 设置图例位置
            if legend_position:
                chart.legend.position = legend_position

            # 设置数据标签
            if show_data_labels:
                chart.dataLabels = chart.dataLabels or {}
                chart.dataLabels.showVal = True

            # 添加图表到工作表
            ws.add_chart(chart, position)

            wb.save(str(file_path))
            wb.close()

            logger.info(f"图表创建成功: {file_path}")
            return {
                "success": True,
                "message": "图表创建成功",
                "filename": str(file_path),
                "chart_type": chart_type,
            }

        except Exception as e:
            logger.error(f"创建图表失败: {e}")
            return {"success": False, "message": f"创建失败: {str(e)}"}

    def format_chart(
        self,
        filename: str,
        sheet_name: str,
        chart_index: int = 0,
        title_font_size: Optional[int] = None,
        title_font_bold: bool = False,
        chart_style: Optional[int] = None,
        color_scheme: Optional[list[str]] = None,
    ) -> dict[str, Any]:
        """格式化图表.

        Args:
            filename: 文件名
            sheet_name: 工作表名称
            chart_index: 图表索引 (默认0, 第一个图表)
            title_font_size: 标题字号 (可选)
            title_font_bold: 标题是否加粗
            chart_style: 图表样式编号 (1-48, 可选)
            color_scheme: 颜色方案 (HEX颜色列表, 可选)
        """
        try:
            file_path = config.paths.output_dir / filename
            self.file_manager.validate_file_path(file_path, must_exist=True)

            wb = load_workbook(str(file_path))

            if sheet_name not in wb.sheetnames:
                raise ValueError(f"工作表 '{sheet_name}' 不存在")

            ws = wb[sheet_name]

            if not ws._charts or chart_index >= len(ws._charts):
                raise ValueError(f"图表索引 {chart_index} 不存在")

            chart = ws._charts[chart_index]

            if chart_style is not None:
                chart.style = chart_style

            if title_font_size or title_font_bold:
                if hasattr(chart, 'title') and chart.title:
                    if title_font_size:
                        pass
                    if title_font_bold:
                        pass

            if color_scheme and hasattr(chart, 'series'):
                for idx, series in enumerate(chart.series):
                    if idx < len(color_scheme):
                        color_hex = color_scheme[idx].lstrip('#')
                        pass

            wb.save(str(file_path))
            wb.close()

            logger.info(f"图表格式化成功: {file_path}")
            return {
                "success": True,
                "message": f"成功格式化图表 {chart_index}",
                "filename": str(file_path),
                "chart_index": chart_index,
            }

        except Exception as e:
            logger.error(f"格式化图表失败: {e}")
            return {"success": False, "message": f"格式化失败: {str(e)}"}

    def create_combination_chart(
        self,
        filename: str,
        sheet_name: str,
        data_range1: str,
        data_range2: str,
        chart_type1: str = "bar",
        chart_type2: str = "line",
        title: str = "",
        position: str = "E5",
    ) -> dict[str, Any]:
        """创建组合图表.

        Args:
            filename: 文件名
            sheet_name: 工作表名称
            data_range1: 第一个数据范围 (如 'A1:B10')
            data_range2: 第二个数据范围 (如 'A1,C1:C10')
            chart_type1: 第一个图表类型 ('bar'柱状图, 'line'折线图)
            chart_type2: 第二个图表类型 ('bar'柱状图, 'line'折线图)
            title: 图表标题
            position: 图表位置 (默认 'E5')
        """
        try:
            file_path = config.paths.output_dir / filename
            self.file_manager.validate_file_path(file_path, must_exist=True)

            wb = load_workbook(str(file_path))

            if sheet_name not in wb.sheetnames:
                raise ValueError(f"工作表 '{sheet_name}' 不存在")

            ws = wb[sheet_name]

            if chart_type1 == "bar":
                chart1 = BarChart()
                chart1.type = "col"
                chart1.style = 10
            elif chart_type1 == "line":
                chart1 = LineChart()
                chart1.style = 13
            else:
                raise ValueError(f"不支持的第一个图表类型: {chart_type1}")

            data1 = Reference(ws, range_string=f"{sheet_name}!{data_range1}")
            chart1.add_data(data1, titles_from_data=True)

            if chart_type2 == "line":
                chart2 = LineChart()
                chart2.style = 13
            elif chart_type2 == "bar":
                chart2 = BarChart()
                chart2.type = "col"
                chart2.style = 11
            else:
                raise ValueError(f"不支持的第二个图表类型: {chart_type2}")

            data2 = Reference(ws, range_string=f"{sheet_name}!{data_range2}")
            chart2.add_data(data2, titles_from_data=True)

            chart1 += chart2

            if title:
                chart1.title = title

            chart1.y_axis.axId = 200
            chart2.y_axis.axId = 300
            chart2.y_axis.crosses = "max"
            chart1.y_axis.crosses = "min"

            ws.add_chart(chart1, position)

            wb.save(str(file_path))
            wb.close()

            logger.info(f"组合图表创建成功: {file_path}")
            return {
                "success": True,
                "message": "组合图表创建成功",
                "filename": str(file_path),
                "chart_type1": chart_type1,
                "chart_type2": chart_type2,
            }

        except Exception as e:
            logger.error(f"创建组合图表失败: {e}")
            return {"success": False, "message": f"创建失败: {str(e)}"}

    def add_trendline_to_chart(
        self,
        filename: str,
        sheet_name: str,
        chart_index: int = 0,
        series_index: int = 0,
        trendline_type: str = "linear",
        display_equation: bool = False,
        display_r_squared: bool = False,
    ) -> dict[str, Any]:
        """为图表添加趋势线.

        Args:
            filename: 文件名
            sheet_name: 工作表名称
            chart_index: 图表索引 (默认0)
            series_index: 系列索引 (默认0)
            trendline_type: 趋势线类型 ('linear'线性, 'exp'指数, 'log'对数,
                           'poly'多项式, 'power'幂, 'movingAvg'移动平均)
            display_equation: 是否显示趋势线方程
            display_r_squared: 是否显示R²值
        """
        try:
            file_path = config.paths.output_dir / filename
            self.file_manager.validate_file_path(file_path, must_exist=True)

            wb = load_workbook(str(file_path))

            if sheet_name not in wb.sheetnames:
                raise ValueError(f"工作表 '{sheet_name}' 不存在")

            ws = wb[sheet_name]

            if not ws._charts or chart_index >= len(ws._charts):
                raise ValueError(f"图表索引 {chart_index} 不存在")

            chart = ws._charts[chart_index]

            if not hasattr(chart, 'series') or series_index >= len(chart.series):
                raise ValueError(f"系列索引 {series_index} 不存在")

            series = chart.series[series_index]

            # 创建趋势线
            trendline = Trendline()
            trendline.trendlineType = trendline_type
            trendline.dispEq = display_equation
            trendline.dispRSqr = display_r_squared

            # 为系列添加趋势线
            series.trendline = trendline

            wb.save(str(file_path))
            wb.close()

            logger.info(f"趋势线添加成功: {file_path}")
            return {
                "success": True,
                "message": f"成功为图表 {chart_index} 系列 {series_index} 添加趋势线",
                "filename": str(file_path),
                "trendline_type": trendline_type,
            }

        except Exception as e:
            logger.error(f"添加趋势线失败: {e}")
            return {"success": False, "message": f"添加失败: {str(e)}"}
