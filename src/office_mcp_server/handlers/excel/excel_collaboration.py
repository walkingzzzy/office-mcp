"""Excel 协作功能模块."""

from typing import Any, Optional
from datetime import datetime

from openpyxl import load_workbook
from openpyxl.comments import Comment
from loguru import logger

from office_mcp_server.config import config
from office_mcp_server.utils.file_manager import FileManager


class ExcelCollaborationOperations:
    """Excel 协作功能操作类."""

    def __init__(self) -> None:
        """初始化协作功能操作类."""
        self.file_manager = FileManager()

    def add_comment(
        self,
        filename: str,
        sheet_name: str,
        cell: str,
        comment_text: str,
        author: str = "User",
    ) -> dict[str, Any]:
        """添加批注.

        Args:
            filename: 文件名
            sheet_name: 工作表名称
            cell: 单元格引用 (如 'A1')
            comment_text: 批注文本
            author: 作者名称 (默认 'User')
        """
        try:
            file_path = config.paths.output_dir / filename
            self.file_manager.validate_file_path(file_path, must_exist=True)

            wb = load_workbook(str(file_path))

            if sheet_name not in wb.sheetnames:
                raise ValueError(f"工作表 '{sheet_name}' 不存在")

            ws = wb[sheet_name]

            comment = Comment(comment_text, author)
            ws[cell].comment = comment

            wb.save(str(file_path))
            wb.close()

            logger.info(f"添加批注成功: {file_path}, 单元格: {cell}")
            return {
                "success": True,
                "message": f"成功在单元格 {cell} 添加批注",
                "filename": str(file_path),
                "sheet_name": sheet_name,
                "cell": cell,
                "author": author,
            }

        except Exception as e:
            logger.error(f"添加批注失败: {e}")
            return {"success": False, "message": f"添加失败: {str(e)}"}

    def get_comment(
        self,
        filename: str,
        sheet_name: str,
        cell: str,
    ) -> dict[str, Any]:
        """获取批注.

        Args:
            filename: 文件名
            sheet_name: 工作表名称
            cell: 单元格引用 (如 'A1')
        """
        try:
            file_path = config.paths.output_dir / filename
            self.file_manager.validate_file_path(file_path, must_exist=True)

            wb = load_workbook(str(file_path))

            if sheet_name not in wb.sheetnames:
                raise ValueError(f"工作表 '{sheet_name}' 不存在")

            ws = wb[sheet_name]
            comment = ws[cell].comment

            if comment:
                result = {
                    "success": True,
                    "filename": str(file_path),
                    "sheet_name": sheet_name,
                    "cell": cell,
                    "text": comment.text,
                    "author": comment.author,
                }
            else:
                result = {
                    "success": True,
                    "filename": str(file_path),
                    "sheet_name": sheet_name,
                    "cell": cell,
                    "text": None,
                    "message": "该单元格没有批注",
                }

            wb.close()

            logger.info(f"获取批注成功: {file_path}, 单元格: {cell}")
            return result

        except Exception as e:
            logger.error(f"获取批注失败: {e}")
            return {"success": False, "message": f"获取失败: {str(e)}"}

    def delete_comment(
        self,
        filename: str,
        sheet_name: str,
        cell: str,
    ) -> dict[str, Any]:
        """删除批注.

        Args:
            filename: 文件名
            sheet_name: 工作表名称
            cell: 单元格引用 (如 'A1')
        """
        try:
            file_path = config.paths.output_dir / filename
            self.file_manager.validate_file_path(file_path, must_exist=True)

            wb = load_workbook(str(file_path))

            if sheet_name not in wb.sheetnames:
                raise ValueError(f"工作表 '{sheet_name}' 不存在")

            ws = wb[sheet_name]
            ws[cell].comment = None

            wb.save(str(file_path))
            wb.close()

            logger.info(f"删除批注成功: {file_path}, 单元格: {cell}")
            return {
                "success": True,
                "message": f"成功删除单元格 {cell} 的批注",
                "filename": str(file_path),
                "sheet_name": sheet_name,
                "cell": cell,
            }

        except Exception as e:
            logger.error(f"删除批注失败: {e}")
            return {"success": False, "message": f"删除失败: {str(e)}"}

    def list_all_comments(
        self,
        filename: str,
        sheet_name: str,
    ) -> dict[str, Any]:
        """列出所有批注.

        Args:
            filename: 文件名
            sheet_name: 工作表名称
        """
        try:
            file_path = config.paths.output_dir / filename
            self.file_manager.validate_file_path(file_path, must_exist=True)

            wb = load_workbook(str(file_path))

            if sheet_name not in wb.sheetnames:
                raise ValueError(f"工作表 '{sheet_name}' 不存在")

            ws = wb[sheet_name]

            comments = []
            for row in ws.iter_rows():
                for cell in row:
                    if cell.comment:
                        comments.append({
                            "cell": cell.coordinate,
                            "text": cell.comment.text,
                            "author": cell.comment.author,
                        })

            wb.close()

            logger.info(f"列出批注成功: {file_path}, 共 {len(comments)} 个")
            return {
                "success": True,
                "filename": str(file_path),
                "sheet_name": sheet_name,
                "count": len(comments),
                "comments": comments,
            }

        except Exception as e:
            logger.error(f"列出批注失败: {e}")
            return {"success": False, "message": f"列出失败: {str(e)}"}
