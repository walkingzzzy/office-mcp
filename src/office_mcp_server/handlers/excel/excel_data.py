"""Excel 数据操作模块."""

from typing import Any, Optional, Union

from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation
from loguru import logger

from office_mcp_server.config import config
from office_mcp_server.utils.file_manager import FileManager


class ExcelDataOperations:
    """Excel 数据操作类."""

    def __init__(self) -> None:
        """初始化数据操作类."""
        self.file_manager = FileManager()

    def sort_data(
        self,
        filename: str,
        sheet_name: str,
        data_range: str,
        sort_by_column: int = 0,
        ascending: bool = True,
        sort_keys: Optional[list[dict[str, Any]]] = None,
    ) -> dict[str, Any]:
        """对数据进行排序.

        Args:
            filename: 文件名
            sheet_name: 工作表名称
            data_range: 数据范围
            sort_by_column: 排序列索引（单列排序时使用）
            ascending: 是否升序（单列排序时使用）
            sort_keys: 多列排序键列表，格式: [{"column": 0, "ascending": True}, {"column": 1, "ascending": False}]
        """
        try:
            file_path = config.paths.output_dir / filename
            self.file_manager.validate_file_path(file_path, must_exist=True)

            wb = load_workbook(str(file_path))

            if sheet_name not in wb.sheetnames:
                raise ValueError(f"工作表 '{sheet_name}' 不存在")

            ws = wb[sheet_name]

            # 获取数据范围
            data_cells = ws[data_range]

            # 将数据转换为列表
            data_list = []
            for row in data_cells:
                row_data = [cell.value for cell in row]
                data_list.append(row_data)

            # 多列排序
            if sort_keys and len(sort_keys) > 0:
                # 验证所有排序列
                for key in sort_keys:
                    col_idx = key.get("column", 0)
                    if col_idx >= len(data_list[0]):
                        raise ValueError(f"排序列索引 {col_idx} 超出范围")

                # 多列排序
                def multi_key(row):
                    keys = []
                    for key in sort_keys:
                        col_idx = key.get("column", 0)
                        asc = key.get("ascending", True)
                        value = row[col_idx] if row[col_idx] is not None else ""
                        # 对于降序，需要反转排序键
                        if not asc:
                            # 对于数字，使用负数；对于字符串，需要特殊处理
                            if isinstance(value, (int, float)):
                                keys.append(-value)
                            else:
                                keys.append((False, value))  # 标记为降序
                        else:
                            if isinstance(value, (int, float)):
                                keys.append(value)
                            else:
                                keys.append((True, value))  # 标记为升序
                    return keys

                # 自定义比较函数处理混合排序
                from functools import cmp_to_key

                def compare(a, b):
                    for key in sort_keys:
                        col_idx = key.get("column", 0)
                        asc = key.get("ascending", True)
                        val_a = a[col_idx] if a[col_idx] is not None else ""
                        val_b = b[col_idx] if b[col_idx] is not None else ""

                        # 比较值
                        if val_a < val_b:
                            result = -1
                        elif val_a > val_b:
                            result = 1
                        else:
                            result = 0

                        # 如果不相等，应用排序方向
                        if result != 0:
                            return result if asc else -result

                    return 0

                sorted_data = sorted(data_list, key=cmp_to_key(compare))
            else:
                # 单列排序
                if sort_by_column >= len(data_list[0]):
                    raise ValueError(f"排序列索引 {sort_by_column} 超出范围")

                sorted_data = sorted(
                    data_list,
                    key=lambda x: x[sort_by_column] if x[sort_by_column] is not None else "",
                    reverse=not ascending
                )

            # 写回数据
            from openpyxl.utils import range_boundaries
            min_col, min_row, max_col, max_row = range_boundaries(data_range)

            for i, row_data in enumerate(sorted_data):
                for j, value in enumerate(row_data):
                    ws.cell(row=min_row + i, column=min_col + j, value=value)

            wb.save(str(file_path))
            wb.close()

            logger.info(f"数据排序成功: {file_path}")
            return {
                "success": True,
                "message": "数据排序成功",
                "filename": str(file_path),
                "sort_by_column": sort_by_column if not sort_keys else None,
                "ascending": ascending if not sort_keys else None,
                "sort_keys": sort_keys if sort_keys else None,
            }

        except Exception as e:
            logger.error(f"排序数据失败: {e}")
            return {"success": False, "message": f"排序失败: {str(e)}"}

    def filter_data(
        self,
        filename: str,
        sheet_name: str,
        data_range: str,
        filter_column: Optional[int] = None,
        filter_value: Optional[str] = None,
        filter_operator: str = "equals",
        enable_autofilter: bool = True,
    ) -> dict[str, Any]:
        """对数据进行筛选."""
        try:
            file_path = config.paths.output_dir / filename
            self.file_manager.validate_file_path(file_path, must_exist=True)

            wb = load_workbook(str(file_path))

            if sheet_name not in wb.sheetnames:
                raise ValueError(f"工作表 '{sheet_name}' 不存在")

            ws = wb[sheet_name]

            # 启用自动筛选
            if enable_autofilter:
                ws.auto_filter.ref = data_range

            # 如果指定了筛选列和值,应用筛选
            if filter_column is not None and filter_value is not None:
                from openpyxl.utils import range_boundaries
                from openpyxl.worksheet.filters import FilterColumn, CustomFilters, CustomFilter

                min_col, min_row, max_col, max_row = range_boundaries(data_range)

                if filter_column >= (max_col - min_col + 1):
                    raise ValueError(f"筛选列索引 {filter_column} 超出范围")

                # 创建筛选条件
                operator_map = {
                    'equals': 'equal',
                    'notEquals': 'notEqual',
                    'greaterThan': 'greaterThan',
                    'lessThan': 'lessThan',
                    'greaterThanOrEqual': 'greaterThanOrEqual',
                    'lessThanOrEqual': 'lessThanOrEqual',
                }

                if filter_operator in operator_map:
                    filter_op = operator_map[filter_operator]
                    custom_filter = CustomFilter(operator=filter_op, val=filter_value)
                    custom_filters = CustomFilters(customFilter=[custom_filter])
                    filter_column_obj = FilterColumn(colId=filter_column, customFilters=custom_filters)
                    ws.auto_filter.filterColumn.append(filter_column_obj)

            wb.save(str(file_path))
            wb.close()

            logger.info(f"数据筛选成功: {file_path}")
            return {
                "success": True,
                "message": "数据筛选成功",
                "filename": str(file_path),
                "data_range": data_range,
            }

        except Exception as e:
            logger.error(f"筛选数据失败: {e}")
            return {"success": False, "message": f"筛选失败: {str(e)}"}

    def set_data_validation(
        self,
        filename: str,
        sheet_name: str,
        cell_range: str,
        validation_type: str,
        operator: Optional[str] = None,
        formula1: Optional[str] = None,
        formula2: Optional[str] = None,
        allow_blank: bool = True,
        show_dropdown: bool = True,
        prompt_title: Optional[str] = None,
        prompt: Optional[str] = None,
        error_title: Optional[str] = None,
        error: Optional[str] = None,
    ) -> dict[str, Any]:
        """设置数据验证."""
        try:
            file_path = config.paths.output_dir / filename
            self.file_manager.validate_file_path(file_path, must_exist=True)

            wb = load_workbook(str(file_path))

            if sheet_name not in wb.sheetnames:
                raise ValueError(f"工作表 '{sheet_name}' 不存在")

            ws = wb[sheet_name]

            # 创建数据验证对象
            if validation_type == 'list':
                # 列表验证
                if not formula1:
                    raise ValueError("列表验证需要指定formula1（逗号分隔的值）")
                dv = DataValidation(
                    type="list",
                    formula1=f'"{formula1}"',
                    allow_blank=allow_blank,
                    showDropDown=not show_dropdown  # 注意:这个参数是反的
                )

            elif validation_type == 'whole':
                # 整数验证
                dv = DataValidation(
                    type="whole",
                    operator=operator or "between",
                    formula1=formula1,
                    formula2=formula2,
                    allow_blank=allow_blank
                )

            elif validation_type == 'decimal':
                # 小数验证
                dv = DataValidation(
                    type="decimal",
                    operator=operator or "between",
                    formula1=formula1,
                    formula2=formula2,
                    allow_blank=allow_blank
                )

            elif validation_type == 'date':
                # 日期验证
                dv = DataValidation(
                    type="date",
                    operator=operator or "between",
                    formula1=formula1,
                    formula2=formula2,
                    allow_blank=allow_blank
                )

            elif validation_type == 'time':
                # 时间验证
                dv = DataValidation(
                    type="time",
                    operator=operator or "between",
                    formula1=formula1,
                    formula2=formula2,
                    allow_blank=allow_blank
                )

            elif validation_type == 'textLength':
                # 文本长度验证
                dv = DataValidation(
                    type="textLength",
                    operator=operator or "between",
                    formula1=formula1,
                    formula2=formula2,
                    allow_blank=allow_blank
                )

            elif validation_type == 'custom':
                # 自定义公式验证
                if not formula1:
                    raise ValueError("自定义验证需要指定formula1")
                dv = DataValidation(
                    type="custom",
                    formula1=formula1,
                    allow_blank=allow_blank
                )

            else:
                raise ValueError(f"不支持的验证类型: {validation_type}")

            # 设置提示信息
            if prompt_title and prompt:
                dv.prompt = prompt
                dv.promptTitle = prompt_title
                dv.showInputMessage = True

            # 设置错误信息
            if error_title and error:
                dv.error = error
                dv.errorTitle = error_title
                dv.showErrorMessage = True

            # 应用数据验证
            ws.add_data_validation(dv)
            dv.add(cell_range)

            wb.save(str(file_path))
            wb.close()

            logger.info(f"数据验证设置成功: {file_path}")
            return {
                "success": True,
                "message": f"数据验证已应用到 {cell_range}",
                "filename": str(file_path),
                "validation_type": validation_type,
                "cell_range": cell_range,
            }

        except Exception as e:
            logger.error(f"设置数据验证失败: {e}")
            return {"success": False, "message": f"设置失败: {str(e)}"}

    def manage_worksheets(
        self,
        filename: str,
        operation: str,
        sheet_name: Optional[str] = None,
        new_name: Optional[str] = None,
        target_index: Optional[int] = None,
    ) -> dict[str, Any]:
        """管理工作表."""
        try:
            file_path = config.paths.output_dir / filename
            self.file_manager.validate_file_path(file_path, must_exist=True)

            wb = load_workbook(str(file_path))

            if operation == 'create':
                # 创建新工作表
                if not sheet_name:
                    sheet_name = f"Sheet{len(wb.sheetnames) + 1}"
                wb.create_sheet(title=sheet_name)
                message = f"工作表 '{sheet_name}' 创建成功"

            elif operation == 'delete':
                # 删除工作表
                if not sheet_name:
                    raise ValueError("删除工作表需要指定sheet_name")
                if sheet_name not in wb.sheetnames:
                    raise ValueError(f"工作表 '{sheet_name}' 不存在")
                wb.remove(wb[sheet_name])
                message = f"工作表 '{sheet_name}' 删除成功"

            elif operation == 'rename':
                # 重命名工作表
                if not sheet_name or not new_name:
                    raise ValueError("重命名工作表需要指定sheet_name和new_name")
                if sheet_name not in wb.sheetnames:
                    raise ValueError(f"工作表 '{sheet_name}' 不存在")
                wb[sheet_name].title = new_name
                message = f"工作表 '{sheet_name}' 已重命名为 '{new_name}'"

            elif operation == 'copy':
                # 复制工作表
                if not sheet_name:
                    raise ValueError("复制工作表需要指定sheet_name")
                if sheet_name not in wb.sheetnames:
                    raise ValueError(f"工作表 '{sheet_name}' 不存在")
                source_sheet = wb[sheet_name]
                if not new_name:
                    new_name = f"{sheet_name}_Copy"
                target_sheet = wb.copy_worksheet(source_sheet)
                target_sheet.title = new_name
                message = f"工作表 '{sheet_name}' 已复制为 '{new_name}'"

            elif operation == 'move':
                # 移动工作表
                if not sheet_name or target_index is None:
                    raise ValueError("移动工作表需要指定sheet_name和target_index")
                if sheet_name not in wb.sheetnames:
                    raise ValueError(f"工作表 '{sheet_name}' 不存在")
                wb.move_sheet(sheet_name, offset=target_index - wb.sheetnames.index(sheet_name))
                message = f"工作表 '{sheet_name}' 已移动到位置 {target_index}"

            else:
                raise ValueError(f"不支持的操作类型: {operation}")

            wb.save(str(file_path))
            wb.close()

            logger.info(f"工作表管理成功: {file_path}")
            return {
                "success": True,
                "message": message,
                "filename": str(file_path),
                "operation": operation,
            }

        except Exception as e:
            logger.error(f"管理工作表失败: {e}")
            return {"success": False, "message": f"操作失败: {str(e)}"}

    def create_table(
        self,
        filename: str,
        sheet_name: str,
        table_range: str,
        table_name: str,
        style: str = "TableStyleMedium9",
        show_header: bool = True,
        show_totals: bool = False,
    ) -> dict[str, Any]:
        """创建Excel表格样式.

        Args:
            filename: Excel文件名
            sheet_name: 工作表名称
            table_range: 表格范围，如 "A1:D10"
            table_name: 表格名称（必须唯一）
            style: 表格样式名称（如 TableStyleMedium9, TableStyleLight1等）
            show_header: 是否显示表头
            show_totals: 是否显示汇总行
        """
        try:
            from openpyxl.worksheet.table import Table, TableStyleInfo

            file_path = config.paths.output_dir / filename
            self.file_manager.validate_file_path(file_path, must_exist=True)

            wb = load_workbook(str(file_path))

            if sheet_name not in wb.sheetnames:
                raise ValueError(f"工作表 '{sheet_name}' 不存在")

            ws = wb[sheet_name]

            # 验证表格名称是否已存在
            for table in ws.tables.values():
                if table.name == table_name:
                    raise ValueError(f"表格名称 '{table_name}' 已存在")

            # 创建表格
            tab = Table(displayName=table_name, ref=table_range)

            # 设置表格样式
            style_info = TableStyleInfo(
                name=style,
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False
            )
            tab.tableStyleInfo = style_info

            # 添加表格到工作表
            ws.add_table(tab)

            wb.save(str(file_path))
            wb.close()

            logger.info(f"Excel表格创建成功: {table_name}")
            return {
                "success": True,
                "message": f"表格 '{table_name}' 创建成功",
                "filename": str(file_path),
                "sheet_name": sheet_name,
                "table_name": table_name,
                "table_range": table_range,
                "style": style,
            }

        except Exception as e:
            logger.error(f"创建Excel表格失败: {e}")
            return {"success": False, "message": f"创建表格失败: {str(e)}"}

    def sort_data_by_color(
        self,
        filename: str,
        sheet_name: str,
        data_range: str,
        sort_by_column: int,
        color_order: list[str],
        sort_by: str = "fill",
    ) -> dict[str, Any]:
        """按颜色排序数据.

        Args:
            filename: 文件名
            sheet_name: 工作表名称
            data_range: 数据范围
            sort_by_column: 排序列索引
            color_order: 颜色顺序列表（HEX格式，如 ['#FF0000', '#00FF00', '#0000FF']）
            sort_by: 排序依据 ('fill'按填充色, 'font'按字体色, 默认 'fill')
        """
        try:
            file_path = config.paths.output_dir / filename
            self.file_manager.validate_file_path(file_path, must_exist=True)

            wb = load_workbook(str(file_path))

            if sheet_name not in wb.sheetnames:
                raise ValueError(f"工作表 '{sheet_name}' 不存在")

            ws = wb[sheet_name]

            # 获取数据范围
            from openpyxl.utils import range_boundaries
            min_col, min_row, max_col, max_row = range_boundaries(data_range)

            # 读取数据及颜色信息
            data_with_color = []
            for row_idx in range(min_row, max_row + 1):
                row_data = []
                for col_idx in range(min_col, max_col + 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    row_data.append(cell.value)

                # 获取排序列的颜色
                sort_cell = ws.cell(row=row_idx, column=min_col + sort_by_column)
                if sort_by == "fill":
                    color = sort_cell.fill.start_color.rgb if sort_cell.fill and sort_cell.fill.start_color else None
                else:  # font
                    color = sort_cell.font.color.rgb if sort_cell.font and sort_cell.font.color else None

                # 转换颜色格式
                if color and len(str(color)) == 8:  # ARGB格式
                    color = f"#{str(color)[2:]}"  # 去掉前两位的Alpha通道
                elif color:
                    color = f"#{color}" if not str(color).startswith('#') else color

                data_with_color.append((row_data, color))

            # 按颜色顺序排序
            def color_sort_key(item):
                row_data, color = item
                if color and color.upper() in [c.upper() for c in color_order]:
                    # 找到颜色在color_order中的索引
                    for i, c in enumerate(color_order):
                        if c.upper() == str(color).upper():
                            return i
                return len(color_order)  # 未匹配的颜色排在最后

            sorted_data_with_color = sorted(data_with_color, key=color_sort_key)

            # 写回数据
            for i, (row_data, color) in enumerate(sorted_data_with_color):
                for j, value in enumerate(row_data):
                    ws.cell(row=min_row + i, column=min_col + j, value=value)

            wb.save(str(file_path))
            wb.close()

            logger.info(f"按颜色排序成功: {file_path}")
            return {
                "success": True,
                "message": "按颜色排序成功",
                "filename": str(file_path),
                "sort_by_column": sort_by_column,
                "color_order": color_order,
                "sort_by": sort_by,
            }

        except Exception as e:
            logger.error(f"按颜色排序失败: {e}")
            return {"success": False, "message": f"排序失败: {str(e)}"}

    def copy_filtered_data(
        self,
        filename: str,
        sheet_name: str,
        source_range: str,
        target_cell: str,
        filter_column: int,
        filter_value: str,
        filter_operator: str = "equals",
    ) -> dict[str, Any]:
        """复制筛选结果到新位置.

        Args:
            filename: 文件名
            sheet_name: 工作表名称
            source_range: 源数据范围
            target_cell: 目标起始单元格
            filter_column: 筛选列索引
            filter_value: 筛选值
            filter_operator: 筛选操作符
        """
        try:
            file_path = config.paths.output_dir / filename
            self.file_manager.validate_file_path(file_path, must_exist=True)

            wb = load_workbook(str(file_path))

            if sheet_name not in wb.sheetnames:
                raise ValueError(f"工作表 '{sheet_name}' 不存在")

            ws = wb[sheet_name]

            # 获取数据范围
            from openpyxl.utils import range_boundaries, coordinate_from_string, column_index_from_string
            min_col, min_row, max_col, max_row = range_boundaries(source_range)

            # 读取数据
            data_list = []
            for row_idx in range(min_row, max_row + 1):
                row_data = []
                for col_idx in range(min_col, max_col + 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    row_data.append(cell.value)
                data_list.append(row_data)

            # 筛选数据
            filtered_data = []
            for row in data_list:
                if filter_column >= len(row):
                    continue

                cell_value = str(row[filter_column]) if row[filter_column] is not None else ""
                filter_val_str = str(filter_value)

                # 应用筛选条件
                match = False
                if filter_operator == "equals":
                    match = cell_value == filter_val_str
                elif filter_operator == "notEquals":
                    match = cell_value != filter_val_str
                elif filter_operator == "contains":
                    match = filter_val_str in cell_value
                elif filter_operator == "notContains":
                    match = filter_val_str not in cell_value
                elif filter_operator == "beginsWith":
                    match = cell_value.startswith(filter_val_str)
                elif filter_operator == "endsWith":
                    match = cell_value.endswith(filter_val_str)
                elif filter_operator == "greaterThan":
                    try:
                        match = float(cell_value) > float(filter_val_str)
                    except (ValueError, TypeError):
                        match = False
                elif filter_operator == "lessThan":
                    try:
                        match = float(cell_value) < float(filter_val_str)
                    except (ValueError, TypeError):
                        match = False

                if match:
                    filtered_data.append(row)

            # 解析目标单元格
            target_col_letter, target_row = coordinate_from_string(target_cell)
            target_col = column_index_from_string(target_col_letter)

            # 写入筛选结果
            for i, row_data in enumerate(filtered_data):
                for j, value in enumerate(row_data):
                    ws.cell(row=target_row + i, column=target_col + j, value=value)

            wb.save(str(file_path))
            wb.close()

            logger.info(f"筛选数据复制成功: {file_path}, 复制了 {len(filtered_data)} 行")
            return {
                "success": True,
                "message": f"成功复制 {len(filtered_data)} 行筛选结果",
                "filename": str(file_path),
                "source_range": source_range,
                "target_cell": target_cell,
                "rows_copied": len(filtered_data),
            }

        except Exception as e:
            logger.error(f"复制筛选数据失败: {e}")
            return {"success": False, "message": f"复制失败: {str(e)}"}

