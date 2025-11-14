"""数据脱敏处理模块."""

import hashlib
import re
from typing import Any, Literal, Optional

from openpyxl import load_workbook
from loguru import logger

from office_mcp_server.config import config
from office_mcp_server.utils.file_manager import FileManager


class DataMasking:
    """数据脱敏处理类."""

    def __init__(self) -> None:
        """初始化数据脱敏处理类."""
        self.file_manager = FileManager()

    def mask_data(
        self,
        filename: str,
        sheet_name: str,
        cell_range: str,
        mask_type: Literal["phone", "email", "id_card", "credit_card", "name", "custom", "partial"],
        mask_char: str = "*",
        keep_first: int = 0,
        keep_last: int = 0,
        custom_pattern: Optional[str] = None,
    ) -> dict[str, Any]:
        """数据脱敏处理.

        Args:
            filename: 文件名
            sheet_name: 工作表名称
            cell_range: 单元格范围
            mask_type: 脱敏类型
            mask_char: 脱敏字符 (默认 '*')
            keep_first: 保留前N位
            keep_last: 保留后N位
            custom_pattern: 自定义正则表达式模式

        Returns:
            操作结果
        """
        try:
            file_path = config.paths.output_dir / filename
            self.file_manager.validate_file_path(file_path, must_exist=True)

            wb = load_workbook(str(file_path))

            if sheet_name not in wb.sheetnames:
                raise ValueError(f"工作表 '{sheet_name}' 不存在")

            ws = wb[sheet_name]

            masked_count = 0

            # 遍历范围内的单元格
            for row in ws[cell_range]:
                for cell in row if isinstance(row, tuple) else [row]:
                    if cell.value is None or not isinstance(cell.value, str):
                        continue

                    original_value = str(cell.value)
                    masked_value = self._apply_mask(
                        original_value,
                        mask_type,
                        mask_char,
                        keep_first,
                        keep_last,
                        custom_pattern,
                    )

                    if masked_value != original_value:
                        cell.value = masked_value
                        masked_count += 1

            wb.save(str(file_path))
            wb.close()

            logger.info(f"数据脱敏完成: {file_path}, 处理 {masked_count} 个单元格")
            return {
                "success": True,
                "message": f"成功脱敏 {masked_count} 个单元格",
                "filename": str(file_path),
                "sheet_name": sheet_name,
                "cell_range": cell_range,
                "masked_count": masked_count,
            }

        except Exception as e:
            logger.error(f"数据脱敏失败: {e}")
            return {"success": False, "message": f"脱敏失败: {str(e)}"}

    def _apply_mask(
        self,
        value: str,
        mask_type: str,
        mask_char: str,
        keep_first: int,
        keep_last: int,
        custom_pattern: Optional[str],
    ) -> str:
        """应用脱敏规则."""
        if mask_type == "phone":
            # 手机号脱敏: 138****5678
            if re.match(r"^\d{11}$", value):
                return value[:3] + mask_char * 4 + value[7:]
            return value

        elif mask_type == "email":
            # 邮箱脱敏: abc***@example.com
            if "@" in value:
                local, domain = value.split("@", 1)
                if len(local) > 3:
                    masked_local = local[:3] + mask_char * 3
                else:
                    masked_local = mask_char * len(local)
                return f"{masked_local}@{domain}"
            return value

        elif mask_type == "id_card":
            # 身份证脱敏: 110***********1234
            if re.match(r"^\d{15}$|^\d{18}$", value):
                return value[:3] + mask_char * (len(value) - 7) + value[-4:]
            return value

        elif mask_type == "credit_card":
            # 信用卡脱敏: 6222 **** **** 1234
            cleaned = re.sub(r"[\s-]", "", value)
            if re.match(r"^\d{16}$", cleaned):
                masked = cleaned[:4] + mask_char * 8 + cleaned[-4:]
                # 格式化为 XXXX **** **** XXXX
                return f"{masked[:4]} {masked[4:8]} {masked[8:12]} {masked[12:]}"
            return value

        elif mask_type == "name":
            # 姓名脱敏: 张*、李**
            if len(value) == 2:
                return value[0] + mask_char
            elif len(value) > 2:
                return value[0] + mask_char * (len(value) - 1)
            return value

        elif mask_type == "partial" or mask_type == "custom":
            # 部分脱敏或自定义脱敏
            if keep_first + keep_last >= len(value):
                return value

            if keep_first == 0 and keep_last == 0:
                return mask_char * len(value)

            middle_len = len(value) - keep_first - keep_last
            return value[:keep_first] + mask_char * middle_len + value[-keep_last:] if keep_last > 0 else value[:keep_first] + mask_char * middle_len

        return value

    def detect_sensitive_data(
        self,
        filename: str,
        sheet_name: str,
        cell_range: Optional[str] = None,
    ) -> dict[str, Any]:
        """检测敏感数据.

        Args:
            filename: 文件名
            sheet_name: 工作表名称
            cell_range: 单元格范围 (可选，默认检测整表)

        Returns:
            检测结果
        """
        try:
            file_path = config.paths.output_dir / filename
            self.file_manager.validate_file_path(file_path, must_exist=True)

            wb = load_workbook(str(file_path))

            if sheet_name not in wb.sheetnames:
                raise ValueError(f"工作表 '{sheet_name}' 不存在")

            ws = wb[sheet_name]

            sensitive_data = {
                "phone": [],
                "email": [],
                "id_card": [],
                "credit_card": [],
            }

            # 确定检测范围
            if cell_range:
                cells = ws[cell_range]
            else:
                cells = ws.iter_rows()

            # 检测敏感数据
            for row in cells:
                for cell in row if isinstance(row, tuple) else [row]:
                    if cell.value is None or not isinstance(cell.value, str):
                        continue

                    value = str(cell.value)
                    cell_coord = cell.coordinate

                    # 检测手机号
                    if re.match(r"^\d{11}$", value):
                        sensitive_data["phone"].append(
                            {"cell": cell_coord, "value": value}
                        )

                    # 检测邮箱
                    elif re.match(r"^[\w\.-]+@[\w\.-]+\.\w+$", value):
                        sensitive_data["email"].append(
                            {"cell": cell_coord, "value": value}
                        )

                    # 检测身份证
                    elif re.match(r"^\d{15}$|^\d{18}$", value):
                        sensitive_data["id_card"].append(
                            {"cell": cell_coord, "value": value}
                        )

                    # 检测信用卡
                    elif re.match(r"^\d{16}$", re.sub(r"[\s-]", "", value)):
                        sensitive_data["credit_card"].append(
                            {"cell": cell_coord, "value": value}
                        )

            wb.close()

            total_count = sum(len(items) for items in sensitive_data.values())

            logger.info(f"敏感数据检测完成: {file_path}, 发现 {total_count} 个敏感数据")
            return {
                "success": True,
                "message": f"发现 {total_count} 个敏感数据",
                "filename": str(file_path),
                "sheet_name": sheet_name,
                "sensitive_data": sensitive_data,
                "total_count": total_count,
            }

        except Exception as e:
            logger.error(f"敏感数据检测失败: {e}")
            return {"success": False, "message": f"检测失败: {str(e)}"}

    def hash_data(
        self,
        filename: str,
        sheet_name: str,
        cell_range: str,
        algorithm: Literal["md5", "sha256"] = "sha256",
    ) -> dict[str, Any]:
        """哈希加密数据.

        Args:
            filename: 文件名
            sheet_name: 工作表名称
            cell_range: 单元格范围
            algorithm: 哈希算法 ('md5' 或 'sha256')

        Returns:
            操作结果
        """
        try:
            file_path = config.paths.output_dir / filename
            self.file_manager.validate_file_path(file_path, must_exist=True)

            wb = load_workbook(str(file_path))

            if sheet_name not in wb.sheetnames:
                raise ValueError(f"工作表 '{sheet_name}' 不存在")

            ws = wb[sheet_name]

            hashed_count = 0

            # 遍历范围内的单元格
            for row in ws[cell_range]:
                for cell in row if isinstance(row, tuple) else [row]:
                    if cell.value is None:
                        continue

                    value_str = str(cell.value)

                    if algorithm == "md5":
                        hashed = hashlib.md5(value_str.encode()).hexdigest()
                    elif algorithm == "sha256":
                        hashed = hashlib.sha256(value_str.encode()).hexdigest()
                    else:
                        raise ValueError(f"不支持的哈希算法: {algorithm}")

                    cell.value = hashed
                    hashed_count += 1

            wb.save(str(file_path))
            wb.close()

            logger.info(f"数据哈希加密完成: {file_path}, 处理 {hashed_count} 个单元格")
            return {
                "success": True,
                "message": f"成功加密 {hashed_count} 个单元格",
                "filename": str(file_path),
                "sheet_name": sheet_name,
                "cell_range": cell_range,
                "hashed_count": hashed_count,
                "algorithm": algorithm,
            }

        except Exception as e:
            logger.error(f"数据哈希加密失败: {e}")
            return {"success": False, "message": f"加密失败: {str(e)}"}
