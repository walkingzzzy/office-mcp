"""Excel 格式化操作模块."""

from typing import Any, Optional, Union

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import (
    ColorScaleRule,
    DataBarRule,
    IconSetRule,
    CellIsRule,
    Rule,
)
from loguru import logger

from office_mcp_server.config import config
from office_mcp_server.utils.file_manager import FileManager
from office_mcp_server.utils.format_helper import ColorUtils


class ExcelFormatOperations:
    """Excel 格式化操作类."""

    def __init__(self) -> None:
        """初始化格式化操作类."""
        self.file_manager = FileManager()

    def format_cell(
        self,
        filename: str,
        sheet_name: str,
        cell: str,
        font_name: Optional[str] = None,
        font_size: Optional[int] = None,
        bold: bool = False,
        color: Optional[str] = None,
        bg_color: Optional[str] = None,
        number_format: Optional[str] = None,
        horizontal_alignment: Optional[str] = None,
        vertical_alignment: Optional[str] = None,
        wrap_text: bool = False,
        border_style: Optional[str] = None,
        border_color: Optional[str] = None,
    ) -> dict[str, Any]:
        """格式化单元格."""
        try:
            file_path = config.paths.output_dir / filename
            self.file_manager.validate_file_path(file_path, must_exist=True)

            wb = load_workbook(str(file_path))

            if sheet_name not in wb.sheetnames:
                raise ValueError(f"工作表 '{sheet_name}' 不存在")

            ws = wb[sheet_name]
            target_cell = ws[cell]

            # 应用字体样式
            font_kwargs = {}
            if font_name:
                font_kwargs["name"] = font_name
            if font_size:
                font_kwargs["size"] = font_size
            if bold:
                font_kwargs["bold"] = True
            if color:
                font_kwargs["color"] = color.lstrip("#")

            if font_kwargs:
                target_cell.font = Font(**font_kwargs)

            # 应用背景色
            if bg_color:
                target_cell.fill = PatternFill(
                    start_color=bg_color.lstrip("#"),
                    end_color=bg_color.lstrip("#"),
                    fill_type="solid",
                )

            # 应用数字格式
            if number_format:
                target_cell.number_format = number_format

            # 应用对齐方式
            alignment_kwargs = {}
            if horizontal_alignment:
                alignment_kwargs["horizontal"] = horizontal_alignment
            if vertical_alignment:
                alignment_kwargs["vertical"] = vertical_alignment
            if wrap_text:
                alignment_kwargs["wrap_text"] = True

            if alignment_kwargs:
                target_cell.alignment = Alignment(**alignment_kwargs)

            # 应用边框
            if border_style:
                border_color_hex = border_color.lstrip("#") if border_color else "000000"
                side = Side(style=border_style, color=border_color_hex)
                target_cell.border = Border(left=side, right=side, top=side, bottom=side)

            wb.save(str(file_path))
            wb.close()

            logger.info(f"单元格 {cell} 格式化成功: {file_path}")
            return {
                "success": True,
                "message": "单元格格式化成功",
                "filename": str(file_path),
            }

        except Exception as e:
            logger.error(f"格式化单元格失败: {e}")
            return {"success": False, "message": f"格式化失败: {str(e)}"}

    def apply_conditional_formatting(
        self,
        filename: str,
        sheet_name: str,
        cell_range: str,
        rule_type: str,
        format_type: str = "fill",
        color: Optional[str] = None,
        operator: Optional[str] = None,
        formula: Optional[str] = None,
        value1: Optional[Union[str, int, float]] = None,
        value2: Optional[Union[str, int, float]] = None,
    ) -> dict[str, Any]:
        """应用条件格式."""
        try:
            file_path = config.paths.output_dir / filename
            self.file_manager.validate_file_path(file_path, must_exist=True)

            wb = load_workbook(str(file_path))

            if sheet_name not in wb.sheetnames:
                raise ValueError(f"工作表 '{sheet_name}' 不存在")

            ws = wb[sheet_name]

            # 根据规则类型创建条件格式规则
            if rule_type == 'colorScale':
                # 色阶规则
                rule = ColorScaleRule(
                    start_type='min',
                    start_color='63BE7B',
                    mid_type='percentile',
                    mid_value=50,
                    mid_color='FFEB84',
                    end_type='max',
                    end_color='F8696B'
                )

            elif rule_type == 'dataBar':
                # 数据条规则
                rule = DataBarRule(
                    start_type='min',
                    start_value=0,
                    end_type='max',
                    end_value=100,
                    color="638EC6"
                )

            elif rule_type == 'iconSet':
                # 图标集规则
                rule = IconSetRule(
                    icon_style='3TrafficLights1',
                    type='percent',
                    values=[0, 33, 67]
                )

            elif rule_type == 'cellIs':
                # 单元格值规则
                if not operator or value1 is None:
                    raise ValueError("cellIs规则需要指定operator和value1")

                fill_color = color.lstrip("#") if color else "FF0000"

                if operator == 'between':
                    if value2 is None:
                        raise ValueError("between操作符需要指定value2")
                    formula_parts = [value1, value2]
                elif operator in ['greaterThan', 'lessThan', 'equal', 'notEqual',
                                'greaterThanOrEqual', 'lessThanOrEqual']:
                    formula_parts = [value1]
                else:
                    raise ValueError(f"不支持的操作符: {operator}")

                if format_type == "fill":
                    fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
                    rule = CellIsRule(operator=operator, formula=formula_parts, fill=fill)
                elif format_type == "font":
                    font = Font(color=fill_color, bold=True)
                    rule = CellIsRule(operator=operator, formula=formula_parts, font=font)
                else:
                    raise ValueError(f"不支持的格式类型: {format_type}")

            elif rule_type == 'aboveAverage':
                # 高于平均值规则
                fill_color = color.lstrip("#") if color else "FF0000"
                fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
                rule = Rule(type='aboveAverage', dxf=fill)

            elif rule_type == 'duplicateValues':
                # 重复值规则
                fill_color = color.lstrip("#") if color else "FF0000"
                fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
                rule = Rule(type='duplicateValues', dxf=fill)

            elif rule_type == 'uniqueValues':
                # 唯一值规则
                fill_color = color.lstrip("#") if color else "00FF00"
                fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
                rule = Rule(type='uniqueValues', dxf=fill)

            elif rule_type == 'expression':
                # 公式规则
                if not formula:
                    raise ValueError("expression规则需要指定formula")
                fill_color = color.lstrip("#") if color else "FF0000"
                fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
                rule = Rule(type='expression', formula=[formula], dxf=fill)

            else:
                raise ValueError(f"不支持的规则类型: {rule_type}")

            # 应用条件格式到指定范围
            ws.conditional_formatting.add(cell_range, rule)

            wb.save(str(file_path))
            wb.close()

            logger.info(f"条件格式应用成功: {file_path}")
            return {
                "success": True,
                "message": f"条件格式 '{rule_type}' 已应用到 {cell_range}",
                "filename": str(file_path),
                "rule_type": rule_type,
                "cell_range": cell_range,
            }

        except Exception as e:
            logger.error(f"应用条件格式失败: {e}")
            return {"success": False, "message": f"应用失败: {str(e)}"}
