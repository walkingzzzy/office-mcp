"""Excel 公式操作模块."""

from typing import Any, Optional

from openpyxl import load_workbook
from loguru import logger

from office_mcp_server.config import config
from office_mcp_server.utils.file_manager import FileManager


class ExcelFormulaOperations:
    """Excel 公式操作类."""

    def __init__(self) -> None:
        """初始化公式操作类."""
        self.file_manager = FileManager()

    def insert_formula(
        self,
        filename: str,
        sheet_name: str,
        cell: str,
        formula: str,
    ) -> dict[str, Any]:
        """插入Excel公式."""
        try:
            file_path = config.paths.output_dir / filename
            self.file_manager.validate_file_path(file_path, must_exist=True)

            wb = load_workbook(str(file_path))

            if sheet_name not in wb.sheetnames:
                raise ValueError(f"工作表 '{sheet_name}' 不存在")

            ws = wb[sheet_name]

            # 插入公式
            if not formula.startswith('='):
                formula = '=' + formula

            ws[cell] = formula

            wb.save(str(file_path))
            wb.close()

            logger.info(f"公式插入成功: {file_path}")
            return {
                "success": True,
                "message": "公式插入成功",
                "filename": str(file_path),
                "cell": cell,
                "formula": formula,
            }

        except Exception as e:
            logger.error(f"插入公式失败: {e}")
            return {"success": False, "message": f"插入失败: {str(e)}"}

    def apply_function(
        self,
        filename: str,
        sheet_name: str,
        cell: str,
        function_name: str,
        range1: Optional[str] = None,
        range2: Optional[str] = None,
        condition: Optional[str] = None,
        value_if_true: Optional[str] = None,
        value_if_false: Optional[str] = None,
        lookup_value: Optional[str] = None,
        table_array: Optional[str] = None,
        col_index: Optional[int] = None,
        range_lookup: bool = False,
    ) -> dict[str, Any]:
        """应用常用Excel函数."""
        try:
            file_path = config.paths.output_dir / filename
            self.file_manager.validate_file_path(file_path, must_exist=True)

            wb = load_workbook(str(file_path))

            if sheet_name not in wb.sheetnames:
                raise ValueError(f"工作表 '{sheet_name}' 不存在")

            ws = wb[sheet_name]

            # 根据函数名称构建公式
            function_name = function_name.upper()

            if function_name in ['SUM', 'AVERAGE', 'MAX', 'MIN', 'COUNT', 'COUNTA']:
                # 基础聚合函数
                if not range1:
                    raise ValueError(f"{function_name}函数需要指定range1")
                formula = f"={function_name}({range1})"

            elif function_name == 'IF':
                # IF条件函数
                if not condition or not value_if_true or not value_if_false:
                    raise ValueError("IF函数需要指定condition, value_if_true, value_if_false")
                formula = f"=IF({condition},{value_if_true},{value_if_false})"

            elif function_name == 'SUMIF':
                # 条件求和
                if not range1 or not condition:
                    raise ValueError("SUMIF函数需要指定range1和condition")
                sum_range = range2 if range2 else range1
                formula = f"=SUMIF({range1},{condition},{sum_range})"

            elif function_name == 'COUNTIF':
                # 条件计数
                if not range1 or not condition:
                    raise ValueError("COUNTIF函数需要指定range1和condition")
                formula = f"=COUNTIF({range1},{condition})"

            elif function_name == 'VLOOKUP':
                # 纵向查找
                if not lookup_value or not table_array or col_index is None:
                    raise ValueError("VLOOKUP函数需要指定lookup_value, table_array, col_index")
                formula = f"=VLOOKUP({lookup_value},{table_array},{col_index},{1 if range_lookup else 0})"

            elif function_name == 'HLOOKUP':
                # 横向查找
                if not lookup_value or not table_array or col_index is None:
                    raise ValueError("HLOOKUP函数需要指定lookup_value, table_array, col_index")
                formula = f"=HLOOKUP({lookup_value},{table_array},{col_index},{1 if range_lookup else 0})"

            elif function_name == 'CONCATENATE':
                # 文本连接
                if not range1:
                    raise ValueError("CONCATENATE函数需要指定range1")
                # range1应该是逗号分隔的单元格引用
                formula = f"=CONCATENATE({range1})"

            elif function_name in ['LEFT', 'RIGHT']:
                # 左侧/右侧文本
                if not range1:
                    raise ValueError(f"{function_name}函数需要指定range1")
                num_chars = condition if condition else "1"
                formula = f"={function_name}({range1},{num_chars})"

            elif function_name == 'MID':
                # 中间文本
                if not range1 or not condition or not value_if_true:
                    raise ValueError("MID函数需要指定range1(文本), condition(起始位置), value_if_true(字符数)")
                formula = f"=MID({range1},{condition},{value_if_true})"

            elif function_name in ['LEN', 'UPPER', 'LOWER', 'TRIM']:
                # 文本处理函数
                if not range1:
                    raise ValueError(f"{function_name}函数需要指定range1")
                formula = f"={function_name}({range1})"

            else:
                raise ValueError(f"不支持的函数: {function_name}")

            # 插入公式
            ws[cell] = formula

            wb.save(str(file_path))
            wb.close()

            logger.info(f"函数应用成功: {file_path}")
            return {
                "success": True,
                "message": f"函数 '{function_name}' 已应用到 {cell}",
                "filename": str(file_path),
                "cell": cell,
                "formula": formula,
                "function": function_name,
            }

        except Exception as e:
            logger.error(f"应用函数失败: {e}")
            return {"success": False, "message": f"应用失败: {str(e)}"}
