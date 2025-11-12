"""Excel 数据导入导出模块."""

import csv
import json
from typing import Any, Optional
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from loguru import logger

from office_mcp_server.config import config
from office_mcp_server.utils.file_manager import FileManager


class ExcelImportExportOperations:
    """Excel 数据导入导出操作类."""

    def __init__(self) -> None:
        """初始化导入导出操作类."""
        self.file_manager = FileManager()

    def import_from_csv(
        self,
        filename: str,
        sheet_name: str,
        csv_file: str,
        start_cell: str = "A1",
        has_header: bool = True,
    ) -> dict[str, Any]:
        """从CSV文件导入数据.

        Args:
            filename: Excel文件名
            sheet_name: 工作表名称
            csv_file: CSV文件路径
            start_cell: 起始单元格 (默认 'A1')
            has_header: 是否包含表头 (默认 True)
        """
        try:
            file_path = config.paths.output_dir / filename
            self.file_manager.validate_file_path(file_path, must_exist=True)

            csv_path = Path(csv_file)
            if not csv_path.is_absolute():
                csv_path = config.paths.output_dir / csv_file

            if not csv_path.exists():
                raise FileNotFoundError(f"CSV文件不存在: {csv_path}")

            wb = load_workbook(str(file_path))

            if sheet_name not in wb.sheetnames:
                raise ValueError(f"工作表 '{sheet_name}' 不存在")

            ws = wb[sheet_name]

            from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
            col_letter, row = coordinate_from_string(start_cell)
            start_col = column_index_from_string(col_letter)

            with open(csv_path, 'r', encoding='utf-8-sig') as f:
                csv_reader = csv.reader(f)
                row_count = 0
                for csv_row in csv_reader:
                    for col_idx, value in enumerate(csv_row):
                        ws.cell(row=row + row_count, column=start_col + col_idx, value=value)
                    row_count += 1

            wb.save(str(file_path))
            wb.close()

            logger.info(f"从CSV导入成功: {csv_path} -> {file_path}")
            return {
                "success": True,
                "message": f"成功从CSV导入 {row_count} 行数据",
                "filename": str(file_path),
                "csv_file": str(csv_path),
                "rows": row_count,
                "start_cell": start_cell,
            }

        except Exception as e:
            logger.error(f"从CSV导入失败: {e}")
            return {"success": False, "message": f"导入失败: {str(e)}"}

    def import_from_json(
        self,
        filename: str,
        sheet_name: str,
        json_file: str,
        start_cell: str = "A1",
        json_path: Optional[str] = None,
    ) -> dict[str, Any]:
        """从JSON文件导入数据.

        Args:
            filename: Excel文件名
            sheet_name: 工作表名称
            json_file: JSON文件路径
            start_cell: 起始单元格 (默认 'A1')
            json_path: JSON数据路径 (用于提取嵌套数据, 可选)
        """
        try:
            file_path = config.paths.output_dir / filename
            self.file_manager.validate_file_path(file_path, must_exist=True)

            json_file_path = Path(json_file)
            if not json_file_path.is_absolute():
                json_file_path = config.paths.output_dir / json_file

            if not json_file_path.exists():
                raise FileNotFoundError(f"JSON文件不存在: {json_file_path}")

            wb = load_workbook(str(file_path))

            if sheet_name not in wb.sheetnames:
                raise ValueError(f"工作表 '{sheet_name}' 不存在")

            ws = wb[sheet_name]

            from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
            col_letter, row = coordinate_from_string(start_cell)
            start_col = column_index_from_string(col_letter)

            with open(json_file_path, 'r', encoding='utf-8') as f:
                data = json.load(f)

            if json_path:
                for path_part in json_path.split('.'):
                    data = data[path_part]

            if isinstance(data, list):
                if data and isinstance(data[0], dict):
                    headers = list(data[0].keys())
                    for col_idx, header in enumerate(headers):
                        ws.cell(row=row, column=start_col + col_idx, value=header)

                    for row_idx, item in enumerate(data):
                        for col_idx, header in enumerate(headers):
                            value = item.get(header, '')
                            ws.cell(row=row + row_idx + 1, column=start_col + col_idx, value=value)

                    row_count = len(data) + 1
                else:
                    for row_idx, value in enumerate(data):
                        ws.cell(row=row + row_idx, column=start_col, value=value)
                    row_count = len(data)
            else:
                raise ValueError("JSON数据必须是数组格式")

            wb.save(str(file_path))
            wb.close()

            logger.info(f"从JSON导入成功: {json_file_path} -> {file_path}")
            return {
                "success": True,
                "message": f"成功从JSON导入 {row_count} 行数据",
                "filename": str(file_path),
                "json_file": str(json_file_path),
                "rows": row_count,
                "start_cell": start_cell,
            }

        except Exception as e:
            logger.error(f"从JSON导入失败: {e}")
            return {"success": False, "message": f"导入失败: {str(e)}"}

    def export_to_csv(
        self,
        filename: str,
        sheet_name: str,
        csv_file: str,
        cell_range: Optional[str] = None,
    ) -> dict[str, Any]:
        """导出为CSV文件.

        Args:
            filename: Excel文件名
            sheet_name: 工作表名称
            csv_file: 输出CSV文件路径
            cell_range: 导出范围 (可选, 如 'A1:D10')
        """
        try:
            file_path = config.paths.output_dir / filename
            self.file_manager.validate_file_path(file_path, must_exist=True)

            csv_path = Path(csv_file)
            if not csv_path.is_absolute():
                csv_path = config.paths.output_dir / csv_file

            self.file_manager.ensure_directory(csv_path.parent)

            wb = load_workbook(str(file_path))

            if sheet_name not in wb.sheetnames:
                raise ValueError(f"工作表 '{sheet_name}' 不存在")

            ws = wb[sheet_name]

            with open(csv_path, 'w', newline='', encoding='utf-8-sig') as f:
                csv_writer = csv.writer(f)

                if cell_range:
                    data_cells = ws[cell_range]
                    for row in data_cells:
                        if isinstance(row, tuple):
                            row_data = [cell.value if cell.value is not None else '' for cell in row]
                        else:
                            row_data = [row.value if row.value is not None else '']
                        csv_writer.writerow(row_data)
                else:
                    for row in ws.iter_rows(values_only=True):
                        csv_writer.writerow([cell if cell is not None else '' for cell in row])

            logger.info(f"导出为CSV成功: {file_path} -> {csv_path}")
            return {
                "success": True,
                "message": f"成功导出为CSV文件",
                "filename": str(file_path),
                "csv_file": str(csv_path),
                "cell_range": cell_range or "全部数据",
            }

        except Exception as e:
            logger.error(f"导出为CSV失败: {e}")
            return {"success": False, "message": f"导出失败: {str(e)}"}

    def export_to_json(
        self,
        filename: str,
        sheet_name: str,
        json_file: str,
        cell_range: Optional[str] = None,
        has_header: bool = True,
        orient: str = "records",
    ) -> dict[str, Any]:
        """导出为JSON文件.

        Args:
            filename: Excel文件名
            sheet_name: 工作表名称
            json_file: 输出JSON文件路径
            cell_range: 导出范围 (可选, 如 'A1:D10')
            has_header: 是否包含表头 (默认 True)
            orient: JSON格式 ('records'记录数组, 'columns'列字典, 'values'值数组)
        """
        try:
            file_path = config.paths.output_dir / filename
            self.file_manager.validate_file_path(file_path, must_exist=True)

            json_path = Path(json_file)
            if not json_path.is_absolute():
                json_path = config.paths.output_dir / json_file

            self.file_manager.ensure_directory(json_path.parent)

            wb = load_workbook(str(file_path))

            if sheet_name not in wb.sheetnames:
                raise ValueError(f"工作表 '{sheet_name}' 不存在")

            ws = wb[sheet_name]

            if cell_range:
                data_cells = ws[cell_range]
                all_data = []
                for row in data_cells:
                    if isinstance(row, tuple):
                        row_data = [cell.value for cell in row]
                    else:
                        row_data = [row.value]
                    all_data.append(row_data)
            else:
                all_data = []
                for row in ws.iter_rows(values_only=True):
                    all_data.append(list(row))

            if not all_data:
                raise ValueError("没有数据可以导出")

            if has_header and orient == "records":
                headers = all_data[0]
                data_rows = all_data[1:]
                result = [dict(zip(headers, row)) for row in data_rows]
            elif orient == "columns":
                if has_header:
                    headers = all_data[0]
                    data_rows = all_data[1:]
                    result = {header: [row[i] for row in data_rows]
                             for i, header in enumerate(headers)}
                else:
                    result = {f"col{i}": [row[i] for row in all_data]
                             for i in range(len(all_data[0]))}
            else:
                result = all_data

            with open(json_path, 'w', encoding='utf-8') as f:
                json.dump(result, f, ensure_ascii=False, indent=2, default=str)

            logger.info(f"导出为JSON成功: {file_path} -> {json_path}")
            return {
                "success": True,
                "message": f"成功导出为JSON文件",
                "filename": str(file_path),
                "json_file": str(json_path),
                "cell_range": cell_range or "全部数据",
                "orient": orient,
            }

        except Exception as e:
            logger.error(f"导出为JSON失败: {e}")
            return {"success": False, "message": f"导出失败: {str(e)}"}

    def export_to_html(
        self,
        filename: str,
        sheet_name: str,
        html_file: str,
        cell_range: Optional[str] = None,
        include_style: bool = True,
    ) -> dict[str, Any]:
        """导出为HTML文件.

        Args:
            filename: Excel文件名
            sheet_name: 工作表名称
            html_file: 输出HTML文件路径
            cell_range: 导出范围 (可选, 如 'A1:D10')
            include_style: 是否包含样式 (默认 True)
        """
        try:
            file_path = config.paths.output_dir / filename
            self.file_manager.validate_file_path(file_path, must_exist=True)

            html_path = Path(html_file)
            if not html_path.is_absolute():
                html_path = config.paths.output_dir / html_file

            self.file_manager.ensure_directory(html_path.parent)

            wb = load_workbook(str(file_path))

            if sheet_name not in wb.sheetnames:
                raise ValueError(f"工作表 '{sheet_name}' 不存在")

            ws = wb[sheet_name]

            html_content = ['<!DOCTYPE html>']
            html_content.append('<html>')
            html_content.append('<head>')
            html_content.append('<meta charset="UTF-8">')
            html_content.append(f'<title>{sheet_name}</title>')

            if include_style:
                html_content.append('<style>')
                html_content.append('table { border-collapse: collapse; width: 100%; }')
                html_content.append('th, td { border: 1px solid #ddd; padding: 8px; text-align: left; }')
                html_content.append('th { background-color: #4CAF50; color: white; }')
                html_content.append('tr:nth-child(even) { background-color: #f2f2f2; }')
                html_content.append('</style>')

            html_content.append('</head>')
            html_content.append('<body>')
            html_content.append('<table>')

            if cell_range:
                data_cells = ws[cell_range]
                for row_idx, row in enumerate(data_cells):
                    html_content.append('<tr>')
                    if isinstance(row, tuple):
                        for cell in row:
                            tag = 'th' if row_idx == 0 else 'td'
                            value = cell.value if cell.value is not None else ''
                            html_content.append(f'<{tag}>{value}</{tag}>')
                    else:
                        tag = 'th' if row_idx == 0 else 'td'
                        value = row.value if row.value is not None else ''
                        html_content.append(f'<{tag}>{value}</{tag}>')
                    html_content.append('</tr>')
            else:
                for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
                    html_content.append('<tr>')
                    for cell in row:
                        tag = 'th' if row_idx == 0 else 'td'
                        value = cell if cell is not None else ''
                        html_content.append(f'<{tag}>{value}</{tag}>')
                    html_content.append('</tr>')

            html_content.append('</table>')
            html_content.append('</body>')
            html_content.append('</html>')

            with open(html_path, 'w', encoding='utf-8') as f:
                f.write('\n'.join(html_content))

            logger.info(f"导出为HTML成功: {file_path} -> {html_path}")
            return {
                "success": True,
                "message": f"成功导出为HTML文件",
                "filename": str(file_path),
                "html_file": str(html_path),
                "cell_range": cell_range or "全部数据",
            }

        except Exception as e:
            logger.error(f"导出为HTML失败: {e}")
            return {"success": False, "message": f"导出失败: {str(e)}"}

    def export_to_pdf(
        self,
        filename: str,
        sheet_name: str,
        pdf_file: str,
        cell_range: Optional[str] = None,
    ) -> dict[str, Any]:
        """导出为PDF文件.

        Args:
            filename: Excel文件名
            sheet_name: 工作表名称
            pdf_file: 输出PDF文件路径
            cell_range: 导出范围 (可选, 如 'A1:D10')

        Note:
            此功能需要安装 reportlab 库: pip install reportlab
            或使用简化的HTML转PDF方案
        """
        try:
            file_path = config.paths.output_dir / filename
            self.file_manager.validate_file_path(file_path, must_exist=True)

            pdf_path = Path(pdf_file)
            if not pdf_path.is_absolute():
                pdf_path = config.paths.output_dir / pdf_file

            self.file_manager.ensure_directory(pdf_path.parent)

            try:
                from reportlab.lib import colors
                from reportlab.lib.pagesizes import letter, A4
                from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
                from reportlab.lib.units import inch

                wb = load_workbook(str(file_path))

                if sheet_name not in wb.sheetnames:
                    raise ValueError(f"工作表 '{sheet_name}' 不存在")

                ws = wb[sheet_name]

                data = []
                if cell_range:
                    data_cells = ws[cell_range]
                    for row in data_cells:
                        if isinstance(row, tuple):
                            row_data = [str(cell.value) if cell.value is not None else '' for cell in row]
                        else:
                            row_data = [str(row.value) if row.value is not None else '']
                        data.append(row_data)
                else:
                    for row in ws.iter_rows(values_only=True):
                        data.append([str(cell) if cell is not None else '' for cell in row])

                if not data:
                    raise ValueError("没有数据可以导出")

                doc = SimpleDocTemplate(str(pdf_path), pagesize=A4)
                elements = []

                table = Table(data)
                table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 12),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black)
                ]))

                elements.append(table)
                doc.build(elements)

                logger.info(f"导出为PDF成功: {file_path} -> {pdf_path}")
                return {
                    "success": True,
                    "message": f"成功导出为PDF文件",
                    "filename": str(file_path),
                    "pdf_file": str(pdf_path),
                    "cell_range": cell_range or "全部数据",
                }

            except ImportError:
                logger.warning("未安装reportlab库,使用备用方案导出PDF")
                html_temp = pdf_path.with_suffix('.html')
                self.export_to_html(filename, sheet_name, str(html_temp), cell_range, True)

                return {
                    "success": True,
                    "message": f"PDF导出需要reportlab库。已导出为HTML文件作为备用: {html_temp}",
                    "filename": str(file_path),
                    "html_file": str(html_temp),
                    "note": "请安装reportlab库以支持PDF导出: pip install reportlab",
                }

        except Exception as e:
            logger.error(f"导出为PDF失败: {e}")
            return {"success": False, "message": f"导出失败: {str(e)}"}
