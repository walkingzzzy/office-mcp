"""Excel 单元格合并操作模块."""

from typing import Any

from openpyxl import load_workbook
from loguru import logger

from office_mcp_server.config import config
from office_mcp_server.utils.file_manager import FileManager


class ExcelMergeOperations:
    """Excel 单元格合并操作类."""

    def __init__(self) -> None:
        """初始化合并操作类."""
        self.file_manager = FileManager()

    def merge_cells(
        self,
        filename: str,
        sheet_name: str,
        cell_range: str,
    ) -> dict[str, Any]:
        """合并单元格.

        Args:
            filename: 文件名
            sheet_name: 工作表名称
            cell_range: 单元格范围 (如 'A1:C3')
        """
        try:
            file_path = config.paths.output_dir / filename
            self.file_manager.validate_file_path(file_path, must_exist=True)

            wb = load_workbook(str(file_path))

            if sheet_name not in wb.sheetnames:
                raise ValueError(f"工作表 '{sheet_name}' 不存在")

            ws = wb[sheet_name]
            ws.merge_cells(cell_range)

            wb.save(str(file_path))
            wb.close()

            logger.info(f"合并单元格成功: {file_path}, 范围: {cell_range}")
            return {
                "success": True,
                "message": f"成功合并单元格范围 {cell_range}",
                "filename": str(file_path),
                "cell_range": cell_range,
            }

        except Exception as e:
            logger.error(f"合并单元格失败: {e}")
            return {"success": False, "message": f"合并单元格失败: {str(e)}"}

    def unmerge_cells(
        self,
        filename: str,
        sheet_name: str,
        cell_range: str,
    ) -> dict[str, Any]:
        """取消合并单元格.

        Args:
            filename: 文件名
            sheet_name: 工作表名称
            cell_range: 单元格范围 (如 'A1:C3')
        """
        try:
            file_path = config.paths.output_dir / filename
            self.file_manager.validate_file_path(file_path, must_exist=True)

            wb = load_workbook(str(file_path))

            if sheet_name not in wb.sheetnames:
                raise ValueError(f"工作表 '{sheet_name}' 不存在")

            ws = wb[sheet_name]
            ws.unmerge_cells(cell_range)

            wb.save(str(file_path))
            wb.close()

            logger.info(f"取消合并单元格成功: {file_path}, 范围: {cell_range}")
            return {
                "success": True,
                "message": f"成功取消合并单元格范围 {cell_range}",
                "filename": str(file_path),
                "cell_range": cell_range,
            }

        except Exception as e:
            logger.error(f"取消合并单元格失败: {e}")
            return {"success": False, "message": f"取消合并单元格失败: {str(e)}"}
