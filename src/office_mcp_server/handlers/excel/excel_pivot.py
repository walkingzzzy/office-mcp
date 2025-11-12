"""Excel 数据透视表模块."""

from typing import Any, Optional

from openpyxl import load_workbook
from loguru import logger

from office_mcp_server.config import config
from office_mcp_server.utils.file_manager import FileManager


class ExcelPivotOperations:
    """Excel 数据透视表操作类."""

    def __init__(self) -> None:
        """初始化数据透视表操作类."""
        self.file_manager = FileManager()

    def create_pivot_table(
        self,
        filename: str,
        source_sheet: str,
        source_range: str,
        pivot_sheet: str,
        pivot_location: str,
        row_fields: list[str],
        col_fields: Optional[list[str]] = None,
        data_fields: Optional[list[dict[str, str]]] = None,
        filter_fields: Optional[list[str]] = None,
    ) -> dict[str, Any]:
        """创建数据透视表.

        Args:
            filename: Excel文件名
            source_sheet: 源数据工作表名称
            source_range: 源数据范围（如 "A1:E100"）
            pivot_sheet: 透视表工作表名称（如果不存在则创建）
            pivot_location: 透视表位置（如 "A3"）
            row_fields: 行字段列表（字段名称）
            col_fields: 列字段列表（可选）
            data_fields: 数据字段列表，每个元素为 {"field": "字段名", "function": "sum/average/count/max/min"}
            filter_fields: 筛选字段列表（可选）
        """
        try:
            file_path = config.paths.output_dir / filename
            self.file_manager.validate_file_path(file_path, must_exist=True)

            # 注意：openpyxl对数据透视表的支持非常有限
            # 完整的数据透视表创建需要使用win32com或xlwings库与Excel应用程序交互
            # 这里我们提供一个基础实现，生成数据透视表的XML结构

            wb = load_workbook(str(file_path))

            if source_sheet not in wb.sheetnames:
                raise ValueError(f"源工作表 '{source_sheet}' 不存在")

            # 创建或获取透视表工作表
            if pivot_sheet not in wb.sheetnames:
                wb.create_sheet(pivot_sheet)

            pivot_ws = wb[pivot_sheet]

            # 由于openpyxl对数据透视表支持有限，我们创建一个简化的统计摘要
            # 实际的数据透视表需要使用xlwings或win32com

            # 尝试使用win32com（仅在Windows上可用）
            try:
                import win32com.client as win32

                # 关闭openpyxl的工作簿，让Excel应用程序接管
                wb.save(str(file_path))
                wb.close()

                # 使用Excel COM接口创建数据透视表
                excel = win32.gencache.EnsureDispatch('Excel.Application')
                excel.Visible = False
                excel.DisplayAlerts = False

                wb_com = excel.Workbooks.Open(str(file_path.absolute()))

                source_ws = wb_com.Worksheets(source_sheet)
                pivot_ws_com = wb_com.Worksheets(pivot_sheet)

                # 创建数据透视表缓存
                source_data = source_ws.Range(source_range)
                pivot_cache = wb_com.PivotCaches().Create(
                    SourceType=1,  # xlDatabase
                    SourceData=source_data
                )

                # 创建数据透视表
                pivot_table = pivot_cache.CreatePivotTable(
                    TableDestination=pivot_ws_com.Range(pivot_location),
                    TableName="PivotTable1"
                )

                # 添加行字段
                for field_name in row_fields:
                    pivot_field = pivot_table.PivotFields(field_name)
                    pivot_field.Orientation = 1  # xlRowField

                # 添加列字段
                if col_fields:
                    for field_name in col_fields:
                        pivot_field = pivot_table.PivotFields(field_name)
                        pivot_field.Orientation = 2  # xlColumnField

                # 添加数据字段
                if data_fields:
                    for data_field_info in data_fields:
                        field_name = data_field_info.get("field")
                        function = data_field_info.get("function", "sum")

                        pivot_field = pivot_table.PivotFields(field_name)
                        pivot_field.Orientation = 4  # xlDataField

                        # 设置聚合函数
                        function_map = {
                            "sum": -4157,      # xlSum
                            "count": -4112,    # xlCount
                            "average": -4106,  # xlAverage
                            "max": -4136,      # xlMax
                            "min": -4139,      # xlMin
                        }
                        data_field = pivot_table.DataFields(1)
                        data_field.Function = function_map.get(function, -4157)

                # 添加筛选字段
                if filter_fields:
                    for field_name in filter_fields:
                        pivot_field = pivot_table.PivotFields(field_name)
                        pivot_field.Orientation = 3  # xlPageField

                wb_com.Save()
                wb_com.Close(SaveChanges=True)
                excel.Quit()

                logger.info(f"数据透视表创建成功: {filename}")
                return {
                    "success": True,
                    "message": "数据透视表创建成功",
                    "filename": str(file_path),
                    "source_sheet": source_sheet,
                    "pivot_sheet": pivot_sheet,
                    "pivot_location": pivot_location,
                    "method": "win32com"
                }

            except ImportError:
                # 如果win32com不可用，返回提示信息
                wb.save(str(file_path))
                wb.close()

                return {
                    "success": False,
                    "message": "数据透视表功能需要 Windows 环境和 Microsoft Excel 应用程序，或安装 pywin32 库 (pip install pywin32)",
                    "alternative": "您可以使用 pandas 进行数据聚合分析作为替代方案",
                    "filename": str(file_path),
                }

        except Exception as e:
            logger.error(f"创建数据透视表失败: {e}")
            return {"success": False, "message": f"创建数据透视表失败: {str(e)}"}

    def refresh_pivot_table(
        self,
        filename: str,
        pivot_sheet: str,
        pivot_table_name: str = "PivotTable1",
    ) -> dict[str, Any]:
        """刷新数据透视表.

        Args:
            filename: Excel文件名
            pivot_sheet: 透视表工作表名称
            pivot_table_name: 透视表名称 (默认 "PivotTable1")

        Note:
            此功能需要 Windows 环境和 Microsoft Excel 应用程序，或安装 pywin32 库
        """
        try:
            file_path = config.paths.output_dir / filename
            self.file_manager.validate_file_path(file_path, must_exist=True)

            try:
                import win32com.client as win32

                # 使用Excel COM接口刷新数据透视表
                excel = win32.gencache.EnsureDispatch('Excel.Application')
                excel.Visible = False
                excel.DisplayAlerts = False

                wb_com = excel.Workbooks.Open(str(file_path.absolute()))

                if pivot_sheet not in [ws.Name for ws in wb_com.Worksheets]:
                    raise ValueError(f"透视表工作表 '{pivot_sheet}' 不存在")

                pivot_ws_com = wb_com.Worksheets(pivot_sheet)

                # 查找并刷新数据透视表
                refreshed = False
                for pivot_table in pivot_ws_com.PivotTables():
                    if pivot_table.Name == pivot_table_name:
                        pivot_table.RefreshTable()
                        refreshed = True
                        break

                if not refreshed:
                    wb_com.Close(SaveChanges=False)
                    excel.Quit()
                    raise ValueError(f"数据透视表 '{pivot_table_name}' 不存在")

                wb_com.Save()
                wb_com.Close(SaveChanges=True)
                excel.Quit()

                logger.info(f"数据透视表刷新成功: {filename}")
                return {
                    "success": True,
                    "message": f"数据透视表 '{pivot_table_name}' 刷新成功",
                    "filename": str(file_path),
                    "pivot_sheet": pivot_sheet,
                    "pivot_table_name": pivot_table_name,
                }

            except ImportError:
                return {
                    "success": False,
                    "message": "数据透视表刷新功能需要 Windows 环境和 Microsoft Excel 应用程序，或安装 pywin32 库 (pip install pywin32)",
                    "filename": str(file_path),
                }

        except Exception as e:
            logger.error(f"刷新数据透视表失败: {e}")
            return {"success": False, "message": f"刷新失败: {str(e)}"}

    def refresh_all_pivot_tables(
        self,
        filename: str,
        sheet_name: Optional[str] = None,
    ) -> dict[str, Any]:
        """刷新所有数据透视表.

        Args:
            filename: Excel文件名
            sheet_name: 工作表名称 (可选，如果指定则只刷新该表中的透视表)

        Note:
            此功能需要 Windows 环境和 Microsoft Excel 应用程序，或安装 pywin32 库
        """
        try:
            file_path = config.paths.output_dir / filename
            self.file_manager.validate_file_path(file_path, must_exist=True)

            try:
                import win32com.client as win32

                # 使用Excel COM接口刷新数据透视表
                excel = win32.gencache.EnsureDispatch('Excel.Application')
                excel.Visible = False
                excel.DisplayAlerts = False

                wb_com = excel.Workbooks.Open(str(file_path.absolute()))

                refreshed_count = 0
                if sheet_name:
                    # 只刷新指定工作表中的透视表
                    if sheet_name not in [ws.Name for ws in wb_com.Worksheets]:
                        raise ValueError(f"工作表 '{sheet_name}' 不存在")

                    ws_com = wb_com.Worksheets(sheet_name)
                    for pivot_table in ws_com.PivotTables():
                        pivot_table.RefreshTable()
                        refreshed_count += 1
                else:
                    # 刷新所有工作表中的所有透视表
                    for ws_com in wb_com.Worksheets:
                        for pivot_table in ws_com.PivotTables():
                            pivot_table.RefreshTable()
                            refreshed_count += 1

                wb_com.Save()
                wb_com.Close(SaveChanges=True)
                excel.Quit()

                logger.info(f"刷新所有数据透视表成功: {filename}, 共刷新 {refreshed_count} 个")
                return {
                    "success": True,
                    "message": f"成功刷新 {refreshed_count} 个数据透视表",
                    "filename": str(file_path),
                    "sheet_name": sheet_name,
                    "refreshed_count": refreshed_count,
                }

            except ImportError:
                return {
                    "success": False,
                    "message": "数据透视表刷新功能需要 Windows 环境和 Microsoft Excel 应用程序，或安装 pywin32 库 (pip install pywin32)",
                    "filename": str(file_path),
                }

        except Exception as e:
            logger.error(f"刷新所有数据透视表失败: {e}")
            return {"success": False, "message": f"刷新失败: {str(e)}"}

    def change_pivot_data_source(
        self,
        filename: str,
        pivot_sheet: str,
        pivot_table_name: str,
        new_source_range: str,
    ) -> dict[str, Any]:
        """更改数据透视表数据源.

        Args:
            filename: 文件名
            pivot_sheet: 数据透视表所在工作表
            pivot_table_name: 数据透视表名称
            new_source_range: 新数据源范围 (如 'Sheet1!A1:E100')

        Returns:
            操作结果
        """
        try:
            file_path = config.paths.output_dir / filename
            self.file_manager.validate_file_path(file_path, must_exist=True)

            try:
                import win32com.client as win32

                # 使用Excel COM接口更改数据源
                excel = win32.gencache.EnsureDispatch('Excel.Application')
                excel.Visible = False
                excel.DisplayAlerts = False

                wb_com = excel.Workbooks.Open(str(file_path.absolute()))

                if pivot_sheet not in [ws.Name for ws in wb_com.Worksheets]:
                    raise ValueError(f"工作表 '{pivot_sheet}' 不存在")

                ws_com = wb_com.Worksheets(pivot_sheet)

                # 查找指定的数据透视表
                pivot_table = None
                for pt in ws_com.PivotTables():
                    if pt.Name == pivot_table_name:
                        pivot_table = pt
                        break

                if pivot_table is None:
                    raise ValueError(f"数据透视表 '{pivot_table_name}' 不存在")

                # 更改数据源
                pivot_table.ChangePivotCache(
                    wb_com.PivotCaches().Create(
                        SourceType=1,  # xlDatabase
                        SourceData=new_source_range
                    )
                )

                # 刷新数据
                pivot_table.RefreshTable()

                wb_com.Save()
                wb_com.Close(SaveChanges=True)
                excel.Quit()

                logger.info(f"更改数据透视表数据源成功: {filename}")
                return {
                    "success": True,
                    "message": "数据源更改成功",
                    "filename": str(file_path),
                    "pivot_table": pivot_table_name,
                    "new_source": new_source_range,
                }

            except ImportError:
                return {
                    "success": False,
                    "message": "更改数据透视表数据源需要 Windows 环境和 Microsoft Excel 应用程序，或安装 pywin32 库 (pip install pywin32)",
                    "filename": str(file_path),
                }

        except Exception as e:
            logger.error(f"更改数据透视表数据源失败: {e}")
            return {"success": False, "message": f"操作失败: {str(e)}"}

