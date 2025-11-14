"""Excel 打印设置模块."""

from typing import Any, Optional

from openpyxl import load_workbook
from openpyxl.worksheet.page import PageMargins, PrintPageSetup
from openpyxl.worksheet.pagebreak import Break
from loguru import logger

from office_mcp_server.config import config
from office_mcp_server.utils.file_manager import FileManager


class ExcelPrintOperations:
    """Excel 打印设置操作类."""

    def __init__(self) -> None:
        """初始化打印设置操作类."""
        self.file_manager = FileManager()

    def set_page_setup(
        self,
        filename: str,
        sheet_name: str,
        orientation: str = "portrait",
        paper_size: int = 9,
        scale: int = 100,
        fit_to_width: Optional[int] = None,
        fit_to_height: Optional[int] = None,
    ) -> dict[str, Any]:
        """设置页面设置.

        Args:
            filename: 文件名
            sheet_name: 工作表名称
            orientation: 页面方向 ('portrait'纵向, 'landscape'横向)
            paper_size: 纸张大小 (1=Letter, 5=Legal, 8=A3, 9=A4, 11=A5, 默认9)
            scale: 缩放比例 (10-400, 默认100)
            fit_to_width: 适合页宽 (可选)
            fit_to_height: 适合页高 (可选)
        """
        try:
            file_path = config.paths.output_dir / filename
            self.file_manager.validate_file_path(file_path, must_exist=True)

            wb = load_workbook(str(file_path))

            if sheet_name not in wb.sheetnames:
                raise ValueError(f"工作表 '{sheet_name}' 不存在")

            ws = wb[sheet_name]

            ws.page_setup.orientation = orientation
            ws.page_setup.paperSize = paper_size
            ws.page_setup.scale = scale

            if fit_to_width is not None:
                ws.page_setup.fitToWidth = fit_to_width
            if fit_to_height is not None:
                ws.page_setup.fitToHeight = fit_to_height

            wb.save(str(file_path))
            wb.close()

            logger.info(f"页面设置成功: {file_path}")
            return {
                "success": True,
                "message": "页面设置成功",
                "filename": str(file_path),
                "sheet_name": sheet_name,
                "orientation": orientation,
                "paper_size": paper_size,
            }

        except Exception as e:
            logger.error(f"页面设置失败: {e}")
            return {"success": False, "message": f"设置失败: {str(e)}"}

    def set_page_margins(
        self,
        filename: str,
        sheet_name: str,
        left: float = 0.75,
        right: float = 0.75,
        top: float = 1.0,
        bottom: float = 1.0,
        header: float = 0.5,
        footer: float = 0.5,
    ) -> dict[str, Any]:
        """设置页边距.

        Args:
            filename: 文件名
            sheet_name: 工作表名称
            left: 左边距 (英寸, 默认0.75)
            right: 右边距 (英寸, 默认0.75)
            top: 上边距 (英寸, 默认1.0)
            bottom: 下边距 (英寸, 默认1.0)
            header: 页眉边距 (英寸, 默认0.5)
            footer: 页脚边距 (英寸, 默认0.5)
        """
        try:
            file_path = config.paths.output_dir / filename
            self.file_manager.validate_file_path(file_path, must_exist=True)

            wb = load_workbook(str(file_path))

            if sheet_name not in wb.sheetnames:
                raise ValueError(f"工作表 '{sheet_name}' 不存在")

            ws = wb[sheet_name]

            ws.page_margins = PageMargins(
                left=left,
                right=right,
                top=top,
                bottom=bottom,
                header=header,
                footer=footer
            )

            wb.save(str(file_path))
            wb.close()

            logger.info(f"页边距设置成功: {file_path}")
            return {
                "success": True,
                "message": "页边距设置成功",
                "filename": str(file_path),
                "sheet_name": sheet_name,
                "margins": {
                    "left": left,
                    "right": right,
                    "top": top,
                    "bottom": bottom,
                }
            }

        except Exception as e:
            logger.error(f"页边距设置失败: {e}")
            return {"success": False, "message": f"设置失败: {str(e)}"}

    def set_print_area(
        self,
        filename: str,
        sheet_name: str,
        print_area: Optional[str] = None,
    ) -> dict[str, Any]:
        """设置打印区域.

        Args:
            filename: 文件名
            sheet_name: 工作表名称
            print_area: 打印区域 (如 'A1:F20', None为清除打印区域)
        """
        try:
            file_path = config.paths.output_dir / filename
            self.file_manager.validate_file_path(file_path, must_exist=True)

            wb = load_workbook(str(file_path))

            if sheet_name not in wb.sheetnames:
                raise ValueError(f"工作表 '{sheet_name}' 不存在")

            ws = wb[sheet_name]

            if print_area:
                ws.print_area = f"{sheet_name}!{print_area}"
                message = f"打印区域已设置为 {print_area}"
            else:
                ws.print_area = None
                message = "打印区域已清除"

            wb.save(str(file_path))
            wb.close()

            logger.info(f"打印区域设置成功: {file_path}")
            return {
                "success": True,
                "message": message,
                "filename": str(file_path),
                "sheet_name": sheet_name,
                "print_area": print_area,
            }

        except Exception as e:
            logger.error(f"打印区域设置失败: {e}")
            return {"success": False, "message": f"设置失败: {str(e)}"}

    def set_print_titles(
        self,
        filename: str,
        sheet_name: str,
        rows: Optional[str] = None,
        cols: Optional[str] = None,
    ) -> dict[str, Any]:
        """设置打印标题(重复打印的行或列).

        Args:
            filename: 文件名
            sheet_name: 工作表名称
            rows: 重复打印的行 (如 '1:1' 或 '1:3')
            cols: 重复打印的列 (如 'A:A' 或 'A:B')
        """
        try:
            file_path = config.paths.output_dir / filename
            self.file_manager.validate_file_path(file_path, must_exist=True)

            wb = load_workbook(str(file_path))

            if sheet_name not in wb.sheetnames:
                raise ValueError(f"工作表 '{sheet_name}' 不存在")

            ws = wb[sheet_name]

            # 使用 defined_names 设置打印标题
            from openpyxl.workbook.defined_name import DefinedName

            # 删除现有的打印标题定义
            if '_xlnm.Print_Titles' in wb.defined_names:
                del wb.defined_names['_xlnm.Print_Titles']

            title_parts = []
            if rows:
                title_parts.append(f"${rows}")
            if cols:
                title_parts.append(f"${cols}")

            if title_parts:
                # 创建打印标题定义
                value = f"'{sheet_name}'!" + ",".join(title_parts)
                defined_name = DefinedName('_xlnm.Print_Titles', attr_text=value)
                defined_name.localSheetId = wb.sheetnames.index(sheet_name)
                wb.defined_names.add(defined_name)
                message = f"打印标题已设置"
            else:
                message = "打印标题已清除"

            wb.save(str(file_path))
            wb.close()

            logger.info(f"打印标题设置成功: {file_path}")
            return {
                "success": True,
                "message": message,
                "filename": str(file_path),
                "sheet_name": sheet_name,
                "rows": rows,
                "cols": cols,
            }

        except Exception as e:
            logger.error(f"打印标题设置失败: {e}")
            return {"success": False, "message": f"设置失败: {str(e)}"}

    def insert_page_break(
        self,
        filename: str,
        sheet_name: str,
        cell: str,
        break_type: str = "row",
    ) -> dict[str, Any]:
        """插入分页符.

        Args:
            filename: 文件名
            sheet_name: 工作表名称
            cell: 起始单元格 (如 'B2')
            break_type: 分页符类型 ('row'行分页符, 'col'列分页符, 默认 'row')
        """
        try:
            import re
            from openpyxl.utils import column_index_from_string

            file_path = config.paths.output_dir / filename
            self.file_manager.validate_file_path(file_path, must_exist=True)

            wb = load_workbook(str(file_path))

            if sheet_name not in wb.sheetnames:
                raise ValueError(f"工作表 '{sheet_name}' 不存在")

            ws = wb[sheet_name]

            # 解析单元格位置
            match = re.match(r'^([A-Z]+)(\d+)$', cell.upper())
            if not match:
                raise ValueError(f"无效的单元格格式: {cell}")
            col_letter = match.group(1)
            row = int(match.group(2))
            col_idx = column_index_from_string(col_letter)

            if break_type == 'row':
                ws.row_breaks.append(Break(id=row))
                message = f"在第 {row} 行插入分页符"
            elif break_type == 'col':
                ws.col_breaks.append(Break(id=col_idx))
                message = f"在第 {col_idx} 列插入分页符"
            else:
                raise ValueError(f"不支持的分页符类型: {break_type}")

            wb.save(str(file_path))
            wb.close()

            logger.info(f"分页符插入成功: {file_path}")
            return {
                "success": True,
                "message": message,
                "filename": str(file_path),
                "sheet_name": sheet_name,
                "break_type": break_type,
                "position": position,
            }

        except Exception as e:
            logger.error(f"分页符插入失败: {e}")
            return {"success": False, "message": f"插入失败: {str(e)}"}

    def delete_page_break(
        self,
        filename: str,
        sheet_name: str,
        break_type: str,
        position: int,
    ) -> dict[str, Any]:
        """删除分页符.

        Args:
            filename: 文件名
            sheet_name: 工作表名称
            break_type: 分页符类型 ('row'行分页符, 'col'列分页符)
            position: 分页符位置 (行号或列号)
        """
        try:
            file_path = config.paths.output_dir / filename
            self.file_manager.validate_file_path(file_path, must_exist=True)

            wb = load_workbook(str(file_path))

            if sheet_name not in wb.sheetnames:
                raise ValueError(f"工作表 '{sheet_name}' 不存在")

            ws = wb[sheet_name]

            if break_type == 'row':
                # 查找并删除指定位置的行分页符
                breaks = [brk for brk in ws.row_breaks if brk.id != position]
                ws.row_breaks.brk = breaks
                message = f"删除第 {position} 行的分页符"
            elif break_type == 'col':
                # 查找并删除指定位置的列分页符
                breaks = [brk for brk in ws.col_breaks if brk.id != position]
                ws.col_breaks.brk = breaks
                message = f"删除第 {position} 列的分页符"
            else:
                raise ValueError(f"不支持的分页符类型: {break_type}")

            wb.save(str(file_path))
            wb.close()

            logger.info(f"分页符删除成功: {file_path}")
            return {
                "success": True,
                "message": message,
                "filename": str(file_path),
                "sheet_name": sheet_name,
                "break_type": break_type,
                "position": position,
            }

        except Exception as e:
            logger.error(f"分页符删除失败: {e}")
            return {"success": False, "message": f"删除失败: {str(e)}"}

    def clear_all_page_breaks(
        self,
        filename: str,
        sheet_name: str,
    ) -> dict[str, Any]:
        """清除所有分页符.

        Args:
            filename: 文件名
            sheet_name: 工作表名称
        """
        try:
            file_path = config.paths.output_dir / filename
            self.file_manager.validate_file_path(file_path, must_exist=True)

            wb = load_workbook(str(file_path))

            if sheet_name not in wb.sheetnames:
                raise ValueError(f"工作表 '{sheet_name}' 不存在")

            ws = wb[sheet_name]

            # 清除所有行分页符
            row_breaks_count = len(ws.row_breaks.brk) if ws.row_breaks else 0
            ws.row_breaks.brk = []

            # 清除所有列分页符
            col_breaks_count = len(ws.col_breaks.brk) if ws.col_breaks else 0
            ws.col_breaks.brk = []

            wb.save(str(file_path))
            wb.close()

            logger.info(f"清除所有分页符成功: {file_path}")
            return {
                "success": True,
                "message": f"成功清除 {row_breaks_count} 个行分页符和 {col_breaks_count} 个列分页符",
                "filename": str(file_path),
                "sheet_name": sheet_name,
                "row_breaks_cleared": row_breaks_count,
                "col_breaks_cleared": col_breaks_count,
            }

        except Exception as e:
            logger.error(f"清除分页符失败: {e}")
            return {"success": False, "message": f"清除失败: {str(e)}"}

