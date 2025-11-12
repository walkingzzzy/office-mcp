"""Excel 报表自动化模块."""

import shutil
from typing import Any, Optional, Dict
from pathlib import Path
from datetime import datetime

from openpyxl import load_workbook
from loguru import logger

from office_mcp_server.config import config
from office_mcp_server.utils.file_manager import FileManager


class ExcelReportAutomation:
    """Excel 报表自动化操作类."""

    def __init__(self) -> None:
        """初始化报表自动化操作类."""
        self.file_manager = FileManager()

    def generate_report_from_template(
        self,
        template_file: str,
        output_file: str,
        data: dict[str, Any],
        mappings: Optional[dict[str, str]] = None,
    ) -> dict[str, Any]:
        """基于模板生成报表.

        Args:
            template_file: 模板文件路径
            output_file: 输出文件路径
            data: 数据字典 (键为数据名, 值为数据)
            mappings: 映射字典 (键为数据名, 值为单元格位置如'A1')
        """
        try:
            template_path = Path(template_file)
            if not template_path.is_absolute():
                template_path = config.paths.output_dir / template_file

            if not template_path.exists():
                raise FileNotFoundError(f"模板文件不存在: {template_path}")

            output_path = config.paths.output_dir / output_file
            self.file_manager.ensure_directory(output_path.parent)

            # 复制模板
            shutil.copy(template_path, output_path)

            # 打开工作簿并填充数据
            wb = load_workbook(str(output_path))
            ws = wb.active

            if mappings:
                # 使用映射填充数据
                for data_key, cell_ref in mappings.items():
                    if data_key in data:
                        ws[cell_ref] = data[data_key]
            else:
                # 自动搜索并替换占位符 (如 {{variable_name}})
                for row in ws.iter_rows():
                    for cell in row:
                        if cell.value and isinstance(cell.value, str):
                            for key, value in data.items():
                                placeholder = f"{{{{{key}}}}}"
                                if placeholder in cell.value:
                                    cell.value = cell.value.replace(placeholder, str(value))

            # 添加生成时间
            ws['A1'].comment = None
            from openpyxl.comments import Comment
            timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            ws['A1'].comment = Comment(f"报表生成于: {timestamp}", "System")

            wb.save(str(output_path))
            wb.close()

            logger.info(f"报表生成成功: {output_path}")
            return {
                "success": True,
                "message": "报表生成成功",
                "template": str(template_path),
                "output_file": str(output_path),
                "data_count": len(data),
                "timestamp": timestamp,
            }

        except Exception as e:
            logger.error(f"报表生成失败: {e}")
            return {"success": False, "message": f"生成失败: {str(e)}"}

    def update_report_data(
        self,
        filename: str,
        sheet_name: str,
        updates: dict[str, Any],
    ) -> dict[str, Any]:
        """更新报表数据.

        Args:
            filename: 文件名
            sheet_name: 工作表名称
            updates: 更新字典 (键为单元格位置, 值为新值)
        """
        try:
            file_path = config.paths.output_dir / filename
            self.file_manager.validate_file_path(file_path, must_exist=True)

            wb = load_workbook(str(file_path))

            if sheet_name not in wb.sheetnames:
                raise ValueError(f"工作表 '{sheet_name}' 不存在")

            ws = wb[sheet_name]

            update_count = 0
            for cell_ref, value in updates.items():
                ws[cell_ref] = value
                update_count += 1

            wb.save(str(file_path))
            wb.close()

            logger.info(f"报表数据更新成功: {file_path}, 更新 {update_count} 个单元格")
            return {
                "success": True,
                "message": f"成功更新 {update_count} 个单元格",
                "filename": str(file_path),
                "sheet_name": sheet_name,
                "update_count": update_count,
            }

        except Exception as e:
            logger.error(f"报表数据更新失败: {e}")
            return {"success": False, "message": f"更新失败: {str(e)}"}

    def consolidate_reports(
        self,
        source_files: list[str],
        output_file: str,
        sheet_name: str = "Consolidated",
        include_source_name: bool = True,
    ) -> dict[str, Any]:
        """合并多个报表.

        Args:
            source_files: 源文件列表
            output_file: 输出文件名
            sheet_name: 目标工作表名称
            include_source_name: 是否包含源文件名列
        """
        try:
            from openpyxl import Workbook

            wb_out = Workbook()
            ws_out = wb_out.active
            ws_out.title = sheet_name

            current_row = 1
            first_file = True

            for source_file in source_files:
                source_path = config.paths.output_dir / source_file
                if not source_path.exists():
                    logger.warning(f"文件不存在,跳过: {source_file}")
                    continue

                wb_source = load_workbook(str(source_path))
                ws_source = wb_source.active

                # 获取数据
                for row_idx, row in enumerate(ws_source.iter_rows(values_only=True), 1):
                    # 跳过第一个文件之后的标题行
                    if not first_file and row_idx == 1:
                        continue

                    row_data = list(row)

                    # 添加源文件名
                    if include_source_name:
                        row_data.insert(0, Path(source_file).stem)

                    ws_out.append(row_data)
                    current_row += 1

                wb_source.close()
                first_file = False

            # 如果添加了源文件名列,添加表头
            if include_source_name and current_row > 1:
                ws_out.insert_rows(1)
                ws_out['A1'] = "Source"

            output_path = config.paths.output_dir / output_file
            wb_out.save(str(output_path))
            wb_out.close()

            logger.info(f"报表合并成功: {output_path}")
            return {
                "success": True,
                "message": f"成功合并 {len(source_files)} 个报表",
                "output_file": str(output_path),
                "source_count": len(source_files),
                "total_rows": current_row - 1,
            }

        except Exception as e:
            logger.error(f"报表合并失败: {e}")
            return {"success": False, "message": f"合并失败: {str(e)}"}

    def schedule_report_generation(
        self,
        template_file: str,
        output_pattern: str,
        data_source: str,
        frequency: str = "daily",
    ) -> dict[str, Any]:
        """计划报表生成.

        Args:
            template_file: 模板文件
            output_pattern: 输出文件名模式 (如 'report_{date}.xlsx')
            data_source: 数据源
            frequency: 频率 ('daily', 'weekly', 'monthly')

        Note:
            此功能需要任务调度器支持,这里仅返回配置信息
        """
        try:
            config_info = {
                "template": template_file,
                "output_pattern": output_pattern,
                "data_source": data_source,
                "frequency": frequency,
                "next_run": "需要外部调度器支持",
            }

            logger.info(f"报表计划创建成功")
            return {
                "success": True,
                "message": "报表计划已创建 (需要外部调度器执行)",
                "config": config_info,
                "note": "此功能需要配合系统任务调度器使用 (如 cron 或 Windows Task Scheduler)"
            }

        except Exception as e:
            logger.error(f"创建报表计划失败: {e}")
            return {"success": False, "message": f"创建失败: {str(e)}"}
