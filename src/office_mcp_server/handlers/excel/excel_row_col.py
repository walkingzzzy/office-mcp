"""Excel 行列操作模块."""

from typing import Any, Optional

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
from loguru import logger

from office_mcp_server.config import config
from office_mcp_server.utils.file_manager import FileManager


class ExcelRowColOperations:
    """Excel 行列操作类."""

    def __init__(self) -> None:
        """初始化行列操作类."""
        self.file_manager = FileManager()

    def insert_rows(
        self,
        filename: str,
        sheet_name: str,
        row_index: int,
        count: int = 1,
    ) -> dict[str, Any]:
        """插入行.

        Args:
            filename: 文件名
            sheet_name: 工作表名称
            row_index: 插入位置的行索引 (从1开始)
            count: 插入的行数 (默认1)
        """
        try:
            file_path = config.paths.output_dir / filename
            self.file_manager.validate_file_path(file_path, must_exist=True)

            wb = load_workbook(str(file_path))

            if sheet_name not in wb.sheetnames:
                raise ValueError(f"工作表 '{sheet_name}' 不存在")

            ws = wb[sheet_name]
            ws.insert_rows(row_index, count)

            wb.save(str(file_path))
            wb.close()

            logger.info(f"插入行成功: {file_path}, 行索引: {row_index}, 数量: {count}")
            return {
                "success": True,
                "message": f"成功在第 {row_index} 行插入 {count} 行",
                "filename": str(file_path),
                "row_index": row_index,
                "count": count,
            }

        except Exception as e:
            logger.error(f"插入行失败: {e}")
            return {"success": False, "message": f"插入行失败: {str(e)}"}

    def delete_rows(
        self,
        filename: str,
        sheet_name: str,
        row_index: int,
        count: int = 1,
    ) -> dict[str, Any]:
        """删除行.

        Args:
            filename: 文件名
            sheet_name: 工作表名称
            row_index: 删除起始行索引 (从1开始)
            count: 删除的行数 (默认1)
        """
        try:
            file_path = config.paths.output_dir / filename
            self.file_manager.validate_file_path(file_path, must_exist=True)

            wb = load_workbook(str(file_path))

            if sheet_name not in wb.sheetnames:
                raise ValueError(f"工作表 '{sheet_name}' 不存在")

            ws = wb[sheet_name]
            ws.delete_rows(row_index, count)

            wb.save(str(file_path))
            wb.close()

            logger.info(f"删除行成功: {file_path}, 行索引: {row_index}, 数量: {count}")
            return {
                "success": True,
                "message": f"成功删除从第 {row_index} 行开始的 {count} 行",
                "filename": str(file_path),
                "row_index": row_index,
                "count": count,
            }

        except Exception as e:
            logger.error(f"删除行失败: {e}")
            return {"success": False, "message": f"删除行失败: {str(e)}"}

    def insert_cols(
        self,
        filename: str,
        sheet_name: str,
        col_index: int,
        count: int = 1,
    ) -> dict[str, Any]:
        """插入列.

        Args:
            filename: 文件名
            sheet_name: 工作表名称
            col_index: 插入位置的列索引 (从1开始, A=1, B=2...)
            count: 插入的列数 (默认1)
        """
        try:
            file_path = config.paths.output_dir / filename
            self.file_manager.validate_file_path(file_path, must_exist=True)

            wb = load_workbook(str(file_path))

            if sheet_name not in wb.sheetnames:
                raise ValueError(f"工作表 '{sheet_name}' 不存在")

            ws = wb[sheet_name]
            ws.insert_cols(col_index, count)

            wb.save(str(file_path))
            wb.close()

            col_letter = get_column_letter(col_index)
            logger.info(f"插入列成功: {file_path}, 列: {col_letter}, 数量: {count}")
            return {
                "success": True,
                "message": f"成功在第 {col_letter} 列插入 {count} 列",
                "filename": str(file_path),
                "col_index": col_index,
                "col_letter": col_letter,
                "count": count,
            }

        except Exception as e:
            logger.error(f"插入列失败: {e}")
            return {"success": False, "message": f"插入列失败: {str(e)}"}

    def delete_cols(
        self,
        filename: str,
        sheet_name: str,
        col_index: int,
        count: int = 1,
    ) -> dict[str, Any]:
        """删除列.

        Args:
            filename: 文件名
            sheet_name: 工作表名称
            col_index: 删除起始列索引 (从1开始, A=1, B=2...)
            count: 删除的列数 (默认1)
        """
        try:
            file_path = config.paths.output_dir / filename
            self.file_manager.validate_file_path(file_path, must_exist=True)

            wb = load_workbook(str(file_path))

            if sheet_name not in wb.sheetnames:
                raise ValueError(f"工作表 '{sheet_name}' 不存在")

            ws = wb[sheet_name]
            ws.delete_cols(col_index, count)

            wb.save(str(file_path))
            wb.close()

            col_letter = get_column_letter(col_index)
            logger.info(f"删除列成功: {file_path}, 列: {col_letter}, 数量: {count}")
            return {
                "success": True,
                "message": f"成功删除从第 {col_letter} 列开始的 {count} 列",
                "filename": str(file_path),
                "col_index": col_index,
                "col_letter": col_letter,
                "count": count,
            }

        except Exception as e:
            logger.error(f"删除列失败: {e}")
            return {"success": False, "message": f"删除列失败: {str(e)}"}

    def hide_rows(
        self,
        filename: str,
        sheet_name: str,
        row_start: int,
        row_end: Optional[int] = None,
    ) -> dict[str, Any]:
        """隐藏行.

        Args:
            filename: 文件名
            sheet_name: 工作表名称
            row_start: 起始行索引 (从1开始)
            row_end: 结束行索引 (可选, 如果不指定则只隐藏单行)
        """
        try:
            file_path = config.paths.output_dir / filename
            self.file_manager.validate_file_path(file_path, must_exist=True)

            wb = load_workbook(str(file_path))

            if sheet_name not in wb.sheetnames:
                raise ValueError(f"工作表 '{sheet_name}' 不存在")

            ws = wb[sheet_name]

            if row_end is None:
                row_end = row_start

            for row_idx in range(row_start, row_end + 1):
                ws.row_dimensions[row_idx].hidden = True

            wb.save(str(file_path))
            wb.close()

            logger.info(f"隐藏行成功: {file_path}, 行范围: {row_start}-{row_end}")
            return {
                "success": True,
                "message": f"成功隐藏第 {row_start} 到 {row_end} 行",
                "filename": str(file_path),
                "row_start": row_start,
                "row_end": row_end,
            }

        except Exception as e:
            logger.error(f"隐藏行失败: {e}")
            return {"success": False, "message": f"隐藏行失败: {str(e)}"}

    def show_rows(
        self,
        filename: str,
        sheet_name: str,
        row_start: int,
        row_end: Optional[int] = None,
    ) -> dict[str, Any]:
        """显示行.

        Args:
            filename: 文件名
            sheet_name: 工作表名称
            row_start: 起始行索引 (从1开始)
            row_end: 结束行索引 (可选, 如果不指定则只显示单行)
        """
        try:
            file_path = config.paths.output_dir / filename
            self.file_manager.validate_file_path(file_path, must_exist=True)

            wb = load_workbook(str(file_path))

            if sheet_name not in wb.sheetnames:
                raise ValueError(f"工作表 '{sheet_name}' 不存在")

            ws = wb[sheet_name]

            if row_end is None:
                row_end = row_start

            for row_idx in range(row_start, row_end + 1):
                ws.row_dimensions[row_idx].hidden = False

            wb.save(str(file_path))
            wb.close()

            logger.info(f"显示行成功: {file_path}, 行范围: {row_start}-{row_end}")
            return {
                "success": True,
                "message": f"成功显示第 {row_start} 到 {row_end} 行",
                "filename": str(file_path),
                "row_start": row_start,
                "row_end": row_end,
            }

        except Exception as e:
            logger.error(f"显示行失败: {e}")
            return {"success": False, "message": f"显示行失败: {str(e)}"}

    def hide_cols(
        self,
        filename: str,
        sheet_name: str,
        col_start: int,
        col_end: Optional[int] = None,
    ) -> dict[str, Any]:
        """隐藏列.

        Args:
            filename: 文件名
            sheet_name: 工作表名称
            col_start: 起始列索引 (从1开始, A=1, B=2...)
            col_end: 结束列索引 (可选, 如果不指定则只隐藏单列)
        """
        try:
            file_path = config.paths.output_dir / filename
            self.file_manager.validate_file_path(file_path, must_exist=True)

            wb = load_workbook(str(file_path))

            if sheet_name not in wb.sheetnames:
                raise ValueError(f"工作表 '{sheet_name}' 不存在")

            ws = wb[sheet_name]

            if col_end is None:
                col_end = col_start

            for col_idx in range(col_start, col_end + 1):
                col_letter = get_column_letter(col_idx)
                ws.column_dimensions[col_letter].hidden = True

            wb.save(str(file_path))
            wb.close()

            col_start_letter = get_column_letter(col_start)
            col_end_letter = get_column_letter(col_end)
            logger.info(f"隐藏列成功: {file_path}, 列范围: {col_start_letter}-{col_end_letter}")
            return {
                "success": True,
                "message": f"成功隐藏第 {col_start_letter} 到 {col_end_letter} 列",
                "filename": str(file_path),
                "col_start": col_start,
                "col_end": col_end,
            }

        except Exception as e:
            logger.error(f"隐藏列失败: {e}")
            return {"success": False, "message": f"隐藏列失败: {str(e)}"}

    def show_cols(
        self,
        filename: str,
        sheet_name: str,
        col_start: int,
        col_end: Optional[int] = None,
    ) -> dict[str, Any]:
        """显示列.

        Args:
            filename: 文件名
            sheet_name: 工作表名称
            col_start: 起始列索引 (从1开始, A=1, B=2...)
            col_end: 结束列索引 (可选, 如果不指定则只显示单列)
        """
        try:
            file_path = config.paths.output_dir / filename
            self.file_manager.validate_file_path(file_path, must_exist=True)

            wb = load_workbook(str(file_path))

            if sheet_name not in wb.sheetnames:
                raise ValueError(f"工作表 '{sheet_name}' 不存在")

            ws = wb[sheet_name]

            if col_end is None:
                col_end = col_start

            for col_idx in range(col_start, col_end + 1):
                col_letter = get_column_letter(col_idx)
                ws.column_dimensions[col_letter].hidden = False

            wb.save(str(file_path))
            wb.close()

            col_start_letter = get_column_letter(col_start)
            col_end_letter = get_column_letter(col_end)
            logger.info(f"显示列成功: {file_path}, 列范围: {col_start_letter}-{col_end_letter}")
            return {
                "success": True,
                "message": f"成功显示第 {col_start_letter} 到 {col_end_letter} 列",
                "filename": str(file_path),
                "col_start": col_start,
                "col_end": col_end,
            }

        except Exception as e:
            logger.error(f"显示列失败: {e}")
            return {"success": False, "message": f"显示列失败: {str(e)}"}

    def set_row_height(
        self,
        filename: str,
        sheet_name: str,
        row_index: int,
        height: float,
    ) -> dict[str, Any]:
        """设置行高.

        Args:
            filename: 文件名
            sheet_name: 工作表名称
            row_index: 行索引 (从1开始)
            height: 行高 (磅值)
        """
        try:
            file_path = config.paths.output_dir / filename
            self.file_manager.validate_file_path(file_path, must_exist=True)

            wb = load_workbook(str(file_path))

            if sheet_name not in wb.sheetnames:
                raise ValueError(f"工作表 '{sheet_name}' 不存在")

            ws = wb[sheet_name]
            ws.row_dimensions[row_index].height = height

            wb.save(str(file_path))
            wb.close()

            logger.info(f"设置行高成功: {file_path}, 行: {row_index}, 高度: {height}")
            return {
                "success": True,
                "message": f"成功设置第 {row_index} 行高度为 {height} 磅",
                "filename": str(file_path),
                "row_index": row_index,
                "height": height,
            }

        except Exception as e:
            logger.error(f"设置行高失败: {e}")
            return {"success": False, "message": f"设置行高失败: {str(e)}"}

    def set_col_width(
        self,
        filename: str,
        sheet_name: str,
        col_index: int,
        width: float,
    ) -> dict[str, Any]:
        """设置列宽.

        Args:
            filename: 文件名
            sheet_name: 工作表名称
            col_index: 列索引 (从1开始, A=1, B=2...)
            width: 列宽 (字符单位)
        """
        try:
            file_path = config.paths.output_dir / filename
            self.file_manager.validate_file_path(file_path, must_exist=True)

            wb = load_workbook(str(file_path))

            if sheet_name not in wb.sheetnames:
                raise ValueError(f"工作表 '{sheet_name}' 不存在")

            ws = wb[sheet_name]
            col_letter = get_column_letter(col_index)
            ws.column_dimensions[col_letter].width = width

            wb.save(str(file_path))
            wb.close()

            logger.info(f"设置列宽成功: {file_path}, 列: {col_letter}, 宽度: {width}")
            return {
                "success": True,
                "message": f"成功设置第 {col_letter} 列宽度为 {width} 字符",
                "filename": str(file_path),
                "col_index": col_index,
                "col_letter": col_letter,
                "width": width,
            }

        except Exception as e:
            logger.error(f"设置列宽失败: {e}")
            return {"success": False, "message": f"设置列宽失败: {str(e)}"}

    def copy_rows(
        self,
        filename: str,
        sheet_name: str,
        source_row: int,
        target_row: int,
        count: int = 1,
    ) -> dict[str, Any]:
        """复制行.

        Args:
            filename: 文件名
            sheet_name: 工作表名称
            source_row: 源行索引 (从1开始)
            target_row: 目标行索引 (从1开始)
            count: 复制行数 (默认 1)
        """
        try:
            file_path = config.paths.output_dir / filename
            self.file_manager.validate_file_path(file_path, must_exist=True)

            wb = load_workbook(str(file_path))

            if sheet_name not in wb.sheetnames:
                raise ValueError(f"工作表 '{sheet_name}' 不存在")

            ws = wb[sheet_name]

            # 在目标位置插入空行
            ws.insert_rows(target_row, count)

            # 复制数据和格式
            for i in range(count):
                src_row_idx = source_row + i
                tgt_row_idx = target_row + i

                for col_idx in range(1, ws.max_column + 1):
                    src_cell = ws.cell(row=src_row_idx, column=col_idx)
                    tgt_cell = ws.cell(row=tgt_row_idx, column=col_idx)

                    # 复制值
                    tgt_cell.value = src_cell.value

                    # 复制格式
                    if src_cell.has_style:
                        tgt_cell.font = src_cell.font.copy()
                        tgt_cell.border = src_cell.border.copy()
                        tgt_cell.fill = src_cell.fill.copy()
                        tgt_cell.number_format = src_cell.number_format
                        tgt_cell.protection = src_cell.protection.copy()
                        tgt_cell.alignment = src_cell.alignment.copy()

            wb.save(str(file_path))
            wb.close()

            logger.info(f"复制行成功: {file_path}, 从第 {source_row} 行复制 {count} 行到第 {target_row} 行")
            return {
                "success": True,
                "message": f"成功从第 {source_row} 行复制 {count} 行到第 {target_row} 行",
                "filename": str(file_path),
                "source_row": source_row,
                "target_row": target_row,
                "count": count,
            }

        except Exception as e:
            logger.error(f"复制行失败: {e}")
            return {"success": False, "message": f"复制行失败: {str(e)}"}

    def copy_cols(
        self,
        filename: str,
        sheet_name: str,
        source_col: int,
        target_col: int,
        count: int = 1,
    ) -> dict[str, Any]:
        """复制列.

        Args:
            filename: 文件名
            sheet_name: 工作表名称
            source_col: 源列索引 (从1开始)
            target_col: 目标列索引 (从1开始)
            count: 复制列数 (默认 1)
        """
        try:
            file_path = config.paths.output_dir / filename
            self.file_manager.validate_file_path(file_path, must_exist=True)

            wb = load_workbook(str(file_path))

            if sheet_name not in wb.sheetnames:
                raise ValueError(f"工作表 '{sheet_name}' 不存在")

            ws = wb[sheet_name]

            # 在目标位置插入空列
            ws.insert_cols(target_col, count)

            # 复制数据和格式
            for i in range(count):
                src_col_idx = source_col + i
                tgt_col_idx = target_col + i

                for row_idx in range(1, ws.max_row + 1):
                    src_cell = ws.cell(row=row_idx, column=src_col_idx)
                    tgt_cell = ws.cell(row=row_idx, column=tgt_col_idx)

                    # 复制值
                    tgt_cell.value = src_cell.value

                    # 复制格式
                    if src_cell.has_style:
                        tgt_cell.font = src_cell.font.copy()
                        tgt_cell.border = src_cell.border.copy()
                        tgt_cell.fill = src_cell.fill.copy()
                        tgt_cell.number_format = src_cell.number_format
                        tgt_cell.protection = src_cell.protection.copy()
                        tgt_cell.alignment = src_cell.alignment.copy()

            wb.save(str(file_path))
            wb.close()

            src_col_letter = get_column_letter(source_col)
            tgt_col_letter = get_column_letter(target_col)
            logger.info(f"复制列成功: {file_path}, 从第 {src_col_letter} 列复制 {count} 列到第 {tgt_col_letter} 列")
            return {
                "success": True,
                "message": f"成功从第 {src_col_letter} 列复制 {count} 列到第 {tgt_col_letter} 列",
                "filename": str(file_path),
                "source_col": source_col,
                "target_col": target_col,
                "count": count,
            }

        except Exception as e:
            logger.error(f"复制列失败: {e}")
            return {"success": False, "message": f"复制列失败: {str(e)}"}

    def move_rows(
        self,
        filename: str,
        sheet_name: str,
        source_row: int,
        target_row: int,
        count: int = 1,
    ) -> dict[str, Any]:
        """移动行.

        Args:
            filename: 文件名
            sheet_name: 工作表名称
            source_row: 源行索引 (从1开始)
            target_row: 目标行索引 (从1开始)
            count: 移动行数 (默认 1)
        """
        try:
            file_path = config.paths.output_dir / filename
            self.file_manager.validate_file_path(file_path, must_exist=True)

            wb = load_workbook(str(file_path))

            if sheet_name not in wb.sheetnames:
                raise ValueError(f"工作表 '{sheet_name}' 不存在")

            ws = wb[sheet_name]

            # 先复制行到目标位置
            ws.insert_rows(target_row, count)

            for i in range(count):
                src_row_idx = source_row + i if source_row < target_row else source_row + i + count
                tgt_row_idx = target_row + i

                for col_idx in range(1, ws.max_column + 1):
                    src_cell = ws.cell(row=src_row_idx, column=col_idx)
                    tgt_cell = ws.cell(row=tgt_row_idx, column=col_idx)

                    tgt_cell.value = src_cell.value

                    if src_cell.has_style:
                        tgt_cell.font = src_cell.font.copy()
                        tgt_cell.border = src_cell.border.copy()
                        tgt_cell.fill = src_cell.fill.copy()
                        tgt_cell.number_format = src_cell.number_format
                        tgt_cell.protection = src_cell.protection.copy()
                        tgt_cell.alignment = src_cell.alignment.copy()

            # 删除源行
            delete_row_idx = source_row if source_row < target_row else source_row + count
            ws.delete_rows(delete_row_idx, count)

            wb.save(str(file_path))
            wb.close()

            logger.info(f"移动行成功: {file_path}, 从第 {source_row} 行移动 {count} 行到第 {target_row} 行")
            return {
                "success": True,
                "message": f"成功从第 {source_row} 行移动 {count} 行到第 {target_row} 行",
                "filename": str(file_path),
                "source_row": source_row,
                "target_row": target_row,
                "count": count,
            }

        except Exception as e:
            logger.error(f"移动行失败: {e}")
            return {"success": False, "message": f"移动行失败: {str(e)}"}

    def move_cols(
        self,
        filename: str,
        sheet_name: str,
        source_col: int,
        target_col: int,
        count: int = 1,
    ) -> dict[str, Any]:
        """移动列.

        Args:
            filename: 文件名
            sheet_name: 工作表名称
            source_col: 源列索引 (从1开始)
            target_col: 目标列索引 (从1开始)
            count: 移动列数 (默认 1)
        """
        try:
            file_path = config.paths.output_dir / filename
            self.file_manager.validate_file_path(file_path, must_exist=True)

            wb = load_workbook(str(file_path))

            if sheet_name not in wb.sheetnames:
                raise ValueError(f"工作表 '{sheet_name}' 不存在")

            ws = wb[sheet_name]

            # 先复制列到目标位置
            ws.insert_cols(target_col, count)

            for i in range(count):
                src_col_idx = source_col + i if source_col < target_col else source_col + i + count
                tgt_col_idx = target_col + i

                for row_idx in range(1, ws.max_row + 1):
                    src_cell = ws.cell(row=row_idx, column=src_col_idx)
                    tgt_cell = ws.cell(row=row_idx, column=tgt_col_idx)

                    tgt_cell.value = src_cell.value

                    if src_cell.has_style:
                        tgt_cell.font = src_cell.font.copy()
                        tgt_cell.border = src_cell.border.copy()
                        tgt_cell.fill = src_cell.fill.copy()
                        tgt_cell.number_format = src_cell.number_format
                        tgt_cell.protection = src_cell.protection.copy()
                        tgt_cell.alignment = src_cell.alignment.copy()

            # 删除源列
            delete_col_idx = source_col if source_col < target_col else source_col + count
            ws.delete_cols(delete_col_idx, count)

            wb.save(str(file_path))
            wb.close()

            src_col_letter = get_column_letter(source_col)
            tgt_col_letter = get_column_letter(target_col)
            logger.info(f"移动列成功: {file_path}, 从第 {src_col_letter} 列移动 {count} 列到第 {tgt_col_letter} 列")
            return {
                "success": True,
                "message": f"成功从第 {src_col_letter} 列移动 {count} 列到第 {tgt_col_letter} 列",
                "filename": str(file_path),
                "source_col": source_col,
                "target_col": target_col,
                "count": count,
            }

        except Exception as e:
            logger.error(f"移动列失败: {e}")
            return {"success": False, "message": f"移动列失败: {str(e)}"}

