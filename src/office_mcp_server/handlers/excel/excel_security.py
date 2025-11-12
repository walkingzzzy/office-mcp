"""Excel 数据安全功能模块."""

from typing import Any, Optional

from openpyxl import load_workbook
from loguru import logger

from office_mcp_server.config import config
from office_mcp_server.utils.file_manager import FileManager


class ExcelSecurityOperations:
    """Excel 数据安全操作类."""

    def __init__(self) -> None:
        """初始化安全操作类."""
        self.file_manager = FileManager()

    def encrypt_workbook(
        self,
        filename: str,
        password: str,
    ) -> dict[str, Any]:
        """加密工作簿.

        Args:
            filename: 文件名
            password: 密码

        Note:
            openpyxl不直接支持加密,此功能需要msoffcrypto-tool库
            pip install msoffcrypto-tool
        """
        try:
            file_path = config.paths.output_dir / filename
            self.file_manager.validate_file_path(file_path, must_exist=True)

            try:
                import msoffcrypto
                import io

                # 读取文件
                with open(file_path, 'rb') as f:
                    office_file = msoffcrypto.OfficeFile(f)
                    office_file.load_key(password=password)

                    # 加密并保存
                    with open(file_path, 'wb') as encrypted_file:
                        office_file.encrypt(password, encrypted_file)

                logger.info(f"工作簿加密成功: {file_path}")
                return {
                    "success": True,
                    "message": "工作簿已成功加密",
                    "filename": str(file_path),
                }

            except ImportError:
                return {
                    "success": False,
                    "message": "此功能需要msoffcrypto-tool库: pip install msoffcrypto-tool",
                    "note": "工作簿加密需要额外依赖"
                }

        except Exception as e:
            logger.error(f"工作簿加密失败: {e}")
            return {"success": False, "message": f"加密失败: {str(e)}"}

    def lock_cells(
        self,
        filename: str,
        sheet_name: str,
        cell_range: str,
        locked: bool = True,
    ) -> dict[str, Any]:
        """锁定/解锁单元格.

        Args:
            filename: 文件名
            sheet_name: 工作表名称
            cell_range: 单元格范围 (如 'A1:D10')
            locked: True为锁定, False为解锁
        """
        try:
            file_path = config.paths.output_dir / filename
            self.file_manager.validate_file_path(file_path, must_exist=True)

            wb = load_workbook(str(file_path))

            if sheet_name not in wb.sheetnames:
                raise ValueError(f"工作表 '{sheet_name}' 不存在")

            ws = wb[sheet_name]
            data_cells = ws[cell_range]

            from openpyxl.styles import Protection

            for row in data_cells:
                if isinstance(row, tuple):
                    for cell in row:
                        cell.protection = Protection(locked=locked)
                else:
                    row.protection = Protection(locked=locked)

            wb.save(str(file_path))
            wb.close()

            action = "锁定" if locked else "解锁"
            logger.info(f"单元格{action}成功: {file_path}")
            return {
                "success": True,
                "message": f"成功{action}单元格范围 {cell_range}",
                "filename": str(file_path),
                "sheet_name": sheet_name,
                "cell_range": cell_range,
                "locked": locked,
                "note": "需要保护工作表后锁定才生效"
            }

        except Exception as e:
            logger.error(f"单元格锁定操作失败: {e}")
            return {"success": False, "message": f"操作失败: {str(e)}"}

    def hide_formulas(
        self,
        filename: str,
        sheet_name: str,
        cell_range: str,
        hidden: bool = True,
    ) -> dict[str, Any]:
        """隐藏/显示公式.

        Args:
            filename: 文件名
            sheet_name: 工作表名称
            cell_range: 单元格范围 (如 'A1:D10')
            hidden: True为隐藏, False为显示
        """
        try:
            file_path = config.paths.output_dir / filename
            self.file_manager.validate_file_path(file_path, must_exist=True)

            wb = load_workbook(str(file_path))

            if sheet_name not in wb.sheetnames:
                raise ValueError(f"工作表 '{sheet_name}' 不存在")

            ws = wb[sheet_name]
            data_cells = ws[cell_range]

            from openpyxl.styles import Protection

            for row in data_cells:
                if isinstance(row, tuple):
                    for cell in row:
                        cell.protection = Protection(hidden=hidden)
                else:
                    row.protection = Protection(hidden=hidden)

            wb.save(str(file_path))
            wb.close()

            action = "隐藏" if hidden else "显示"
            logger.info(f"公式{action}成功: {file_path}")
            return {
                "success": True,
                "message": f"成功{action}公式 {cell_range}",
                "filename": str(file_path),
                "sheet_name": sheet_name,
                "cell_range": cell_range,
                "hidden": hidden,
                "note": "需要保护工作表后隐藏才生效"
            }

        except Exception as e:
            logger.error(f"公式隐藏操作失败: {e}")
            return {"success": False, "message": f"操作失败: {str(e)}"}
