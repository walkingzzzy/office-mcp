"""Excel 工作簿高级操作模块."""

import shutil
from typing import Any, Optional
from pathlib import Path

from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.protection import SheetProtection
from loguru import logger

from office_mcp_server.config import config
from office_mcp_server.utils.file_manager import FileManager


class ExcelWorkbookAdvancedOperations:
    """Excel 工作簿高级操作类."""

    def __init__(self) -> None:
        """初始化高级操作类."""
        self.file_manager = FileManager()

    def create_from_template(
        self,
        template_file: str,
        new_filename: str,
        sheet_name: Optional[str] = None,
    ) -> dict[str, Any]:
        """基于模板创建工作簿.

        Args:
            template_file: 模板文件路径
            new_filename: 新文件名
            sheet_name: 新工作表名称 (可选)
        """
        try:
            template_path = Path(template_file)
            if not template_path.is_absolute():
                template_path = config.paths.output_dir / template_file

            if not template_path.exists():
                raise FileNotFoundError(f"模板文件不存在: {template_path}")

            output_path = config.paths.output_dir / new_filename
            self.file_manager.ensure_directory(output_path.parent)

            shutil.copy(template_path, output_path)

            if sheet_name:
                wb = load_workbook(str(output_path))
                if wb.active:
                    wb.active.title = sheet_name
                wb.save(str(output_path))
                wb.close()

            logger.info(f"基于模板创建工作簿成功: {template_path} -> {output_path}")
            return {
                "success": True,
                "message": f"成功基于模板创建工作簿",
                "template": str(template_path),
                "filename": str(output_path),
                "sheet_name": sheet_name,
            }

        except Exception as e:
            logger.error(f"基于模板创建工作簿失败: {e}")
            return {"success": False, "message": f"创建失败: {str(e)}"}

    def copy_workbook(
        self,
        source_file: str,
        new_filename: str,
    ) -> dict[str, Any]:
        """复制工作簿.

        Args:
            source_file: 源文件路径
            new_filename: 新文件名
        """
        try:
            source_path = config.paths.output_dir / source_file
            self.file_manager.validate_file_path(source_path, must_exist=True)

            output_path = config.paths.output_dir / new_filename
            self.file_manager.ensure_directory(output_path.parent)

            shutil.copy(source_path, output_path)

            logger.info(f"复制工作簿成功: {source_path} -> {output_path}")
            return {
                "success": True,
                "message": f"成功复制工作簿",
                "source": str(source_path),
                "filename": str(output_path),
            }

        except Exception as e:
            logger.error(f"复制工作簿失败: {e}")
            return {"success": False, "message": f"复制失败: {str(e)}"}

    def protect_sheet(
        self,
        filename: str,
        sheet_name: str,
        password: Optional[str] = None,
        enable: bool = True,
    ) -> dict[str, Any]:
        """保护/取消保护工作表.

        Args:
            filename: 文件名
            sheet_name: 工作表名称
            password: 密码 (可选)
            enable: True为保护, False为取消保护
        """
        try:
            file_path = config.paths.output_dir / filename
            self.file_manager.validate_file_path(file_path, must_exist=True)

            wb = load_workbook(str(file_path))

            if sheet_name not in wb.sheetnames:
                raise ValueError(f"工作表 '{sheet_name}' 不存在")

            ws = wb[sheet_name]

            if enable:
                if password:
                    ws.protection.password = password
                ws.protection.sheet = True
                ws.protection.enable()
                message = f"工作表 '{sheet_name}' 已保护"
            else:
                ws.protection.sheet = False
                ws.protection.disable()
                message = f"工作表 '{sheet_name}' 已取消保护"

            wb.save(str(file_path))
            wb.close()

            logger.info(f"工作表保护操作成功: {file_path}")
            return {
                "success": True,
                "message": message,
                "filename": str(file_path),
                "sheet_name": sheet_name,
                "protected": enable,
            }

        except Exception as e:
            logger.error(f"工作表保护操作失败: {e}")
            return {"success": False, "message": f"操作失败: {str(e)}"}

    def freeze_panes(
        self,
        filename: str,
        sheet_name: str,
        cell: Optional[str] = None,
        freeze_rows: int = 0,
        freeze_cols: int = 0,
    ) -> dict[str, Any]:
        """冻结窗格.

        Args:
            filename: 文件名
            sheet_name: 工作表名称
            cell: 冻结位置单元格 (如 'B2' 表示冻结第1行和第A列)
            freeze_rows: 冻结的行数 (从顶部开始)
            freeze_cols: 冻结的列数 (从左侧开始)
        """
        try:
            file_path = config.paths.output_dir / filename
            self.file_manager.validate_file_path(file_path, must_exist=True)

            wb = load_workbook(str(file_path))

            if sheet_name not in wb.sheetnames:
                raise ValueError(f"工作表 '{sheet_name}' 不存在")

            ws = wb[sheet_name]

            # 如果指定了单元格，使用单元格位置
            if cell:
                ws.freeze_panes = cell
                freeze_info = f"单元格 {cell}"
            # 否则根据行列数计算
            elif freeze_rows > 0 or freeze_cols > 0:
                from openpyxl.utils import get_column_letter
                col_letter = get_column_letter(freeze_cols + 1)
                row_num = freeze_rows + 1
                ws.freeze_panes = f"{col_letter}{row_num}"
                freeze_info = f"{freeze_rows} 行, {freeze_cols} 列"
            else:
                # 取消冻结
                ws.freeze_panes = None
                freeze_info = "已取消"

            wb.save(str(file_path))
            wb.close()

            logger.info(f"冻结窗格成功: {file_path}")
            return {
                "success": True,
                "message": f"冻结窗格成功: {freeze_info}",
                "filename": str(file_path),
                "sheet_name": sheet_name,
                "freeze_info": freeze_info,
            }

        except Exception as e:
            logger.error(f"冻结窗格失败: {e}")
            return {"success": False, "message": f"操作失败: {str(e)}"}

    def auto_save_workbook(
        self,
        filename: str,
        backup_dir: Optional[str] = None,
        version_suffix: Optional[str] = None,
    ) -> dict[str, Any]:
        """自动保存工作簿（创建备份）.

        Args:
            filename: 文件名
            backup_dir: 备份目录 (可选，默认在同目录下创建backup文件夹)
            version_suffix: 版本后缀 (如 '_v1', '_backup', 默认使用时间戳)
        """
        try:
            from datetime import datetime

            file_path = config.paths.output_dir / filename
            self.file_manager.validate_file_path(file_path, must_exist=True)

            # 确定备份目录
            if backup_dir:
                backup_path = config.paths.output_dir / backup_dir
            else:
                backup_path = config.paths.output_dir / "backup"

            # 创建备份目录
            backup_path.mkdir(parents=True, exist_ok=True)

            # 生成备份文件名
            if version_suffix:
                backup_filename = f"{file_path.stem}{version_suffix}{file_path.suffix}"
            else:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                backup_filename = f"{file_path.stem}_backup_{timestamp}{file_path.suffix}"

            backup_file_path = backup_path / backup_filename

            # 复制文件作为备份
            shutil.copy(str(file_path), str(backup_file_path))

            logger.info(f"自动保存工作簿成功: {backup_file_path}")
            return {
                "success": True,
                "message": "自动保存成功",
                "original_file": str(file_path),
                "backup_file": str(backup_file_path),
            }

        except Exception as e:
            logger.error(f"自动保存失败: {e}")
            return {"success": False, "message": f"保存失败: {str(e)}"}
