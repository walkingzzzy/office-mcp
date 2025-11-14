"""Excel 处理器主模块 - 门面模式."""

from typing import Any, Optional, Union

from loguru import logger

from office_mcp_server.config import config
from office_mcp_server.handlers.excel.excel_basic import ExcelBasicOperations
from office_mcp_server.handlers.excel.excel_format import ExcelFormatOperations
from office_mcp_server.handlers.excel.excel_data import ExcelDataOperations
from office_mcp_server.handlers.excel.excel_chart import ExcelChartOperations
from office_mcp_server.handlers.excel.excel_formula import ExcelFormulaOperations
from office_mcp_server.handlers.excel.excel_pivot import ExcelPivotOperations
from office_mcp_server.handlers.excel.excel_row_col import ExcelRowColOperations
from office_mcp_server.handlers.excel.excel_merge import ExcelMergeOperations
from office_mcp_server.handlers.excel.excel_autofill import ExcelAutoFillOperations
from office_mcp_server.handlers.excel.excel_import_export import ExcelImportExportOperations
from office_mcp_server.handlers.excel.excel_workbook_advanced import ExcelWorkbookAdvancedOperations
from office_mcp_server.handlers.excel.excel_print import ExcelPrintOperations
from office_mcp_server.handlers.excel.excel_batch import ExcelBatchOperations
from office_mcp_server.handlers.excel.excel_analysis import ExcelAnalysisOperations
from office_mcp_server.handlers.excel.excel_collaboration import ExcelCollaborationOperations
from office_mcp_server.handlers.excel.excel_security import ExcelSecurityOperations
from office_mcp_server.handlers.excel.excel_report import ExcelReportAutomation
from office_mcp_server.handlers.excel.excel_cell_advanced import ExcelCellAdvancedOperations
from office_mcp_server.handlers.excel.excel_data_masking import DataMasking


class ExcelHandler:
    """Excel 处理器类 - 门面模式.

    将所有Excel操作委托给相应的子模块处理。
    """

    def __init__(self) -> None:
        """初始化 Excel 处理器."""
        self.basic_ops = ExcelBasicOperations()
        self.format_ops = ExcelFormatOperations()
        self.data_ops = ExcelDataOperations()
        self.chart_ops = ExcelChartOperations()
        self.formula_ops = ExcelFormulaOperations()
        self.pivot_ops = ExcelPivotOperations()
        self.row_col_ops = ExcelRowColOperations()
        self.merge_ops = ExcelMergeOperations()
        self.autofill_ops = ExcelAutoFillOperations()
        self.import_export_ops = ExcelImportExportOperations()
        self.workbook_advanced_ops = ExcelWorkbookAdvancedOperations()
        self.print_ops = ExcelPrintOperations()
        self.batch_ops = ExcelBatchOperations()
        self.analysis_ops = ExcelAnalysisOperations()
        self.collaboration_ops = ExcelCollaborationOperations()
        self.security_ops = ExcelSecurityOperations()
        self.report_ops = ExcelReportAutomation()
        self.cell_advanced_ops = ExcelCellAdvancedOperations()
        self.data_masking_ops = DataMasking()
        logger.info("Excel 处理器初始化完成 - 已加载所有功能模块")

    # ========== 基础操作 ==========
    def create_workbook(self, filename: str, sheet_name: Optional[str] = None) -> dict[str, Any]:
        """创建工作簿."""
        return self.basic_ops.create_workbook(filename, sheet_name)

    def write_cell(
        self, filename: str, sheet_name: str, cell: str, value: Union[str, int, float]
    ) -> dict[str, Any]:
        """写入单元格."""
        return self.basic_ops.write_cell(filename, sheet_name, cell, value)

    def write_range(
        self, filename: str, sheet_name: str, start_cell: str, data: list[list[Any]]
    ) -> dict[str, Any]:
        """批量写入数据."""
        return self.basic_ops.write_range(filename, sheet_name, start_cell, data)

    def read_cell(self, filename: str, sheet_name: str, cell: str) -> dict[str, Any]:
        """读取单元格."""
        return self.basic_ops.read_cell(filename, sheet_name, cell)

    def get_workbook_info(self, filename: str) -> dict[str, Any]:
        """获取工作簿信息."""
        return self.basic_ops.get_workbook_info(filename)

    # ========== 格式化操作 ==========
    def format_cell(
        self,
        filename: str,
        sheet_name: str,
        cell: str,
        font_name: Optional[str] = None,
        font_size: Optional[int] = None,
        bold: bool = False,
        color: Optional[str] = None,
        bg_color: Optional[str] = None,
        number_format: Optional[str] = None,
        horizontal_alignment: Optional[str] = None,
        vertical_alignment: Optional[str] = None,
        wrap_text: bool = False,
        border_style: Optional[str] = None,
        border_color: Optional[str] = None,
    ) -> dict[str, Any]:
        """格式化单元格."""
        return self.format_ops.format_cell(
            filename, sheet_name, cell, font_name, font_size, bold, color, bg_color,
            number_format, horizontal_alignment, vertical_alignment, wrap_text,
            border_style, border_color
        )

    def apply_conditional_formatting(
        self,
        filename: str,
        sheet_name: str,
        cell_range: str,
        rule_type: str,
        format_type: str = "fill",
        color: Optional[str] = None,
        operator: Optional[str] = None,
        formula: Optional[str] = None,
        value1: Optional[Union[str, int, float]] = None,
        value2: Optional[Union[str, int, float]] = None,
    ) -> dict[str, Any]:
        """应用条件格式."""
        return self.format_ops.apply_conditional_formatting(
            filename, sheet_name, cell_range, rule_type, format_type,
            color, operator, formula, value1, value2
        )

    # ========== 数据操作 ==========
    def sort_data(
        self,
        filename: str,
        sheet_name: str,
        data_range: str,
        sort_by_column: int = 0,
        ascending: bool = True,
        sort_keys: Optional[list[dict[str, Any]]] = None,
    ) -> dict[str, Any]:
        """排序数据."""
        return self.data_ops.sort_data(filename, sheet_name, data_range, sort_by_column, ascending, sort_keys)

    def sort_data_by_color(
        self,
        filename: str,
        sheet_name: str,
        data_range: str,
        sort_by_column: int,
        color_order: list[str],
        sort_by: str = "fill",
    ) -> dict[str, Any]:
        """按颜色排序数据."""
        return self.data_ops.sort_data_by_color(filename, sheet_name, data_range, sort_by_column, color_order, sort_by)

    def copy_filtered_data(
        self,
        filename: str,
        sheet_name: str,
        source_range: str,
        target_cell: str,
        filter_column: int,
        filter_value: str,
        filter_operator: str = "equals",
    ) -> dict[str, Any]:
        """复制筛选结果."""
        return self.data_ops.copy_filtered_data(
            filename, sheet_name, source_range, target_cell, filter_column, filter_value, filter_operator
        )

    def filter_data(
        self,
        filename: str,
        sheet_name: str,
        data_range: str,
        filter_column: Optional[int] = None,
        filter_value: Optional[str] = None,
        filter_operator: str = "equals",
        enable_autofilter: bool = True,
    ) -> dict[str, Any]:
        """筛选数据."""
        return self.data_ops.filter_data(
            filename, sheet_name, data_range, filter_column,
            filter_value, filter_operator, enable_autofilter
        )

    def set_data_validation(
        self,
        filename: str,
        sheet_name: str,
        cell_range: str,
        validation_type: str,
        operator: Optional[str] = None,
        formula1: Optional[str] = None,
        formula2: Optional[str] = None,
        allow_blank: bool = True,
        show_dropdown: bool = True,
        prompt_title: Optional[str] = None,
        prompt: Optional[str] = None,
        error_title: Optional[str] = None,
        error: Optional[str] = None,
    ) -> dict[str, Any]:
        """设置数据验证."""
        return self.data_ops.set_data_validation(
            filename, sheet_name, cell_range, validation_type, operator,
            formula1, formula2, allow_blank, show_dropdown,
            prompt_title, prompt, error_title, error
        )

    def manage_worksheets(
        self,
        filename: str,
        operation: str,
        sheet_name: Optional[str] = None,
        new_name: Optional[str] = None,
        target_index: Optional[int] = None,
    ) -> dict[str, Any]:
        """管理工作表."""
        return self.data_ops.manage_worksheets(
            filename, operation, sheet_name, new_name, target_index
        )

    def create_table(
        self,
        filename: str,
        sheet_name: str,
        table_range: str,
        table_name: str,
        style: str = "TableStyleMedium9",
        show_header: bool = True,
        show_totals: bool = False,
    ) -> dict[str, Any]:
        """创建Excel表格样式."""
        return self.data_ops.create_table(
            filename, sheet_name, table_range, table_name,
            style, show_header, show_totals
        )

    # ========== 图表操作 ==========
    def create_chart(
        self,
        filename: str,
        sheet_name: str,
        chart_type: str,
        data_range: str,
        title: str = "",
        position: str = "E5",
        x_axis_title: Optional[str] = None,
        y_axis_title: Optional[str] = None,
        legend_position: Optional[str] = None,
        show_data_labels: bool = False,
        grouping: Optional[str] = None,
    ) -> dict[str, Any]:
        """创建图表."""
        return self.chart_ops.create_chart(
            filename, sheet_name, chart_type, data_range, title, position,
            x_axis_title, y_axis_title, legend_position, show_data_labels, grouping
        )

    # ========== 公式操作 ==========
    def insert_formula(
        self,
        filename: str,
        sheet_name: str,
        cell: str,
        formula: str,
    ) -> dict[str, Any]:
        """插入公式."""
        return self.formula_ops.insert_formula(filename, sheet_name, cell, formula)

    def apply_function(
        self,
        filename: str,
        sheet_name: str,
        cell: str,
        function_name: str,
        range1: Optional[str] = None,
        range2: Optional[str] = None,
        condition: Optional[str] = None,
        value_if_true: Optional[str] = None,
        value_if_false: Optional[str] = None,
        lookup_value: Optional[str] = None,
        table_array: Optional[str] = None,
        col_index: Optional[int] = None,
        range_lookup: bool = False,
    ) -> dict[str, Any]:
        """应用常用函数."""
        return self.formula_ops.apply_function(
            filename, sheet_name, cell, function_name, range1, range2,
            condition, value_if_true, value_if_false, lookup_value,
            table_array, col_index, range_lookup
        )

    # ========== 数据透视表操作 ==========
    def create_pivot_table(
        self,
        filename: str,
        source_sheet: str,
        source_range: str,
        pivot_sheet: str,
        pivot_location: str,
        row_fields: list[str],
        col_fields: Optional[list[str]] = None,
        data_fields: Optional[list[dict[str, str]]] = None,
        filter_fields: Optional[list[str]] = None,
    ) -> dict[str, Any]:
        """创建数据透视表."""
        return self.pivot_ops.create_pivot_table(
            filename, source_sheet, source_range, pivot_sheet, pivot_location,
            row_fields, col_fields, data_fields, filter_fields
        )

    def refresh_pivot_table(
        self,
        filename: str,
        pivot_sheet: str,
        pivot_table_name: str = "PivotTable1",
    ) -> dict[str, Any]:
        """刷新数据透视表."""
        return self.pivot_ops.refresh_pivot_table(filename, pivot_sheet, pivot_table_name)

    def refresh_all_pivot_tables(
        self,
        filename: str,
        sheet_name: Optional[str] = None,
    ) -> dict[str, Any]:
        """刷新所有数据透视表."""
        return self.pivot_ops.refresh_all_pivot_tables(filename, sheet_name)

    # ========== 行列操作 ==========
    def insert_rows(self, filename: str, sheet_name: str, row_index: int, count: int = 1) -> dict[str, Any]:
        """插入行."""
        return self.row_col_ops.insert_rows(filename, sheet_name, row_index, count)

    def delete_rows(self, filename: str, sheet_name: str, row_index: int, count: int = 1) -> dict[str, Any]:
        """删除行."""
        return self.row_col_ops.delete_rows(filename, sheet_name, row_index, count)

    def insert_cols(self, filename: str, sheet_name: str, col_index: int, count: int = 1) -> dict[str, Any]:
        """插入列."""
        return self.row_col_ops.insert_cols(filename, sheet_name, col_index, count)

    def delete_cols(self, filename: str, sheet_name: str, col_index: int, count: int = 1) -> dict[str, Any]:
        """删除列."""
        return self.row_col_ops.delete_cols(filename, sheet_name, col_index, count)

    def hide_rows(self, filename: str, sheet_name: str, row_start: int, row_end: Optional[int] = None) -> dict[str, Any]:
        """隐藏行."""
        return self.row_col_ops.hide_rows(filename, sheet_name, row_start, row_end)

    def show_rows(self, filename: str, sheet_name: str, row_start: int, row_end: Optional[int] = None) -> dict[str, Any]:
        """显示行."""
        return self.row_col_ops.show_rows(filename, sheet_name, row_start, row_end)

    def hide_cols(self, filename: str, sheet_name: str, col_start: int, col_end: Optional[int] = None) -> dict[str, Any]:
        """隐藏列."""
        return self.row_col_ops.hide_cols(filename, sheet_name, col_start, col_end)

    def show_cols(self, filename: str, sheet_name: str, col_start: int, col_end: Optional[int] = None) -> dict[str, Any]:
        """显示列."""
        return self.row_col_ops.show_cols(filename, sheet_name, col_start, col_end)

    def set_row_height(self, filename: str, sheet_name: str, row_index: int, height: float) -> dict[str, Any]:
        """设置行高."""
        return self.row_col_ops.set_row_height(filename, sheet_name, row_index, height)

    def set_col_width(self, filename: str, sheet_name: str, col_index: int, width: float) -> dict[str, Any]:
        """设置列宽."""
        return self.row_col_ops.set_col_width(filename, sheet_name, col_index, width)

    def copy_rows(self, filename: str, sheet_name: str, source_row: int, target_row: int, count: int = 1) -> dict[str, Any]:
        """复制行."""
        return self.row_col_ops.copy_rows(filename, sheet_name, source_row, target_row, count)

    def copy_cols(self, filename: str, sheet_name: str, source_col: int, target_col: int, count: int = 1) -> dict[str, Any]:
        """复制列."""
        return self.row_col_ops.copy_cols(filename, sheet_name, source_col, target_col, count)

    def move_rows(self, filename: str, sheet_name: str, source_row: int, target_row: int, count: int = 1) -> dict[str, Any]:
        """移动行."""
        return self.row_col_ops.move_rows(filename, sheet_name, source_row, target_row, count)

    def move_cols(self, filename: str, sheet_name: str, source_col: int, target_col: int, count: int = 1) -> dict[str, Any]:
        """移动列."""
        return self.row_col_ops.move_cols(filename, sheet_name, source_col, target_col, count)

    # ========== 单元格合并操作 ==========
    def merge_cells(self, filename: str, sheet_name: str, cell_range: str) -> dict[str, Any]:
        """合并单元格."""
        return self.merge_ops.merge_cells(filename, sheet_name, cell_range)

    def unmerge_cells(self, filename: str, sheet_name: str, cell_range: str) -> dict[str, Any]:
        """取消合并单元格."""
        return self.merge_ops.unmerge_cells(filename, sheet_name, cell_range)

    # ========== 数据读取增强 ==========
    def read_range(self, filename: str, sheet_name: str, cell_range: str) -> dict[str, Any]:
        """读取范围."""
        return self.basic_ops.read_range(filename, sheet_name, cell_range)

    def read_row(self, filename: str, sheet_name: str, row_index: int) -> dict[str, Any]:
        """读取整行."""
        return self.basic_ops.read_row(filename, sheet_name, row_index)

    def read_column(self, filename: str, sheet_name: str, col_index: int) -> dict[str, Any]:
        """读取整列."""
        return self.basic_ops.read_column(filename, sheet_name, col_index)

    def read_all_data(self, filename: str, sheet_name: str, include_empty: bool = False) -> dict[str, Any]:
        """读取整表数据."""
        return self.basic_ops.read_all_data(filename, sheet_name, include_empty)

    def clear_cell(self, filename: str, sheet_name: str, cell: str) -> dict[str, Any]:
        """清除单元格."""
        return self.basic_ops.clear_cell(filename, sheet_name, cell)

    def clear_range(self, filename: str, sheet_name: str, cell_range: str) -> dict[str, Any]:
        """清除范围."""
        return self.basic_ops.clear_range(filename, sheet_name, cell_range)

    # ========== 自动填充操作 ==========
    def fill_series(self, filename: str, sheet_name: str, start_cell: str, end_cell: str,
                    fill_type: str = "linear", start_value: Union[int, float] = 1,
                    step: Union[int, float] = 1) -> dict[str, Any]:
        """序列填充."""
        return self.autofill_ops.fill_series(filename, sheet_name, start_cell, end_cell,
                                             fill_type, start_value, step)

    def copy_fill(self, filename: str, sheet_name: str, source_cell: str, target_range: str) -> dict[str, Any]:
        """复制填充."""
        return self.autofill_ops.copy_fill(filename, sheet_name, source_cell, target_range)

    def formula_fill(self, filename: str, sheet_name: str, start_cell: str, formula: str,
                     fill_direction: str = "down", count: int = 10) -> dict[str, Any]:
        """公式填充."""
        return self.autofill_ops.formula_fill(filename, sheet_name, start_cell, formula,
                                              fill_direction, count)

    # ========== 导入导出操作 ==========
    def import_from_csv(self, filename: str, sheet_name: str, csv_file: str,
                        start_cell: str = "A1", has_header: bool = True) -> dict[str, Any]:
        """从CSV导入."""
        return self.import_export_ops.import_from_csv(filename, sheet_name, csv_file,
                                                       start_cell, has_header)

    def import_from_json(self, filename: str, sheet_name: str, json_file: str,
                         start_cell: str = "A1", json_path: Optional[str] = None) -> dict[str, Any]:
        """从JSON导入."""
        return self.import_export_ops.import_from_json(filename, sheet_name, json_file,
                                                        start_cell, json_path)

    def export_to_csv(self, filename: str, sheet_name: str, csv_file: str,
                      cell_range: Optional[str] = None) -> dict[str, Any]:
        """导出为CSV."""
        return self.import_export_ops.export_to_csv(filename, sheet_name, csv_file, cell_range)

    def export_to_json(self, filename: str, sheet_name: str, json_file: str,
                       cell_range: Optional[str] = None, has_header: bool = True,
                       orient: str = "records") -> dict[str, Any]:
        """导出为JSON."""
        return self.import_export_ops.export_to_json(filename, sheet_name, json_file,
                                                       cell_range, has_header, orient)

    def export_to_pdf(self, filename: str, sheet_name: str, pdf_file: str,
                      cell_range: Optional[str] = None) -> dict[str, Any]:
        """导出为PDF."""
        return self.import_export_ops.export_to_pdf(filename, sheet_name, pdf_file, cell_range)

    def export_to_html(self, filename: str, sheet_name: str, html_file: str,
                       cell_range: Optional[str] = None, include_style: bool = True) -> dict[str, Any]:
        """导出为HTML."""
        return self.import_export_ops.export_to_html(filename, sheet_name, html_file,
                                                       cell_range, include_style)

    # ========== 工作簿高级操作 ==========
    def create_from_template(self, template_file: str, new_filename: str,
                            sheet_name: Optional[str] = None) -> dict[str, Any]:
        """基于模板创建."""
        return self.workbook_advanced_ops.create_from_template(template_file, new_filename, sheet_name)

    def copy_workbook(self, source_file: str, new_filename: str) -> dict[str, Any]:
        """复制工作簿."""
        return self.workbook_advanced_ops.copy_workbook(source_file, new_filename)

    def protect_sheet(self, filename: str, sheet_name: str, password: Optional[str] = None,
                      enable: bool = True) -> dict[str, Any]:
        """保护工作表."""
        return self.workbook_advanced_ops.protect_sheet(filename, sheet_name, password, enable)

    def freeze_panes(self, filename: str, sheet_name: str, cell: Optional[str] = None,
                     freeze_rows: int = 0, freeze_cols: int = 0) -> dict[str, Any]:
        """冻结窗格."""
        return self.workbook_advanced_ops.freeze_panes(filename, sheet_name, cell, freeze_rows, freeze_cols)

    def auto_save_workbook(self, filename: str, backup_dir: Optional[str] = None,
                           version_suffix: Optional[str] = None) -> dict[str, Any]:
        """自动保存工作簿."""
        return self.workbook_advanced_ops.auto_save_workbook(filename, backup_dir, version_suffix)

    # ========== 图表高级操作 ==========
    def format_chart(self, filename: str, sheet_name: str, chart_index: int = 0,
                     title_font_size: Optional[int] = None, title_font_bold: bool = False,
                     chart_style: Optional[int] = None, color_scheme: Optional[list[str]] = None) -> dict[str, Any]:
        """格式化图表."""
        return self.chart_ops.format_chart(filename, sheet_name, chart_index,
                                           title_font_size, title_font_bold, chart_style, color_scheme)

    def create_combination_chart(self, filename: str, sheet_name: str, data_range1: str, data_range2: str,
                                 chart_type1: str = "bar", chart_type2: str = "line",
                                 title: str = "", position: str = "E5") -> dict[str, Any]:
        """创建组合图表."""
        return self.chart_ops.create_combination_chart(filename, sheet_name, data_range1, data_range2,
                                                        chart_type1, chart_type2, title, position)

    def add_trendline_to_chart(self, filename: str, sheet_name: str, chart_index: int = 0,
                               series_index: int = 0, trendline_type: str = "linear",
                               display_equation: bool = False, display_r_squared: bool = False) -> dict[str, Any]:
        """为图表添加趋势线."""
        return self.chart_ops.add_trendline_to_chart(filename, sheet_name, chart_index, series_index,
                                                       trendline_type, display_equation, display_r_squared)

    # ========== 打印设置操作 ==========
    def set_page_setup(self, filename: str, sheet_name: str, orientation: str = "portrait",
                       paper_size: int = 9, scale: int = 100, fit_to_width: Optional[int] = None,
                       fit_to_height: Optional[int] = None) -> dict[str, Any]:
        """设置页面."""
        return self.print_ops.set_page_setup(filename, sheet_name, orientation, paper_size,
                                              scale, fit_to_width, fit_to_height)

    def set_page_margins(self, filename: str, sheet_name: str, left: float = 0.75,
                        right: float = 0.75, top: float = 1.0, bottom: float = 1.0,
                        header: float = 0.5, footer: float = 0.5) -> dict[str, Any]:
        """设置页边距."""
        return self.print_ops.set_page_margins(filename, sheet_name, left, right, top,
                                                bottom, header, footer)

    def set_print_area(self, filename: str, sheet_name: str, print_area: Optional[str] = None) -> dict[str, Any]:
        """设置打印区域."""
        return self.print_ops.set_print_area(filename, sheet_name, print_area)

    def set_print_titles(self, filename: str, sheet_name: str, rows: Optional[str] = None,
                         cols: Optional[str] = None) -> dict[str, Any]:
        """设置打印标题."""
        return self.print_ops.set_print_titles(filename, sheet_name, rows, cols)

    def insert_page_break(self, filename: str, sheet_name: str, cell: str, break_type: str = "row") -> dict[str, Any]:
        """插入分页符."""
        return self.print_ops.insert_page_break(filename, sheet_name, cell, break_type)

    def delete_page_break(self, filename: str, sheet_name: str, break_type: str, position: int) -> dict[str, Any]:
        """删除分页符."""
        return self.print_ops.delete_page_break(filename, sheet_name, break_type, position)

    def clear_all_page_breaks(self, filename: str, sheet_name: str) -> dict[str, Any]:
        """清除所有分页符."""
        return self.print_ops.clear_all_page_breaks(filename, sheet_name)

    # ========== 批量处理操作 ==========
    def batch_process_files(self, pattern: str, operation: str, **kwargs: Any) -> dict[str, Any]:
        """批量处理."""
        return self.batch_ops.batch_process_files(pattern, operation, **kwargs)

    def merge_workbooks(self, source_files: list[str], output_file: str, merge_mode: str = "sheets") -> dict[str, Any]:
        """合并工作簿."""
        return self.batch_ops.merge_workbooks(source_files, output_file, merge_mode)

    # ========== 数据分析操作 ==========
    def descriptive_statistics(self, filename: str, sheet_name: str, data_range: str, output_cell: Optional[str] = None) -> dict[str, Any]:
        """描述性统计."""
        # 注意：当前实现不使用 output_cell 参数，但保留以保持接口一致性
        return self.analysis_ops.descriptive_statistics(filename, sheet_name, data_range)

    def correlation_analysis(self, filename: str, sheet_name: str, data_range1: str, data_range2: str) -> dict[str, Any]:
        """相关性分析."""
        return self.analysis_ops.correlation_analysis(filename, sheet_name, data_range1, data_range2)

    def goal_seek(self, filename: str, sheet_name: str, target_cell: str, target_value: float,
                  variable_cell: str, max_iterations: int = 100, tolerance: float = 0.001) -> dict[str, Any]:
        """单变量求解."""
        return self.analysis_ops.goal_seek(filename, sheet_name, target_cell, target_value,
                                            variable_cell, max_iterations, tolerance)

    # ========== 协作功能操作 ==========
    def add_comment(self, filename: str, sheet_name: str, cell: str, comment_text: str,
                    author: str = "User") -> dict[str, Any]:
        """添加批注."""
        return self.collaboration_ops.add_comment(filename, sheet_name, cell, comment_text, author)

    def get_comment(self, filename: str, sheet_name: str, cell: str) -> dict[str, Any]:
        """获取批注."""
        return self.collaboration_ops.get_comment(filename, sheet_name, cell)

    def delete_comment(self, filename: str, sheet_name: str, cell: str) -> dict[str, Any]:
        """删除批注."""
        return self.collaboration_ops.delete_comment(filename, sheet_name, cell)

    def list_all_comments(self, filename: str, sheet_name: Optional[str] = None) -> dict[str, Any]:
        """列出所有批注."""
        # 如果没有指定工作表，使用第一个工作表
        if sheet_name is None:
            from openpyxl import load_workbook
            file_path = config.paths.output_dir / filename
            wb = load_workbook(str(file_path))
            sheet_name = wb.sheetnames[0]
            wb.close()
        return self.collaboration_ops.list_all_comments(filename, sheet_name)

    # ========== 安全功能操作 ==========
    def encrypt_workbook(self, filename: str, password: str) -> dict[str, Any]:
        """加密工作簿."""
        return self.security_ops.encrypt_workbook(filename, password)

    def lock_cells(self, filename: str, sheet_name: str, cell_range: str, locked: bool = True) -> dict[str, Any]:
        """锁定单元格."""
        return self.security_ops.lock_cells(filename, sheet_name, cell_range, locked)

    def hide_formulas(self, filename: str, sheet_name: str, cell_range: str, hidden: bool = True) -> dict[str, Any]:
        """隐藏公式."""
        return self.security_ops.hide_formulas(filename, sheet_name, cell_range, hidden)

    # ========== 报表自动化操作 ==========
    def generate_report_from_template(self, template_file: str, output_file: str, data: dict[str, Any],
                                      mappings: Optional[dict[str, str]] = None) -> dict[str, Any]:
        """生成报表."""
        return self.report_ops.generate_report_from_template(template_file, output_file, data, mappings)

    def update_report_data(self, filename: str, sheet_name: str, updates: dict[str, Any]) -> dict[str, Any]:
        """更新报表."""
        return self.report_ops.update_report_data(filename, sheet_name, updates)

    def consolidate_reports(self, source_files: list[str], output_file: str, sheet_name: str = "Consolidated",
                           include_source_name: bool = True) -> dict[str, Any]:
        """合并报表."""
        return self.report_ops.consolidate_reports(source_files, output_file, sheet_name, include_source_name)

    def schedule_report_generation(self, template_file: str, output_pattern: str, data_source: str,
                                    frequency: str = "daily") -> dict[str, Any]:
        """计划报表生成."""
        return self.report_ops.schedule_report_generation(template_file, output_pattern, data_source, frequency)

    # ========== 高级统计分析操作 ==========
    def regression_analysis(self, filename: str, sheet_name: str, x_range: str, y_range: str,
                            regression_type: str = "linear") -> dict[str, Any]:
        """回归分析."""
        return self.analysis_ops.regression_analysis(filename, sheet_name, x_range, y_range, regression_type)

    def anova_analysis(self, filename: str, sheet_name: str, *group_ranges: str) -> dict[str, Any]:
        """方差分析."""
        return self.analysis_ops.anova_analysis(filename, sheet_name, *group_ranges)

    def t_test(self, filename: str, sheet_name: str, group1_range: str, group2_range: str,
               test_type: str = "independent") -> dict[str, Any]:
        """t检验."""
        return self.analysis_ops.t_test(filename, sheet_name, group1_range, group2_range, test_type)

    def chi_square_test(self, filename: str, sheet_name: str, observed_range: str) -> dict[str, Any]:
        """卡方检验."""
        return self.analysis_ops.chi_square_test(filename, sheet_name, observed_range)

    def trend_analysis(self, filename: str, sheet_name: str, data_range: str,
                       periods_ahead: int = 5, output_cell: Optional[str] = None) -> dict[str, Any]:
        """趋势分析."""
        # 注意：当前实现不使用 output_cell 参数，但保留以保持接口一致性
        return self.analysis_ops.trend_analysis(filename, sheet_name, data_range, periods_ahead)

    def moving_average(self, filename: str, sheet_name: str, data_range: str, window: int = 3,
                       output_cell: Optional[str] = None) -> dict[str, Any]:
        """移动平均."""
        return self.analysis_ops.moving_average(filename, sheet_name, data_range, window, output_cell)

    def exponential_smoothing(self, filename: str, sheet_name: str, data_range: str, alpha: float = 0.3,
                              output_cell: Optional[str] = None) -> dict[str, Any]:
        """指数平滑."""
        return self.analysis_ops.exponential_smoothing(filename, sheet_name, data_range, alpha, output_cell)

    # ========== 单元格高级操作 ==========
    def insert_cells(self, filename: str, sheet_name: str, cell: str, shift: str = "down") -> dict[str, Any]:
        """插入单元格并移动."""
        return self.cell_advanced_ops.insert_cells(filename, sheet_name, cell, shift)

    def delete_cells(self, filename: str, sheet_name: str, cell: str, shift: str = "up") -> dict[str, Any]:
        """删除单元格并移动."""
        return self.cell_advanced_ops.delete_cells(filename, sheet_name, cell, shift)

    def insert_cell_range(self, filename: str, sheet_name: str, start_cell: str, end_cell: str,
                          shift: str = "down") -> dict[str, Any]:
        """插入单元格范围并移动."""
        return self.cell_advanced_ops.insert_cell_range(filename, sheet_name, start_cell, end_cell, shift)

    def delete_cell_range(self, filename: str, sheet_name: str, start_cell: str, end_cell: str,
                          shift: str = "up") -> dict[str, Any]:
        """删除单元格范围并移动."""
        return self.cell_advanced_ops.delete_cell_range(filename, sheet_name, start_cell, end_cell, shift)

    # ========== 数据透视表增强 ==========
    def change_pivot_data_source(self, filename: str, pivot_sheet: str, pivot_table_name: str,
                                  new_source_range: str) -> dict[str, Any]:
        """更改数据透视表数据源."""
        return self.pivot_ops.change_pivot_data_source(filename, pivot_sheet, pivot_table_name, new_source_range)

    # ========== 数据脱敏操作 ==========
    def mask_data(self, filename: str, sheet_name: str, cell_range: str, mask_type: str,
                  mask_char: str = "*", keep_first: int = 0, keep_last: int = 0,
                  custom_pattern: Optional[str] = None) -> dict[str, Any]:
        """数据脱敏."""
        return self.data_masking_ops.mask_data(filename, sheet_name, cell_range, mask_type,
                                                mask_char, keep_first, keep_last, custom_pattern)

    def detect_sensitive_data(self, filename: str, sheet_name: str,
                              cell_range: Optional[str] = None) -> dict[str, Any]:
        """检测敏感数据."""
        return self.data_masking_ops.detect_sensitive_data(filename, sheet_name, cell_range)

    def hash_data(self, filename: str, sheet_name: str, cell_range: str,
                  algorithm: str = "sha256") -> dict[str, Any]:
        """哈希加密数据."""
        return self.data_masking_ops.hash_data(filename, sheet_name, cell_range, algorithm)

